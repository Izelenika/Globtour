import re
from flask import send_file
from functools import wraps
import io
from flask import Flask,render_template,request,redirect,url_for,flash,send_file,jsonify,send_from_directory,abort,session
import sqlite3,os,io,uuid,hashlib
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment
from openpyxl.utils import get_column_letter
BASE=os.path.dirname(os.path.abspath(__file__)); DB=os.path.join(BASE,"raspored.db")
app=Flask(__name__)
def hours_hm(value):
 try:
  total_minutes=int(round(float(value or 0)*60))
 except (TypeError,ValueError):
  total_minutes=0
 h=total_minutes//60
 m=total_minutes%60
 return f"{h} h {m} min"
app.jinja_env.filters["hours_hm"]=hours_hm
@app.context_processor
def inject_today():
 return {'today':date.today().isoformat()}
app.secret_key="globtour-local"

# Authentication uses a browser-session cookie. Closing the browser removes it.
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

UPLOAD_DIR=os.path.join(BASE,"uploads","permits")
os.makedirs(UPLOAD_DIR,exist_ok=True)
ALLOWED_PERMIT_EXTENSIONS={"pdf","jpg","jpeg","png","webp"}
def permit_file_allowed(name):
    return "." in name and name.rsplit(".",1)[1].lower() in ALLOWED_PERMIT_EXTENSIONS


@app.template_filter("datefmt")
def datefmt_filter(value):
    try:
        return date.fromisoformat(str(value)[:10]).strftime("%d.%m.%Y.")
    except Exception:
        return str(value or "")

@app.template_filter("blank")
def blank_filter(value):
    return "" if value is None or str(value).strip().lower() == "none" else value

try:
    _c=sqlite3.connect(DB)
    _cols=[r[1] for r in _c.execute("PRAGMA table_info(lines)").fetchall()]
    if "group_type" not in _cols:
        _c.execute("ALTER TABLE lines ADD COLUMN group_type TEXT DEFAULT 'D0'")
    _c.execute("""CREATE TABLE IF NOT EXISTS line_permits (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        line_id INTEGER NOT NULL,
        permit TEXT NOT NULL,
        valid_until TEXT,
        file_name TEXT
    )""")
    _pcols=[r[1] for r in _c.execute("PRAGMA table_info(line_permits)").fetchall()]
    if "valid_until" not in _pcols:
        _c.execute("ALTER TABLE line_permits ADD COLUMN valid_until TEXT")
    if "file_name" not in _pcols:
        _c.execute("ALTER TABLE line_permits ADD COLUMN file_name TEXT")
    _vcols=[r[1] for r in _c.execute("PRAGMA table_info(vehicles)").fetchall()]
    for _col,_typ in [
        ("vehicle_type","TEXT"),
        ("seats","INTEGER"),
        ("production_year","INTEGER"),
        ("registration_date","TEXT"),
        ("registration_expiry","TEXT"),
        ("chassis_number","TEXT"), ("axles","INTEGER"), ("tachograph_type","TEXT"),
        ("tachograph_expiry","TEXT"), ("periodic_expiry","TEXT"), ("fire_extinguisher_expiry","TEXT"),
        ("permit_file_name","TEXT"),
    ]:
        if _col not in _vcols:
            _c.execute(f"ALTER TABLE vehicles ADD COLUMN {_col} {_typ}")
    _dcols=[r[1] for r in _c.execute("PRAGMA table_info(drivers)").fetchall()]
    for _col,_typ in [
        ("prijava","TEXT"),
        ("baza","TEXT"),
        ("grad","TEXT"),
        ("adresa","TEXT"),
    ]:
        if _col not in _dcols:
            _c.execute(f"ALTER TABLE drivers ADD COLUMN {_col} {_typ}")
    # Internal return lines and duration used only for statistics.
    _lcols=[r[1] for r in _c.execute("PRAGMA table_info(lines)").fetchall()]
    for _col,_typ in [("internal_return","INTEGER NOT NULL DEFAULT 0"),("duration_days","INTEGER NOT NULL DEFAULT 1"),("distance_km","REAL NOT NULL DEFAULT 0")]:
        if _col not in _lcols: _c.execute(f"ALTER TABLE lines ADD COLUMN {_col} {_typ}")
    _scols=[r[1] for r in _c.execute("PRAGMA table_info(schedules)").fetchall()]
    for _col,_typ in [("is_return","INTEGER NOT NULL DEFAULT 0"),("return_of","INTEGER"),("hidden_from_schedule","INTEGER NOT NULL DEFAULT 0")]:
        if _col not in _scols: _c.execute(f"ALTER TABLE schedules ADD COLUMN {_col} {_typ}")
    _c.commit()
    _c.close()
except Exception:
    pass
# Korisnici i prava pristupa
def hash_password(password):
    return hashlib.sha256((password or "").encode("utf-8")).hexdigest()

PERMISSION_LABELS={
 "dashboard":"Početna",
 "schedule_view":"Raspored – pregled","schedule_edit":"Raspored – unos/uređivanje","schedule_delete":"Raspored – brisanje","schedule_export":"Raspored – Excel",
 "drivers_view":"Vozači – pregled","drivers_edit":"Vozači – unos/uređivanje","drivers_delete":"Vozači – brisanje","drivers_export":"Vozači – Excel",
 "vehicles_view":"Vozila – pregled","vehicles_edit":"Vozila – unos/uređivanje","vehicles_delete":"Vozila – brisanje","vehicles_export":"Vozila – Excel",
 "lines_view":"Linije – pregled","lines_edit":"Linije – unos/uređivanje","lines_delete":"Linije – brisanje","lines_export":"Linije – Excel",
 "tires_view":"Gume – pregled","tires_edit":"Gume – unos/uređivanje","tires_delete":"Gume – brisanje","tires_export":"Gume – Excel",
 "free_rides_view":"Slobodne vožnje – pregled","free_rides_edit":"Slobodne vožnje – unos/uređivanje","free_rides_delete":"Slobodne vožnje – brisanje","users_view":"Korisnici – pregled","users_edit":"Korisnici – unos/uređivanje","users_delete":"Korisnici – brisanje"
}

def clean_row(row):
    if row is None:
        return None
    return {k: ("" if v is None or (isinstance(v,str) and v.strip().lower()=="none") else v)
            for k,v in dict(row).items()}

def clean_rows(rows):
    return [clean_row(r) for r in rows]

def ensure_return_schema():
    c=sqlite3.connect(DB)
    try:
        lc=[r[1] for r in c.execute("PRAGMA table_info(lines)").fetchall()]
        if "internal_return" not in lc: c.execute("ALTER TABLE lines ADD COLUMN internal_return INTEGER NOT NULL DEFAULT 0")
        if "duration_days" not in lc: c.execute("ALTER TABLE lines ADD COLUMN duration_days INTEGER NOT NULL DEFAULT 1")
        if "duration_hours" not in lc: c.execute("ALTER TABLE lines ADD COLUMN duration_hours REAL NOT NULL DEFAULT 0")
        if "duration_hours_int" not in lc: c.execute("ALTER TABLE lines ADD COLUMN duration_hours_int INTEGER NOT NULL DEFAULT 0")
        if "duration_minutes" not in lc: c.execute("ALTER TABLE lines ADD COLUMN duration_minutes INTEGER NOT NULL DEFAULT 0")
        if "distance_km" not in lc: c.execute("ALTER TABLE lines ADD COLUMN distance_km REAL NOT NULL DEFAULT 0")
        sc=[r[1] for r in c.execute("PRAGMA table_info(schedules)").fetchall()]
        if "is_return" not in sc: c.execute("ALTER TABLE schedules ADD COLUMN is_return INTEGER NOT NULL DEFAULT 0")
        if "return_of" not in sc: c.execute("ALTER TABLE schedules ADD COLUMN return_of INTEGER")
        if "hidden_from_schedule" not in sc: c.execute("ALTER TABLE schedules ADD COLUMN hidden_from_schedule INTEGER NOT NULL DEFAULT 0")
        # Stari status "Demontirana - provjera" više se ne koristi:
        # svaka tako označena guma smatra se rashodovanom.
        c.execute("UPDATE tires SET status='Rashod', warehouse=NULL, vehicle=NULL, position=NULL WHERE status='Demontirana - provjera'")
        c.execute("""CREATE TABLE IF NOT EXISTS free_rides (
            id INTEGER PRIMARY KEY AUTOINCREMENT, kind TEXT NOT NULL DEFAULT 'reserved',
            client TEXT, relation TEXT, date_from TEXT, date_to TEXT, passengers TEXT,
            vehicle TEXT, driver1 TEXT, driver2 TEXT, status TEXT DEFAULT 'Rezervirano',
            payment_status TEXT, document_no TEXT, amount REAL DEFAULT 0, notes TEXT,
            issued_by TEXT, created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )""")
        existing={r[1] for r in c.execute("PRAGMA table_info(free_rides)").fetchall()}
        if "issued_by" not in existing:
            c.execute("ALTER TABLE free_rides ADD COLUMN issued_by TEXT")
        c.commit()
    finally: c.close()

def seed_known_line_durations(c):
    # Trajanja preuzeta iz Dnevnice.xlsx / List1:
    # stanice + noćna vožnja + dnevna vožnja + neproduktivne minute.
    known = {
        "Sarajevo-Kopar": (42,35),
        "Međugorje-Novi Sad": (43,10),
        "Dubrovnik-Vinkovci": (44,25),
        "Trebinje-Banja Luka": (38,15),
        "Sarajevo-Split-Mostar": (42,35),
        "Visoko-Dubrovnik": (15,45),
        "Čapljina-Frankfurt": (70,55),
        "Vrgoračka": (27,50),
        "Sarajevo-Split-1500": (25,25),
        "Sarajevo-Foča": (10,15),
        "Bihać-Zagreb": (17,20),
        "Trebinje-Gacko": (12,15),
        "Međugorje-Zagreb-Varaždinska": (49,55),
    }
    def norm(x):
        import re
        x=(x or "").lower().replace("č","c").replace("ć","c").replace("ž","z").replace("š","s").replace("đ","d")
        return re.sub(r"[^a-z0-9]+","-",x).strip("-")
    rows=c.execute("SELECT id,name,duration_hours_int,duration_minutes FROM lines").fetchall()
    for r in rows:
        key=norm(r["name"])
        match=None
        if key in known:
            match=known[key]
        elif key=="sarajevo-split":
            # U postojećoj bazi ova linija nema oznaku 15:00; koristi trajanje Sarajevo–Split (15:00).
            match=(25,25)
        elif key=="medugorje-zagreb":
            # U postojećoj bazi odgovara Varaždinskoj liniji iz List1.
            match=(49,55)
        if match and int(r["duration_hours_int"] or 0)==0 and int(r["duration_minutes"] or 0)==0:
            h,m=match
            c.execute("UPDATE lines SET duration_hours_int=?, duration_minutes=?, duration_hours=? WHERE id=?",
                      (h,m,h+m/60,r["id"]))
    c.commit()

def db():
 ensure_return_schema()
 c=sqlite3.connect(DB); c.row_factory=sqlite3.Row
 seed_known_line_durations(c)
 return c

def ensure_proformas_table():
    c=db()
    try: c.execute("ALTER TABLE proformas ADD COLUMN issued_by TEXT")
    except Exception: pass
    c.execute("""CREATE TABLE IF NOT EXISTS proformas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        number TEXT,
        free_ride_id INTEGER,
        client TEXT,
        relation TEXT,
        issue_date TEXT,
        due_date TEXT,
        amount REAL DEFAULT 0,
        vat_text TEXT,
        currency TEXT DEFAULT 'BAM',
        client_address TEXT,
        client_city TEXT,
        client_id TEXT,
        km_total REAL DEFAULT 0,
        km_bih REAL DEFAULT 0,
        km_hr REAL DEFAULT 0,
        km_ino REAL DEFAULT 0,
        price_per_km REAL DEFAULT 0,
        note TEXT,
        status TEXT DEFAULT 'Otvoren',
        issued_by TEXT
    )""")
    existing={row["name"] for row in c.execute("PRAGMA table_info(proformas)").fetchall()}
    if "issued_by" not in existing:
        c.execute("ALTER TABLE proformas ADD COLUMN issued_by TEXT")
        existing.add("issued_by")
    required={
        "number":"TEXT","free_ride_id":"INTEGER","client":"TEXT","relation":"TEXT",
        "issue_date":"TEXT","due_date":"TEXT","amount":"REAL DEFAULT 0","vat_text":"TEXT",
        "currency":"TEXT DEFAULT 'BAM'","client_address":"TEXT","client_city":"TEXT",
        "client_id":"TEXT","km_total":"REAL DEFAULT 0","km_bih":"REAL DEFAULT 0",
        "km_hr":"REAL DEFAULT 0","km_ino":"REAL DEFAULT 0","price_per_km":"REAL DEFAULT 0",
        "note":"TEXT","status":"TEXT DEFAULT 'Otvoren'"
    }
    for col, typ in required.items():
        if col not in existing:
            c.execute(f"ALTER TABLE proformas ADD COLUMN {col} {typ}")
    # v148: payment method and C-invoice support for free rides
    for _sql in [
        "ALTER TABLE free_rides ADD COLUMN payment_method TEXT DEFAULT 'account'",
        "ALTER TABLE free_rides ADD COLUMN c_invoice_number TEXT",
        "ALTER TABLE free_rides ADD COLUMN c_invoice_note TEXT",
        "ALTER TABLE free_rides ADD COLUMN currency TEXT DEFAULT 'BAM'",
        "ALTER TABLE free_rides ADD COLUMN issue_date TEXT",
    ]:
        try:
            c.execute(_sql)
        except Exception:
            pass
    c.execute("CREATE TABLE IF NOT EXISTS customers (id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT NOT NULL UNIQUE,id_number TEXT,vat_number TEXT,address TEXT,city TEXT,country TEXT,contact_person TEXT,phone TEXT,email TEXT,note TEXT,active TEXT DEFAULT 'Da')")
    c.commit()
    c.close()

ensure_proformas_table()
# Compatibility: store both ID and PDV number on newer predračuni.
try:
    _c=db()
    _c.execute("ALTER TABLE proformas ADD COLUMN client_vat_number TEXT")
    _c.commit()
    _c.close()
except Exception:
    try: _c.close()
    except Exception: pass


def ensure_proforma_items_table():
    c=db()
    c.execute("""CREATE TABLE IF NOT EXISTS proforma_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        proforma_id INTEGER NOT NULL,
        item_date TEXT,
        relation TEXT,
        km_total REAL DEFAULT 0,
        price_per_km REAL DEFAULT 0,
        amount REAL DEFAULT 0,
        sort_order INTEGER DEFAULT 0
    )""")
    c.commit()
    c.close()

ensure_proforma_items_table()

def ensure_manual_c_invoices_table():
    c=db()
    try: c.execute("ALTER TABLE manual_c_invoices ADD COLUMN issued_by TEXT")
    except Exception: pass
    c.execute("""CREATE TABLE IF NOT EXISTS manual_c_invoices (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        document_no TEXT UNIQUE,
        client TEXT,
        issue_date TEXT,
        currency TEXT DEFAULT 'BAM',
        note TEXT,
        issued_by TEXT,
        created_at TEXT
    )""")
    existing={r["name"] for r in c.execute("PRAGMA table_info(manual_c_invoices)").fetchall()}
    if "issued_by" not in existing:
        c.execute("ALTER TABLE manual_c_invoices ADD COLUMN issued_by TEXT")
    if "payment_status" not in existing:
        c.execute("ALTER TABLE manual_c_invoices ADD COLUMN payment_status TEXT DEFAULT 'Neplaćeno'")
    c.execute("""CREATE TABLE IF NOT EXISTS manual_c_invoice_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        invoice_id INTEGER NOT NULL,
        description TEXT,
        amount REAL DEFAULT 0,
        sort_order INTEGER DEFAULT 0
    )""")
    c.commit(); c.close()

ensure_manual_c_invoices_table()

def _issued_by_name():
    """Return the currently logged-in user's full name for documents."""
    try:
        u = auth_user()
        if u is not None:
            try:
                return str(u["full_name"] or u["username"] or "").strip()
            except Exception:
                return str(getattr(u, "full_name", "") or getattr(u, "username", "") or "").strip()
    except Exception:
        pass
    return ""

def _stored_document_issuer(row):
    try:
        name=str(dict(row).get("issued_by") or "").strip()
        if name:
            return name
    except Exception:
        pass
    # Legacy documents created before issued_by existed.
    return _document_issuer_name()


def _document_issuer_name():
    try:
        u=auth_user()
        if u is not None:
            data=dict(u)
            name=str(data.get("full_name") or data.get("name") or data.get("username") or "").strip()
            if name:
                return name
    except Exception:
        pass
    # Fallback for cases where another uncommitted DB transaction temporarily
    # prevents auth_user from opening a second connection.
    return str(session.get("user_full_name") or session.get("username") or "").strip()

def _document_signature_footer(styles, issuer=""):
    """Visible two-column signature block for all PDF documents."""
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
    issued=str(issuer or "").strip()
    left=Paragraph("Preuzeo:<br/><br/><br/>______________________________", styles["Normal"])
    right=Paragraph((f"Izdao: {issued}" if issued else "Izdao:")+"<br/><br/><br/>______________________________", styles["Normal"])
    t=Table([[left,right]], colWidths=[245,245], rowHeights=[62], hAlign="CENTER")
    t.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("ALIGN",(0,0),(0,0),"LEFT"),
        ("ALIGN",(1,0),(1,0),"RIGHT"),
        ("LEFTPADDING",(0,0),(-1,-1),0),
        ("RIGHTPADDING",(0,0),(-1,-1),0),
        ("TOPPADDING",(0,0),(-1,-1),0),
        ("BOTTOMPADDING",(0,0),(-1,-1),0),
    ]))
    return [Spacer(1,10), t]


def _next_document_number(conn, document_type, year=None):
    """Return the next sequential document number in the form NNN/YYYY."""
    year = int(year or datetime.now().year)
    if document_type == "proforma":
        table, column, start_no = "proformas", "number", 40
    else:
        table, column, start_no = "free_rides", "document_no", 0

    try:
        rows = conn.execute(
            f"SELECT {column} AS n FROM {table} WHERE {column} IS NOT NULL AND TRIM({column})<>''"
        ).fetchall()
        if document_type == "c_invoice":
            try:
                rows += conn.execute("SELECT document_no AS n FROM manual_c_invoices WHERE document_no IS NOT NULL AND TRIM(document_no)<>''").fetchall()
            except Exception:
                pass
    except Exception:
        rows = []

    highest = start_no
    suffix = f"/{year}"
    for row in rows:
        value = str(row["n"] or "").strip()
        if not value.endswith(suffix):
            continue
        first = value.split("/", 1)[0].strip()
        # Accept existing forms such as 040/2026, PR-040/2026 or C-040/2026.
        match = re.search(r"(\d+)$", first)
        if match:
            highest = max(highest, int(match.group(1)))

    return f"{highest + 1:03d}/{year}"


def _num(v):
    try:
        return float(str(v or 0).replace(",", "."))
    except Exception:
        return 0.0

def _save_items(pid, form, connection=None, price_override=None):
    own_connection = connection is None
    c = connection or db()
    rels=form.getlist("item_relation[]")
    dates=form.getlist("item_date[]")
    kms=form.getlist("item_km[]")
    prices=form.getlist("item_price[]")
    amounts=form.getlist("item_amount[]")

    # If the edit form did not submit item rows, keep existing rows intact.
    if not rels and not dates and not kms and not prices:
        total_row=c.execute("SELECT COALESCE(SUM(amount),0) AS total FROM proforma_items WHERE proforma_id=?",(pid,)).fetchone()
        total=_num(total_row["total"] if total_row else 0)
        if own_connection:
            c.close()
        return total

    c.execute("DELETE FROM proforma_items WHERE proforma_id=?", (pid,))
    total=0.0
    for i in range(max(len(rels),len(dates),len(kms),len(prices),len(amounts))):
        rel=(rels[i] if i < len(rels) else "").strip()
        date=dates[i] if i < len(dates) else ""
        km=_num(kms[i] if i < len(kms) else 0)
        price=_num(prices[i] if i < len(prices) else 0)
        # Cijena se uvijek ponovno računa iz trenutne kilometraže i cijene po km.
        amount=km*price
        if not rel and not date and km==0 and price==0 and amount==0:
            continue
        total += amount
        c.execute("""INSERT INTO proforma_items
            (proforma_id,item_date,relation,km_total,price_per_km,amount,sort_order)
            VALUES(?,?,?,?,?,?,?)""",(pid,date,rel,km,price,amount,i))
    if own_connection:
        c.commit(); c.close()
    return total

def ensure_users_table():
    c=db()
    c.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        full_name TEXT,
        role TEXT NOT NULL DEFAULT 'user',
        permissions TEXT DEFAULT '',
        active INTEGER NOT NULL DEFAULT 1
    )""")
    # Migrate any existing users table safely.
    cols={r[1] for r in c.execute("PRAGMA table_info(users)").fetchall()}
    for col,typ in {
        "password_hash":"TEXT",
        "full_name":"TEXT",
        "role":"TEXT DEFAULT 'user'",
        "permissions":"TEXT DEFAULT ''",
        "active":"INTEGER NOT NULL DEFAULT 1"
    }.items():
        if col not in cols:
            c.execute(f"ALTER TABLE users ADD COLUMN {col} {typ}")
    # Always ensure the built-in admin account is usable.
    c.execute("""INSERT INTO users(username,password_hash,full_name,role,permissions,active)
                 VALUES(?,?,?,?,?,1)
                 ON CONFLICT(username) DO UPDATE SET
                   password_hash=excluded.password_hash,
                   full_name=excluded.full_name,
                   role='admin',
                   permissions='*',
                   active=1""",
              ("admin",hash_password("admin"),"Administrator","admin","*"))
    c.commit()
    c.close()

ensure_users_table()

def ensure_audit_table():
    c=db()
    c.execute("""CREATE TABLE IF NOT EXISTS audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        created_at TEXT NOT NULL,
        user_id TEXT,
        username TEXT,
        full_name TEXT,
        action TEXT NOT NULL,
        endpoint TEXT,
        method TEXT,
        path TEXT,
        details TEXT
    )""")
    c.commit()
    c.close()

def audit_log(action, details="", endpoint=None, method=None, path=None):
    try:
        u=auth_user()
        if u:
            c=db()
            c.execute("""INSERT INTO audit_log
                (created_at,user_id,username,full_name,action,endpoint,method,path,details)
                VALUES(?,?,?,?,?,?,?,?,?)""",
                (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                 str(u["id"]), str(u["username"]), str(u["full_name"] or ""),
                 action, endpoint or request.endpoint or "", method or request.method,
                 path or request.path, details or ""))
            c.commit(); c.close()
    except Exception:
        pass

ensure_audit_table()

def init_tire_db():
 c=db()
 c.executescript("""
 CREATE TABLE IF NOT EXISTS tire_purchases (
   id INTEGER PRIMARY KEY AUTOINCREMENT,
   purchase_date TEXT NOT NULL,
   supplier TEXT,
   invoice TEXT,
   warehouse TEXT NOT NULL DEFAULT 'Međugorje',
   dimension TEXT NOT NULL,
   brand TEXT,
   tire_type TEXT,
   quantity INTEGER NOT NULL DEFAULT 1,
   note TEXT
 );
 CREATE TABLE IF NOT EXISTS tires (
   id INTEGER PRIMARY KEY AUTOINCREMENT,
   serial_number TEXT UNIQUE,
   dimension TEXT NOT NULL,
   brand TEXT,
   tire_type TEXT,
   status TEXT NOT NULL DEFAULT 'Skladište',
   warehouse TEXT,
   vehicle TEXT,
   position TEXT,
   mounted_date TEXT,
   mounted_km INTEGER,
   demounted_date TEXT,
   demounted_km INTEGER,
   responsible TEXT,
   note TEXT
 );
 CREATE TABLE IF NOT EXISTS tire_events (
   id INTEGER PRIMARY KEY AUTOINCREMENT,
   tire_id INTEGER,
   event_type TEXT NOT NULL,
   event_date TEXT NOT NULL,
   vehicle TEXT,
   position TEXT,
   km INTEGER,
   warehouse TEXT,
   responsible TEXT,
   reason TEXT,
   note TEXT,
   FOREIGN KEY(tire_id) REFERENCES tires(id)
 );
 """)
 c.commit(); c.close()

try:
 init_tire_db()
except Exception:
 pass

WEEKDAYS=[("PON","Ponedjeljak"),("UTO","Utorak"),("SRI","Srijeda"),("ČET","Četvrtak"),("PET","Petak"),("SUB","Subota"),("NED","Nedjelja")]
def line_runs_on(line, target_date):
    days=(line["schedule_day"] or "").strip()
    if not days:
        return False
    allowed={x.strip().upper() for x in days.split(",") if x.strip()}
    code={0:"PON",1:"UTO",2:"SRI",3:"ČET",4:"PET",5:"SUB",6:"NED"}[target_date.weekday()]
    return code in allowed
def line_group(line):
    g=(line["group_type"] or "").strip()
    return g if g in ("D0","D+1") else "D0"
def planned_rows_for_dates(c, selected):
 from datetime import timedelta
 next_day=selected+timedelta(days=1)
 lines=c.execute("""SELECT * FROM lines
                    WHERE active='Da' AND group_type IN ('D0','D+1') AND COALESCE(internal_return,0)=0
                    ORDER BY departure,name""").fetchall()
 saved=c.execute("""SELECT * FROM schedules
                    WHERE date IN (?,?)
                      AND COALESCE(hidden_from_schedule,0)=0
                      AND (source_date IS NULL OR source_date=?)
                    ORDER BY date,time,line""",
                 (selected.isoformat(),next_day.isoformat(),
                  selected.isoformat())).fetchall()
 saved_map={(r["date"],r["line"].strip()):r for r in saved}
 rows=[]
 for line in lines:
  group=line_group(line)
  target=selected if group=="D0" else next_day
  if not line_runs_on(line,target):
   continue
  key=(target.isoformat(),line["name"].strip())
  r=saved_map.get(key)
  if r:
   rows.append(dict(r))
  else:
   rows.append({"id":None,"date":target.isoformat(),"line":line["name"].strip(),
    "time":line["departure"] or "","driver1":"","driver2":"","vehicle":"",
    "note":"","group_name":"Danas" if group=="D0" else "Sutra","planned":True})
 rows.sort(key=lambda r:(r["date"],r["time"] or "99:99",(r["line"] or "").strip().lower()))
 return rows,next_day

@app.context_processor
def customer_select_options():
    try:
        c=db()
        rows=c.execute("SELECT id,name,id_number,vat_number,address,city,country,contact_person,phone,email,note FROM customers WHERE COALESCE(active,'Da')!='Ne' ORDER BY name COLLATE NOCASE").fetchall()
        c.close()
    except Exception:
        rows=[]
    return {"customer_rows":rows, "free_ride_customer_options":rows}

@app.context_processor
def inject_customer_lookup():
    def customer_by_name(name):
        return _customer_data(name)
    return {"customer_by_name":customer_by_name}

@app.context_processor
def inject_customers():
    try:
        c=db()
        rows=c.execute("SELECT name FROM customers WHERE COALESCE(active,'Da')!='Ne' ORDER BY name").fetchall()
        c.close()
    except Exception:
        rows=[]
    return {"customer_rows":rows}

@app.context_processor
def ctx(): return {"today":date.today().isoformat()}

TIRE_WAREHOUSES=["Međugorje","Kiseljak","Zagreb"]
TIRE_TYPES=["Upravljačka","Pogonska","Prateća"]
TIRE_POSITIONS=["Upravljačka","Pogonska","Prateća"]


def permission_required(view_permission, edit_permission=None):
    """Enforce permissions on the server.
    For edit routes: GET requires view, POST requires edit.
    Routes that are always write operations should pass only edit_permission
    (or use the same permission as both arguments).
    """
    def decorator(func):
        @wraps(func)
        def wrapped(*args, **kwargs):
            user=auth_user()
            required=edit_permission if (edit_permission and request.method in ("POST","PUT","PATCH","DELETE")) else view_permission
            if not has_permission(user, required):
                return render_template("403.html"), 403
            return func(*args, **kwargs)
        return wrapped
    return decorator

@permission_required("tires_export")

def sync_free_ride_clients_to_customers():
    """Copy distinct existing free-ride client names into master customers once."""
    try:
        c = db()
        try:
            free_cols = {r["name"] for r in c.execute("PRAGMA table_info(free_rides)").fetchall()}
            cust_cols = {r["name"] for r in c.execute("PRAGMA table_info(customers)").fetchall()}
            client_col = next((x for x in ("client", "customer", "client_name") if x in free_cols), None)
            name_col = next((x for x in ("name", "client", "customer_name") if x in cust_cols), None)
            if client_col and name_col:
                names = c.execute(
                    f'SELECT DISTINCT TRIM("{client_col}") AS n FROM free_rides '
                    f'WHERE TRIM(COALESCE("{client_col}", "")) <> ""'
                ).fetchall()
                for row in names:
                    n = row["n"]
                    exists = c.execute(
                        f'SELECT 1 FROM customers WHERE LOWER(TRIM("{name_col}"))=LOWER(TRIM(?)) LIMIT 1',
                        (n,)
                    ).fetchone()
                    if not exists:
                        c.execute(f'INSERT INTO customers ("{name_col}") VALUES (?)', (n,))
                c.commit()
        finally:
            c.close()
    except Exception:
        pass


try:
    sync_free_ride_clients_to_customers()
except Exception:
    pass


def free_ride_first_date(conn, ride_id, fallback=None):
    """Return the earliest actual date from ride items for a free ride."""
    try:
        row = conn.execute(
            """
            SELECT item_date
            FROM free_ride_items
            WHERE free_ride_id=?
              AND item_date IS NOT NULL
              AND TRIM(item_date) <> ''
            ORDER BY sort_order ASC, id ASC
            LIMIT 1
            """,
            (ride_id,)
        ).fetchone()
        if row and row[0]:
            return row[0]
    except Exception:
        pass
    return fallback


def enrich_free_ride_rows(conn, rows):
    """Ensure list views show the real issuer and total amount for free rides."""
    out=[]
    for row in rows:
        data=dict(row)
        try:
            if not str(data.get("issued_by") or "").strip():
                data["issued_by"]=_stored_document_issuer(row)
        except Exception:
            pass

        # Older free rides can have amount=0 although item rows contain the
        # actual KM × price calculation. Recalculate for display.
        try:
            current=_num(data.get("amount"))
            if current <= 0:
                total_row=conn.execute(
                    """SELECT COALESCE(SUM(
                           CASE
                             WHEN COALESCE(amount,0)<>0 THEN amount
                             ELSE COALESCE(km_total,0)*COALESCE(price_per_km,0)
                           END
                       ),0) AS total
                       FROM free_ride_items WHERE free_ride_id=?""",
                    (data.get("id"),)
                ).fetchone()
                total=_num(total_row["total"] if total_row else 0)
                if total > 0:
                    data["amount"]=total
        except Exception:
            pass
        out.append(data)
    return out

def apply_free_ride_first_dates(conn, rows):
    out = []
    for row in rows:
        try:
            d = free_ride_first_date(conn, row["id"], row["date"] if "date" in row.keys() else None)
            if hasattr(row, "keys") and "date" in row.keys() and d:
                row = dict(row)
                row["date"] = d
        except Exception:
            pass
        out.append(row)
    return out

@app.route("/gume/izvoz")
def tire_export():
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.page import PageMargins

    d1s=request.args.get("date_from") or request.args.get("from") or "2026-01-01"
    d2s=request.args.get("date_to") or request.args.get("to") or date.today().isoformat()
    try:
        d1=datetime.strptime(d1s,"%Y-%m-%d").date()
        d2=datetime.strptime(d2s,"%Y-%m-%d").date()
    except ValueError:
        d1=date.today(); d2=d1

    c=db()
    purchases=c.execute("""SELECT * FROM tire_purchases
                           WHERE purchase_date BETWEEN ? AND ?
                           ORDER BY purchase_date,id""",
                        (d1.isoformat(),d2.isoformat())).fetchall()
    events=c.execute("""SELECT e.*,t.serial_number,t.dimension,t.brand,t.tire_type,t.status
                        FROM tire_events e LEFT JOIN tires t ON t.id=e.tire_id
                        WHERE e.event_date BETWEEN ? AND ?
                        ORDER BY e.event_date,e.id""",
                     (d1.isoformat(),d2.isoformat())).fetchall()
    all_tires=c.execute("""SELECT * FROM tires ORDER BY status,warehouse,vehicle,position,id""").fetchall()
    c.close()

    wb=Workbook()
    ws=wb.active; ws.title="Pregled"; ws.sheet_view.showGridLines=False
    navy="17365D"; blue="2F75B5"; light="D9EAF7"; border_color="AAB7C4"
    def setup(ws,headers):
        for col,h in enumerate(headers,1):
            cell=ws.cell(1,col,h); cell.font=Font(name="Arial",bold=True,size=11,color="FFFFFF")
            cell.fill=PatternFill("solid",fgColor=blue); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{chr(64+len(headers))}1"
        for i in range(1,len(headers)+1): ws.column_dimensions[chr(64+i)].width=20
    def add_rows(ws,headers,rows):
        setup(ws,headers); thin=Side(style="thin",color=border_color)
        for rr,row in enumerate(rows,2):
            for cc,val in enumerate(row,1):
                cell=ws.cell(rr,cc,"" if val is None else str(val))
                cell.font=Font(name="Arial",size=11)
                cell.border=Border(left=thin,right=thin,top=thin,bottom=thin)
                cell.alignment=Alignment(vertical="center",wrap_text=True)
        for cc in range(1,len(headers)+1):
            mx=max([len(str(ws.cell(rr,cc).value or "")) for rr in range(1,ws.max_row+1)] or [10])
            ws.column_dimensions[chr(64+cc)].width=min(45,max(14,mx+2))

    headers=["Datum","Vrsta","Guma/Dimenzija","Marka","Serijski broj","Autobus","Pozicija","KM","Skladište","Odgovorna osoba","Razlog"]
    overview=[]
    for e in events:
        overview.append([e["event_date"],e["event_type"],e["dimension"],e["brand"],e["serial_number"],e["vehicle"],e["position"],e["km"],e["warehouse"],e["responsible"],e["reason"]])
    for pch in purchases:
        overview.append([pch["purchase_date"],"Nabava",pch["dimension"],pch["brand"],"", "",pch["tire_type"],pch["quantity"],pch["warehouse"],"",pch["supplier"]])
    overview.sort(key=lambda x:(x[0] or "",x[1] or ""))
    add_rows(ws,headers,overview)

    ws2=wb.create_sheet("Nabave")
    add_rows(ws2,["Datum","Dobavljač","Račun","Skladište","Dimenzija","Marka","Tip","Količina","Napomena"],
             [[p["purchase_date"],p["supplier"],p["invoice"],p["warehouse"],p["dimension"],p["brand"],p["tire_type"],p["quantity"],p["note"]] for p in purchases])

    ws3=wb.create_sheet("Montaže")
    add_rows(ws3,["Datum","Dimenzija","Marka","Serijski broj","Autobus","Pozicija","KM","Skladište","Odgovorna osoba","Napomena"],
             [[e["event_date"],e["dimension"],e["brand"],e["serial_number"],e["vehicle"],e["position"],e["km"],e["warehouse"],e["responsible"],e["note"]] for e in events if e["event_type"]=="Montaža"])

    ws4=wb.create_sheet("Demontaže")
    add_rows(ws4,["Datum","Dimenzija","Marka","Serijski broj","Autobus","Pozicija","KM","Skladište","Odgovorna osoba","Razlog","Napomena"],
             [[e["event_date"],e["dimension"],e["brand"],e["serial_number"],e["vehicle"],e["position"],e["km"],e["warehouse"],e["responsible"],e["reason"],e["note"]] for e in events if e["event_type"]=="Demontaža"])

    ws5=wb.create_sheet("Sve gume")
    add_rows(ws5,["ID","Serijski broj","Dimenzija","Marka","Tip","Status","Skladište","Autobus","Pozicija","Datum montaže","KM montaže","Datum demontaže","KM demontaže","Odgovorna osoba","Napomena"],
             [[t["id"],t["serial_number"],t["dimension"],t["brand"],t["tire_type"],t["status"],t["warehouse"],t["vehicle"],t["position"],t["mounted_date"],t["mounted_km"],t["demounted_date"],t["demounted_km"],t["responsible"],t["note"]] for t in all_tires])

    # Separate warehouse tabs: only tires that are currently physically in stock.
    # A tire is marked as "Demontirana – vraćena" when its latest event is a
    # demount event and it is currently back in a warehouse. Otherwise it is
    # treated as a tire that entered stock through purchase/manual entry.
    warehouse_rows={}
    for warehouse in TIRE_WAREHOUSES:
        warehouse_rows[warehouse]=c_rows=[]
        # c is already closed above, so query through a fresh DB connection.
        cx=db()
        rows_wh=cx.execute("""
            SELECT t.*,
                   (SELECT e.event_type FROM tire_events e
                    WHERE e.tire_id=t.id ORDER BY e.id DESC LIMIT 1) AS last_event_type,
                   (SELECT e.event_date FROM tire_events e
                    WHERE e.tire_id=t.id ORDER BY e.id DESC LIMIT 1) AS last_event_date
            FROM tires t
            WHERE t.status='Skladište' AND COALESCE(t.warehouse,'')=?
            ORDER BY t.dimension,t.brand,t.serial_number,t.id
        """,(warehouse,)).fetchall()
        cx.close()
        for r in rows_wh:
            entry="Demontirana – vraćena u skladište" if (r["last_event_type"] or "")=="Demontaža" else "Nabavljena / unesena"
            warehouse_rows[warehouse].append([
                r["serial_number"],r["dimension"],r["brand"],r["tire_type"],
                r["warehouse"],entry,r["last_event_date"] if entry.startswith("Demontirana") else "",
                r["responsible"],r["note"]
            ])

    for warehouse in TIRE_WAREHOUSES:
        safe_name=f"Skladište {warehouse}"[:31]
        ws_wh=wb.create_sheet(safe_name)
        add_rows(ws_wh,
                 ["Serijski broj","Dimenzija","Marka","Tip","Skladište","Način ulaska u skladište","Datum demontaže","Odgovorna osoba","Napomena"],
                 warehouse_rows[warehouse])

    bio=io.BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio,as_attachment=True,
                     download_name=f"izvoz_guma_{d1.strftime('%d-%m-%Y')}_{d2.strftime('%d-%m-%Y')}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@permission_required("tires_view")
@app.route("/gume")
def tires_dashboard():
 c=db()
 stock=c.execute("""SELECT dimension,COALESCE(brand,'') brand,COALESCE(tire_type,'') tire_type,
                    COALESCE(warehouse,'') warehouse,COUNT(*) quantity
                    FROM tires WHERE status='Skladište'
                    GROUP BY dimension,brand,tire_type,warehouse
                    ORDER BY dimension,brand,warehouse""").fetchall()

 # Pregled ukupnog stanja po skladištima. Uvijek prikazujemo poznata skladišta,
 # čak i kada trenutno nemaju nijednu gumu.
 known_warehouses=list(TIRE_WAREHOUSES)
 extra=[r[0] for r in c.execute("""
     SELECT DISTINCT warehouse FROM tires
     WHERE status='Skladište' AND COALESCE(TRIM(warehouse),'')<>''
     ORDER BY warehouse
 """).fetchall()]
 warehouses=[]
 for warehouse in known_warehouses + extra:
  if warehouse not in [x["name"] for x in warehouses]:
   warehouses.append({
    "name": warehouse,
    "quantity": c.execute(
     "SELECT COUNT(*) FROM tires WHERE status='Skladište' AND COALESCE(warehouse,'')=?",
     (warehouse,)
    ).fetchone()[0]
   })

 stats={
  "stock":c.execute("SELECT COUNT(*) FROM tires WHERE status='Skladište'").fetchone()[0],
  "mounted":c.execute("SELECT COUNT(*) FROM tires WHERE status='Montirana'").fetchone()[0],
  "scrapped":c.execute("SELECT COUNT(*) FROM tires WHERE status='Rashod'").fetchone()[0],
  "review":c.execute("SELECT COUNT(*) FROM tire_events WHERE event_type='Demontaža'").fetchone()[0],
  "purchased":c.execute("SELECT COALESCE(SUM(quantity),0) FROM tire_purchases").fetchone()[0],
  "total":c.execute("SELECT COUNT(*) FROM tires").fetchone()[0]
 }
 recent=c.execute("""SELECT e.*,t.serial_number,t.dimension,t.brand
                     FROM tire_events e LEFT JOIN tires t ON t.id=e.tire_id
                     ORDER BY e.id DESC LIMIT 15""").fetchall()
 c.close()
 return render_template("tires_dashboard.html",stock=stock,stats=stats,
                        warehouses=warehouses,recent=recent)

@permission_required("tires_view")
@app.route("/gume/skladista")
def tire_warehouses():
 c=db()
 warehouses=[]
 for warehouse in TIRE_WAREHOUSES:
  quantity=c.execute(
   "SELECT COUNT(*) FROM tires WHERE status='Skladište' AND COALESCE(TRIM(warehouse),'')=?",
   (warehouse,)
  ).fetchone()[0]
  warehouses.append({"name":warehouse,"quantity":quantity})
 c.close()
 return render_template("tire_warehouses.html",warehouses=warehouses)

@permission_required("tires_view")
@app.route("/gume/skladiste/<path:warehouse>")
def tire_warehouse(warehouse):
 c=db()
 rows=c.execute("""
     SELECT *
     FROM tires
     WHERE status='Skladište' AND COALESCE(warehouse,'')=?
     ORDER BY dimension, brand, tire_type, serial_number, id
 """,(warehouse,)).fetchall()
 c.close()
 u=auth_user()
 return render_template("tire_warehouse.html",warehouse=warehouse,rows=rows,can_edit=has_permission(u,"tires_edit"),is_admin=bool(u and str(u["role"]).lower()=="admin"))

@permission_required("tires_view")
@app.route("/gume/gume")
def tire_list():
 q=request.args.get("q","").strip()
 c=db()
 rows=c.execute("""SELECT * FROM tires
                   WHERE serial_number LIKE ? OR dimension LIKE ? OR brand LIKE ? OR vehicle LIKE ?
                   ORDER BY status,dimension,brand,serial_number""",
                tuple(["%"+q+"%"]*4)).fetchall()
 c.close()
 return render_template("tires.html",rows=rows,q=q)

@permission_required("tires_edit")
@app.route("/gume/nabava",methods=["GET","POST"])
def tire_purchase():
 c=db()
 if request.method=="POST":
  d=request.form.get("purchase_date") or date.today().isoformat()
  supplier=request.form.get("supplier","").strip()
  invoice=request.form.get("invoice","").strip()
  warehouse=request.form.get("warehouse","Međugorje")
  dim=request.form.get("dimension","").strip()
  brand=request.form.get("brand","").strip()
  typ=request.form.get("tire_type","").strip()
  qty=max(1,int(request.form.get("quantity","1") or 1))
  note=request.form.get("note","").strip()
  serials=[x.strip() for x in request.form.getlist("serial_numbers") if x.strip()]
  if not dim:
   flash("Dimenzija je obavezna.","danger")
  elif serials and len(serials)!=qty:
   flash(f"Ako unosiš serijske brojeve, potrebno je unijeti točno {qty} serijski broj(a), jedan po redu.","danger")
  elif len(set(serials))!=len(serials):
   flash("Serijski broj ne smije biti ponovljen u istoj nabavi.","danger")
  elif any(c.execute("SELECT 1 FROM tires WHERE serial_number=?",(x,)).fetchone() for x in serials):
   flash("Jedan od unesenih serijskih brojeva već postoji u bazi.","danger")
  else:
   c.execute("""INSERT INTO tire_purchases(purchase_date,supplier,invoice,warehouse,dimension,brand,tire_type,quantity,note)
                VALUES(?,?,?,?,?,?,?,?,?)""",(d,supplier,invoice,warehouse,dim,brand,typ,qty,note))
   for idx in range(qty):
    serial=serials[idx] if serials else None
    c.execute("""INSERT INTO tires(serial_number,dimension,brand,tire_type,status,warehouse,note)
                 VALUES(?,?,?,?,?,?,?)""",(serial,dim,brand,typ,"Skladište",warehouse,note))
   c.commit(); c.close()
   flash(f"Dodano {qty} guma u skladište.","success")
   return redirect(url_for("tire_list"))
 c.close()
 return render_template("tire_purchase_form.html",warehouses=TIRE_WAREHOUSES,types=TIRE_TYPES,today=date.today().isoformat())

@permission_required("tires_edit")
@app.route("/gume/dodaj",methods=["GET","POST"])
def tire_add():
 c=db()
 if request.method=="POST":
  serial=request.form.get("serial_number","").strip() or None
  dim=request.form.get("dimension","").strip()
  brand=request.form.get("brand","").strip()
  typ=request.form.get("tire_type","").strip()
  warehouse=request.form.get("warehouse","Međugorje")
  note=request.form.get("note","").strip()
  if not dim: flash("Dimenzija je obavezna.","danger")
  elif serial and c.execute("SELECT 1 FROM tires WHERE serial_number=?",(serial,)).fetchone():
   flash("Serijski broj već postoji.","danger")
  else:
   c.execute("""INSERT INTO tires(serial_number,dimension,brand,tire_type,status,warehouse,note)
                VALUES(?,?,?,?,?,?,?)""",(serial,dim,brand,typ,"Skladište",warehouse,note))
   c.commit();c.close();flash("Guma je dodana u skladište.","success");return redirect(url_for("tire_list"))
 c.close();return render_template("tire_add_form.html",warehouses=TIRE_WAREHOUSES,types=TIRE_TYPES)

@permission_required("tires_view", "tires_edit")
@app.route("/gume/uredi/<int:tire_id>",methods=["GET","POST"])
def tire_edit(tire_id):
 c=db()
 tire=c.execute("SELECT * FROM tires WHERE id=?",(tire_id,)).fetchone()
 if not tire:
  c.close(); flash("Guma nije pronađena.","danger"); return redirect(url_for("tire_list"))
 if request.method=="POST":
  serial=request.form.get("serial_number","").strip() or None
  dim=request.form.get("dimension","").strip()
  brand=request.form.get("brand","").strip()
  typ=request.form.get("tire_type","").strip()
  status=request.form.get("status","Skladište").strip() or "Skladište"
  warehouse=request.form.get("warehouse","").strip() or None
  vehicle=request.form.get("vehicle","").strip() or None
  position=request.form.get("position","").strip() or None
  mounted_date=request.form.get("mounted_date","").strip() or None
  mounted_km=request.form.get("mounted_km","").strip() or None
  demounted_date=request.form.get("demounted_date","").strip() or None
  demounted_km=request.form.get("demounted_km","").strip() or None
  responsible=request.form.get("responsible","").strip() or None
  note=request.form.get("note","").strip() or None
  duplicate=c.execute("SELECT id FROM tires WHERE serial_number=? AND id<>?",(serial,tire_id)).fetchone() if serial else None
  if not dim:
   flash("Dimenzija je obavezna.","danger")
  elif duplicate:
   flash("Serijski broj već postoji na drugoj gumi.","danger")
  else:
   c.execute("""UPDATE tires SET serial_number=?,dimension=?,brand=?,tire_type=?,status=?,warehouse=?,
                vehicle=?,position=?,mounted_date=?,mounted_km=?,demounted_date=?,demounted_km=?,
                responsible=?,note=? WHERE id=?""",
             (serial,dim,brand,typ,status,warehouse,vehicle,position,mounted_date,mounted_km,
              demounted_date,demounted_km,responsible,note,tire_id))
   c.commit(); c.close()
   flash("Podaci o gumi su izmijenjeni.","success")
   nxt=request.form.get("next","").strip()
   return redirect(nxt if nxt.startswith("/") else url_for("tire_list"))
 vehicles=c.execute("SELECT registration, seats FROM vehicles ORDER BY registration").fetchall()
 c.close()
 return render_template("tire_edit_form.html",tire=tire,warehouses=TIRE_WAREHOUSES,types=TIRE_TYPES,
                        positions=TIRE_POSITIONS,vehicles=vehicles)

@app.route("/gume/obrisi/<int:tire_id>",methods=["POST"])
def tire_delete(tire_id):
 u=auth_user()
 if not u or str(u["role"]).lower()!="admin": return render_template("403.html"),403
 c=db(); c.execute("DELETE FROM tires WHERE id=?",(tire_id,)); c.commit(); c.close()
 flash("Guma je obrisana.","success"); return redirect(url_for("tire_list"))

@permission_required("tires_edit")
@app.route("/gume/montaza",methods=["GET","POST"])
def tire_mount():
 c=db()
 vehicles=c.execute("SELECT registration FROM vehicles WHERE active='Da' ORDER BY registration").fetchall()
 available=c.execute("""SELECT * FROM tires WHERE status='Skladište'
                        ORDER BY tire_type,dimension,brand,serial_number""").fetchall()
 if request.method=="POST":
  tire_ids=request.form.getlist("tire_id")
  vehicle=request.form.get("vehicle","").strip()
  posi=request.form.get("position","").strip()
  d=request.form.get("event_date") or date.today().isoformat()
  km=int(request.form.get("km","0") or 0)
  resp=request.form.get("responsible","").strip()
  if not tire_ids or not vehicle or not posi:
   flash("Odaberi gumu, vozilo i poziciju.","danger")
  else:
   for tid in tire_ids:
    tire=c.execute("SELECT * FROM tires WHERE id=? AND status='Skladište'",(tid,)).fetchone()
    if not tire:
     continue
    # Sigurnosna provjera: tip gume mora odgovarati odabranoj poziciji.
    if (tire["tire_type"] or "").strip().casefold() != posi.strip().casefold():
     c.close()
     flash(f"Ne možeš montirati gumu tipa {tire['tire_type'] or 'nije određeno'} na poziciju {posi}.","danger")
     return redirect(url_for("tire_mount"))
    c.execute("""UPDATE tires SET status='Montirana',warehouse=NULL,vehicle=?,position=?,
                 mounted_date=?,mounted_km=?,demounted_date=NULL,demounted_km=NULL,responsible=?
                 WHERE id=?""",(vehicle,posi,d,km,resp,tid))
    c.execute("""INSERT INTO tire_events(tire_id,event_type,event_date,vehicle,position,km,responsible)
                 VALUES(?,?,?,?,?,?,?)""",(tid,"Montaža",d,vehicle,posi,km,resp))
   c.commit();c.close();flash("Montaža je evidentirana.","success");return redirect(url_for("tire_list"))
 c.close()
 return render_template("tire_mount_form.html",vehicles=vehicles,tires=available,positions=TIRE_POSITIONS)

@permission_required("tires_edit")
@app.route("/gume/demontaza",methods=["GET","POST"])
def tire_demount():
    c=db()
    vehicles=c.execute("""SELECT DISTINCT vehicle FROM tires
                          WHERE status='Montirana' AND vehicle IS NOT NULL AND TRIM(vehicle)!=''
                          ORDER BY vehicle""").fetchall()
    selected_vehicle=request.args.get("vehicle","").strip()

    if request.method=="POST":
        selected_vehicle=request.form.get("vehicle","").strip()
        tid=request.form.get("tire_id")
        d=request.form.get("event_date") or date.today().isoformat()
        km=int(request.form.get("km","0") or 0)
        action=request.form.get("action","warehouse")
        warehouse=request.form.get("warehouse","Međugorje")
        resp=request.form.get("responsible","").strip()
        reason=request.form.get("reason","").strip()

        tire=c.execute("""SELECT * FROM tires
                          WHERE id=? AND status='Montirana' AND vehicle=?""",
                       (tid,selected_vehicle)).fetchone()

        if not tire:
            flash("Odabrana guma nije montirana na odabranom autobusu.","danger")
        elif not selected_vehicle:
            flash("Prvo odaberi autobus.","danger")
        else:
            # Demontirana guma može se vratiti u skladište kao stara/demontirana
            # ili se može trajno rashodovati.
            if action=="scrap":
                status="Rashod"; wh=None
            else:
                status="Skladište"; wh=warehouse

            c.execute("""UPDATE tires SET status=?,warehouse=?,vehicle=NULL,position=NULL,
                         demounted_date=?,demounted_km=?,responsible=?
                         WHERE id=?""",
                      (status,wh,d,km,resp,tid))
            c.execute("""INSERT INTO tire_events
                         (tire_id,event_type,event_date,vehicle,position,km,warehouse,responsible,reason)
                         VALUES(?,?,?,?,?,?,?,?,?)""",
                      (tid,"Demontaža",d,tire["vehicle"],tire["position"],km,wh,resp,reason))
            c.commit(); c.close()
            flash("Demontaža je evidentirana.","success")
            return redirect(url_for("tire_demount"))

    mounted=[]
    if selected_vehicle:
        mounted=c.execute("""SELECT * FROM tires
                             WHERE status='Montirana' AND vehicle=?
                             ORDER BY position,dimension,brand,serial_number""",
                          (selected_vehicle,)).fetchall()

    c.close()
    return render_template("tire_demount_form.html",
                           tires=mounted, vehicles=vehicles,
                           selected_vehicle=selected_vehicle,
                           warehouses=TIRE_WAREHOUSES)

@permission_required("tires_view")
@app.route("/gume/vozilo/<path:vehicle>")
def tire_vehicle(vehicle):
 c=db()
 rows=c.execute("SELECT * FROM tires WHERE status='Montirana' AND vehicle=? ORDER BY position,id",(vehicle,)).fetchall()
 c.close()
 return render_template("tire_vehicle.html",vehicle=vehicle,rows=rows)

def vehicle_expiry_alerts(c, days=10):
    """Return expiry alerts without ever breaking the homepage."""
    try:
        from datetime import datetime, timedelta
        today=date.today()
        limit=today+timedelta(days=days)
        rows=c.execute("""
            SELECT registration, registration_expiry, tachograph_expiry,
                   periodic_expiry, fire_extinguisher_expiry
            FROM vehicles
            WHERE active='Da'
            ORDER BY registration
        """).fetchall()
    except Exception:
        # Older databases are still allowed to open the homepage.
        return []
    alerts=[]
    for row in rows:
        registration=row[0] or ""
        for label,raw in (
            ("Registracija",row[1]),
            ("Tahograf",row[2]),
            ("Periodični",row[3]),
            ("PP aparati",row[4]),
        ):
            if not raw:
                continue
            try:
                expiry=datetime.strptime(str(raw)[:10],"%Y-%m-%d").date()
            except (ValueError,TypeError):
                continue
            left=(expiry-today).days
            if left<=days:
                alerts.append({
                    "registration":str(registration),
                    "type":label,
                    "date":expiry.strftime("%d.%m.%Y"),
                    "days":left,
                    "status":"Isteklo" if left<0 else ("Istječe danas" if left==0 else f"{left} dana")
                })
    return sorted(alerts,key=lambda x:(x["days"],x["registration"],x["type"]))

def line_permit_expiry_alerts(c, days=10):
    """Return line permit expiry alerts for permits expiring within the next N days or already expired."""
    try:
        from datetime import datetime, timedelta
        today=date.today()
        limit=today+timedelta(days=days)
        ensure_line_permits_table(c)
        rows=c.execute("""
            SELECT lp.id, lp.line_id, lp.permit, lp.valid_until, l.name AS line_name
            FROM line_permits lp
            LEFT JOIN lines l ON l.id=lp.line_id
            WHERE lp.valid_until IS NOT NULL AND TRIM(lp.valid_until) <> ''
            ORDER BY lp.valid_until, l.name, lp.permit
        """).fetchall()
    except Exception:
        return []
    alerts=[]
    for row in rows:
        try:
            expiry=datetime.strptime(str(row["valid_until"])[:10],"%Y-%m-%d").date()
        except (ValueError,TypeError):
            continue
        left=(expiry-today).days
        if left<=days:
            alerts.append({
                "id":row["id"],
                "line_id":row["line_id"],
                "line":str(row["line_name"] or ""),
                "permit":str(row["permit"] or ""),
                "date":expiry.strftime("%d.%m.%Y"),
                "days":left,
                "status":"Istekla" if left<0 else ("Istječe danas" if left==0 else f"{left} dana")
            })
    return sorted(alerts,key=lambda x:(x["days"],x["line"].lower(),x["permit"].lower()))

@permission_required("lines_view")
@app.route("/rokovi-dozvole")
def permit_expiry_page():
    c=db()
    alerts=line_permit_expiry_alerts(c,10) or []
    c.close()
    return render_template("permit_expiry.html", expiry_alerts=alerts)

@permission_required("vehicles_view")
@app.route("/rokovi-vozila")
def vehicle_expiry_page():
    # Optional filter from the dashboard: Registracija, PP aparati,
    # Periodični or Tahograf. Without a filter all vehicle warnings are shown.
    selected_type=request.args.get("type","").strip()
    allowed={"Registracija","PP aparati","Periodični","Tahograf"}
    if selected_type not in allowed:
        selected_type=""
    try:
        c=db()
        alerts=vehicle_expiry_alerts(c,10) or []
        c.close()
    except Exception:
        alerts=[]
    if selected_type:
        alerts=[a for a in alerts if a["type"]==selected_type]
    return render_template("vehicle_expiry.html", expiry_alerts=alerts,
                           selected_type=selected_type)

def auth_user():
    uid=session.get("user_id")
    if not uid:
        return None
    if uid == "__admin__":
        class BootstrapAdmin(dict):
            def __getattr__(self, name):
                try:
                    return self[name]
                except KeyError:
                    raise AttributeError(name)
        return BootstrapAdmin(
            id="__admin__", username="admin",
            password_hash=hash_password("admin"),
            full_name="Administrator", role="admin",
            permissions="*", active=1
        )
    try:
        c=db()
        row=c.execute("SELECT * FROM users WHERE id=? AND active=1",(uid,)).fetchone()
        c.close()
        return row
    except Exception:
        return None

def has_permission(user, key):
    if user is None:
        return False
    try:
        role=str(user["role"] or "").strip().lower()
        permissions=str(user["permissions"] or "")
        # Administrator always has access. "*" is also treated as full access
        # so both the built-in admin and database admin behave identically.
        if role == "admin" or permissions.strip() == "*":
            return True
        granted={x.strip() for x in permissions.split(",") if x.strip()}
        if key in granted:
            return True
        # Grupne dozvole i kompatibilnost sa postojećim rutama.
        aliases={
            "schedule_view":{"statistics_view"},
            "free_rides_view":{"statistics_view"},
            "customers_view":{"customers_view"},
        }
        return bool(aliases.get(key,set()) & granted)
    except Exception:
        return False

@app.context_processor
def auth_context():
    return {"current_user": auth_user(), "has_permission": has_permission}

@app.before_request
def auth_guard():
    endpoint=request.endpoint or ""
    if endpoint in ("login", "static"):
        return None
    if not session.get("user_id"):
        return redirect(url_for("login"))
    return None

@app.after_request
def record_user_activity(response):
    # Record meaningful actions, not every page view.
    try:
        if session.get("user_id") and (
            request.method in ("POST","PUT","PATCH","DELETE")
            or "export" in (request.endpoint or "").lower()
            or request.endpoint in ("activities_export",)
        ):
            action_map={
                "login":"Prijava","logout":"Odjava",
                "add_user":"Dodao korisnika","edit_user":"Uredio korisnika","delete_user":"Obrisao korisnika",
                "add_driver":"Dodao vozača","edit_driver":"Uredio vozača","delete_driver":"Obrisao vozača",
                "add_vehicle":"Dodao vozilo","edit_vehicle":"Uredio vozilo","delete_vehicle":"Obrisao vozilo",
                "add_line":"Dodao liniju","edit_line":"Uredio liniju",
                "add_schedule":"Dodao raspored","edit_schedule":"Uredio raspored",
                "tire_purchase":"Dodao nabavu guma","tire_mount":"Montirao gumu","tire_demount":"Demontirao gumu",
                "tire_edit":"Uredio gumu"
            }
            action=action_map.get(request.endpoint or "", "Izvršio radnju")
            details=f"{request.method} {request.path}"
            audit_log(action, details, endpoint=request.endpoint, method=request.method, path=request.path)
    except Exception:
        pass
    return response

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username=(request.form.get("username") or "").strip()
        password=request.form.get("password") or ""

        # Built-in administrator does not depend on the database.
        if username == "admin" and password == "admin":
            session.clear()
            session.permanent=False
            session["user_id"]="__admin__"
            audit_log("Prijava", "Administrator se prijavio")

            # Database synchronization is only a convenience. Even if an
            # existing/legacy database has a problem, admin login must still
            # succeed and redirect to the application.
            try:
                c=db()
                c.execute("""CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    full_name TEXT,
                    role TEXT NOT NULL DEFAULT 'user',
                    permissions TEXT DEFAULT '',
                    active INTEGER NOT NULL DEFAULT 1
                )""")
                row=c.execute("SELECT id FROM users WHERE username='admin' LIMIT 1").fetchone()
                if row:
                    c.execute("""UPDATE users SET password_hash=?,full_name=?,
                                 role='admin',permissions='*',active=1 WHERE id=?""",
                              (hash_password("admin"),"Administrator",row["id"]))
                else:
                    c.execute("""INSERT INTO users
                        (username,password_hash,full_name,role,permissions,active)
                        VALUES(?,?,?,?,?,1)""",
                        ("admin",hash_password("admin"),"Administrator","admin","*"))
                c.commit()
                c.close()
            except Exception:
                pass

            # IMPORTANT: this redirect is outside the database try/except.
            return redirect(url_for("index"))

        # Normal users.
        try:
            c=db()
            row=c.execute(
                "SELECT id,password_hash,active FROM users WHERE username=? LIMIT 1",
                (username,)
            ).fetchone()
            c.close()
            if row and int(row["active"] or 0)==1 and row["password_hash"]==hash_password(password):
                session.clear()
                session.permanent=False
                session["user_id"]=row["id"]
                audit_log("Prijava", "Korisnik se prijavio")
                return redirect(url_for("index"))
        except Exception:
            pass

        return render_template("login.html", error="Pogrešno korisničko ime ili lozinka.")
    return render_template("login.html", error="")

@app.route("/logout", methods=["GET"])
def logout():
    audit_log("Odjava", "Korisnik se odjavio")
    session.clear()
    return redirect(url_for("login"))

@app.route("/aktivnosti")
def activities():
    u=auth_user()
    if not u or str(u["role"]).lower()!="admin":
        return render_template("403.html"),403
    date_from=request.args.get("date_from","").strip()
    date_to=request.args.get("date_to","").strip() or date.today().isoformat()
    q="SELECT * FROM audit_log WHERE 1=1"
    params=[]
    if date_from:
        q+=" AND created_at >= ?"; params.append(date_from+" 00:00:00")
    if date_to:
        q+=" AND created_at <= ?"; params.append(date_to+" 23:59:59")
    q+=" ORDER BY id DESC"
    c=db(); rows=c.execute(q,params).fetchall(); c.close()
    return render_template("activities.html",rows=rows,date_from=date_from,date_to=date_to)

@app.route("/aktivnosti/izvoz")
def activities_export():
    u=auth_user()
    if not u or str(u["role"]).lower()!="admin":
        return render_template("403.html"),403
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    date_from=request.args.get("date_from","").strip()
    date_to=request.args.get("date_to","").strip() or date.today().isoformat()
    q="SELECT created_at,username,full_name,action,details,endpoint,method,path FROM audit_log WHERE 1=1"
    params=[]
    if date_from:
        q+=" AND created_at >= ?"; params.append(date_from+" 00:00:00")
    if date_to:
        q+=" AND created_at <= ?"; params.append(date_to+" 23:59:59")
    q+=" ORDER BY id DESC"
    c=db(); rows=c.execute(q,params).fetchall(); c.close()
    wb=Workbook(); ws=wb.active; ws.title="Aktivnosti"
    ws.append(["Datum i vrijeme","Korisničko ime","Korisnik","Radnja","Detalji","Stranica","Metoda","Putanja"])
    for cell in ws[1]:
        cell.font=Font(bold=True); cell.alignment=Alignment(horizontal="center")
    for r in rows:
        ws.append([r["created_at"],r["username"],r["full_name"],r["action"],r["details"],r["endpoint"],r["method"],r["path"]])
    widths=[20,18,25,28,40,24,10,35]
    for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width=w
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    from io import BytesIO
    out=BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,as_attachment=True,download_name="aktivnosti_globtour.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@permission_required("users_view")
@app.route("/korisnici")
def users():
    u=auth_user()
    if not has_permission(u,"users_view"): return render_template("403.html"),403
    c=db(); rows=c.execute("SELECT id,username,full_name,role,active FROM users ORDER BY username").fetchall(); c.close()
    return render_template("users.html",rows=rows)

@permission_required("users_edit")
@app.route("/korisnici/novi",methods=["GET","POST"])
def add_user():
    if not has_permission(auth_user(),"users_edit"): return render_template("403.html"),403
    if request.method=="POST":
        username=request.form.get("username","").strip()
        password=request.form.get("password","")
        full_name=request.form.get("full_name","").strip()
        role=request.form.get("role","user")
        perms=request.form.getlist("permissions")
        if role == "admin":
            perms=["*"]
        if not username or not password:
            flash("Korisničko ime i lozinka su obavezni.","danger")
        else:
            try:
                c=db(); c.execute("INSERT INTO users(username,password_hash,full_name,role,permissions,active) VALUES(?,?,?,?,?,1)",
                    (username,hash_password(password),full_name,role,",".join(perms))); c.commit(); c.close()
                flash("Korisnik je dodan.","success"); return redirect(url_for("users"))
            except sqlite3.IntegrityError:
                flash("Korisničko ime već postoji.","danger")
    return render_template("user_form.html",user=None,permission_labels=PERMISSION_LABELS)

@permission_required("users_view", "users_edit")
@app.route("/korisnici/<int:id>/uredi",methods=["GET","POST"])
def edit_user(id):
    if not has_permission(auth_user(),"users_edit"): return render_template("403.html"),403
    c=db(); user=c.execute("SELECT * FROM users WHERE id=?",(id,)).fetchone()
    if not user: c.close(); abort(404)
    if request.method=="POST":
        full_name=request.form.get("full_name","").strip()
        role=request.form.get("role","user")
        active=1 if request.form.get("active")=="1" else 0
        perms=request.form.getlist("permissions")
        password=request.form.get("password","")
        if role == "admin":
            perms=["*"]
        if user["role"]=="admin" and user["username"]=="admin":
            role="admin"; active=1
        if password:
            c.execute("UPDATE users SET full_name=?,role=?,permissions=?,active=?,password_hash=? WHERE id=?",
                      (full_name,role,",".join(perms),active,hash_password(password),id))
        else:
            c.execute("UPDATE users SET full_name=?,role=?,permissions=?,active=? WHERE id=?",
                      (full_name,role,",".join(perms),active,id))
        c.commit(); c.close()
        flash("Korisnik je izmijenjen.","success"); return redirect(url_for("users"))
    c.close()
    return render_template("user_form.html",user=user,permission_labels=PERMISSION_LABELS)

@permission_required("users_delete")
@app.route("/korisnici/<int:id>/izbrisi",methods=["POST"])
def delete_user(id):
    if not has_permission(auth_user(),"users_delete"): return render_template("403.html"),403
    c=db(); user=c.execute("SELECT * FROM users WHERE id=?",(id,)).fetchone()
    if not user: c.close(); abort(404)
    if user["username"]=="admin" and user["role"]=="admin":
        c.close(); flash("Glavni administrator se ne može izbrisati.","danger"); return redirect(url_for("users"))
    c.execute("DELETE FROM users WHERE id=?",(id,)); c.commit(); c.close()
    flash("Korisnik je obrisan.","success"); return redirect(url_for("users"))

@permission_required("dashboard")
@app.route("/")
def index():
 from datetime import datetime
 d=request.args.get("date") or date.today().isoformat()
 try: selected=datetime.strptime(d,"%Y-%m-%d").date()
 except ValueError: selected=date.today()
 c=db()
 counts=[c.execute("SELECT COUNT(*) FROM "+t+" WHERE "+col+"='Da'").fetchone()[0]
         for t,col in [("drivers","active"),("vehicles","active"),("lines","active")]]
 rows,next_day=planned_rows_for_dates(c,selected)
 expiry_alerts=vehicle_expiry_alerts(c,10) or []
 permit_expiry_alerts=line_permit_expiry_alerts(c,10) or []
 # Alarm za nadolazeće slobodne vožnje: danas i sljedećih 7 dana.
 # Koristimo stvarni datum slobodne vožnje (date_from), koji se kod više
 # vožnji automatski postavlja na datum prve unesene vožnje.
 free_ride_alerts=c.execute("""
     SELECT id, client, date_from, relation, vehicle, kind, payment_status
     FROM free_rides
     WHERE TRIM(COALESCE(date_from,'')) != ''
       AND date(date_from) >= date('now','localtime')
       AND date(date_from) <= date('now','localtime','+7 days')
       AND COALESCE(kind,'reserved') = 'reserved'
     ORDER BY date(date_from) ASC, id ASC
 """).fetchall()
 alarm_counts={
     "Registracija":sum(1 for x in expiry_alerts if x["type"]=="Registracija"),
     "PP aparati":sum(1 for x in expiry_alerts if x["type"]=="PP aparati"),
     "Periodični":sum(1 for x in expiry_alerts if x["type"]=="Periodični"),
     "Tahograf":sum(1 for x in expiry_alerts if x["type"]=="Tahograf"),
     "Dozvole":len(permit_expiry_alerts),
     "Slobodne vožnje":len(free_ride_alerts)
 }
 recent_activity=[]
 for r in c.execute("""SELECT created_at,username,full_name,action,details
                      FROM audit_log ORDER BY id DESC LIMIT 8""").fetchall():
     action=(r["action"] or "Aktivnost")
     low=action.lower()
     icon_type="user"
     if "vozilo" in low: icon_type="vehicle"
     elif "voza" in low: icon_type="driver"
     elif "linij" in low: icon_type="line"
     elif "raspored" in low: icon_type="schedule"
     elif "gum" in low or "mont" in low or "demont" in low: icon_type="tire"
     elif "korisnik" in low or "prijav" in low or "odjav" in low: icon_type="user"
     recent_activity.append({
         "icon":icon_type,
         "title":action,
         "name":r["full_name"] or r["username"] or "",
         "detail":r["details"] or "",
         "created_at":r["created_at"]
     })
 tire_stock=c.execute("SELECT COUNT(*) FROM tires WHERE status='Skladište'").fetchone()[0]
 tire_mounted=c.execute("SELECT COUNT(*) FROM tires WHERE status='Montirana'").fetchone()[0]
 tire_purchased=c.execute("SELECT COALESCE(SUM(quantity),0) FROM tire_purchases").fetchone()[0]
 tire_demounted=c.execute("SELECT COUNT(*) FROM tire_events WHERE event_type='Demontaža'").fetchone()[0]
 c.close()
 return render_template("index.html",rows=rows,selected_date=selected.isoformat(),
                        tomorrow_date=next_day.isoformat(),drivers=counts[0],
                        vehicles=counts[1],lines=counts[2],expiry_alerts=expiry_alerts,
                        permit_expiry_alerts=permit_expiry_alerts,
                        alarm_counts=alarm_counts,free_ride_alerts=free_ride_alerts,
                        recent_activity=recent_activity,
                        tire_stock=tire_stock,tire_mounted=tire_mounted,
                        tire_purchased=tire_purchased,tire_demounted=tire_demounted)

@permission_required("schedule_view")
@app.route("/statistika/excel")
def driver_statistics_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from flask import send_file
    import io

    date_from=request.args.get("date_from","").strip()
    date_to=request.args.get("date_to","").strip() or date.today().isoformat()
    driver_filter=request.args.get("driver","").strip()

    # Reuse the statistics page logic by collecting summary from the schedule table.
    c=db()
    try:
        try:
            sched=c.execute("SELECT * FROM schedule").fetchall()
        except Exception:
            sched=c.execute("SELECT * FROM schedules").fetchall()
        try:
            lines=c.execute("SELECT * FROM lines").fetchall()
        except Exception:
            lines=[]
    finally:
        c.close()

    line_map={}
    for ln in lines:
        try: line_map[(ln["name"] or "").strip()]=ln
        except Exception: pass

    data={}
    for r in sched:
        keys=r.keys()
        raw_date=next((r[col] for col in ("date","work_date","schedule_date","datum") if col in keys and r[col]), None)
        if not raw_date: continue
        d=str(raw_date)[:10]
        if date_from and d<date_from: continue
        if date_to and d>date_to: continue

        assigned=[]
        for col in ("driver","driver1","driver2","vozac","vozac1","vozac2"):
            if col in keys and r[col]:
                n=str(r[col]).strip()
                if n and n not in assigned: assigned.append(n)

        relation=next((str(r[col]).strip() for col in ("line","line_name","relation","route") if col in keys and r[col]), "")
        ln=line_map.get(relation)

        minutes=0
        for col in ("duration_total_minutes","duration_minutes","duration_min","minutes"):
            if col in keys and r[col] not in (None,""):
                try: minutes=float(r[col]); break
                except Exception: pass
        if not minutes and ln:
            lk=ln.keys()
            try:
                if "duration_hours_int" in lk or "duration_minutes" in lk:
                    minutes=float(ln["duration_hours_int"] or 0)*60+float(ln["duration_minutes"] or 0)
            except Exception: pass
            if not minutes:
                for col in ("duration_hours","duration"):
                    if col in lk and ln[col] not in (None,""):
                        try: minutes=float(ln[col])*60; break
                        except Exception: pass

        for n in assigned:
            if driver_filter and n != driver_filter: continue
            x=data.setdefault(n, {"rides":0,"dates":set(),"minutes":0})
            x["rides"]+=1
            x["dates"].add(d)
            x["minutes"]+=minutes/len(assigned) if assigned else 0

    wb=Workbook()
    ws=wb.active
    ws.title="Statistika vozača"
    ws.append(["GLOBTOUR d.o.o. Međugorje"])
    ws.append(["STATISTIKA VOZAČA"])
    period=f"Period: {date_from or 'početak'} - {date_to or 'danas'}"
    if driver_filter: period += f" | Vozač: {driver_filter}"
    ws.append([period])
    ws.append([])
    ws.append(["Vozač","Broj vožnji","Dani zauzeća","Odvezeni sati"])
    for cell in ws[1]+ws[2]+ws[3]:
        cell.font=Font(bold=True)
    for cell in ws[5]:
        cell.font=Font(bold=True)
    for name,x in sorted(data.items(), key=lambda kv: kv[0].lower()):
        mins=round(x["minutes"])
        ws.append([name,x["rides"],len(x["dates"]),f"{mins//60} h {mins%60} min"])
    for col,width in {"A":30,"B":16,"C":18,"D":20}.items():
        ws.column_dimensions[col].width=width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment=Alignment(vertical="center")
    buf=io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="statistika_vozaca.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@permission_required("schedule_view")
@app.route("/statistika/vozac/excel")
def driver_statistics_detail_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from flask import send_file
    import io

    name=request.args.get("driver","").strip()
    date_from=request.args.get("date_from","").strip() or date.today().replace(day=1).isoformat()
    date_to=request.args.get("date_to","").strip() or date.today().isoformat() or date.today().isoformat()

    if not name:
        return redirect(url_for("driver_statistics", date_from=date_from, date_to=date_to))

    c=db()
    try:
        try:
            rows=c.execute("SELECT * FROM schedule ORDER BY date,id").fetchall()
        except Exception:
            rows=c.execute("SELECT * FROM schedules ORDER BY date,id").fetchall()
        try:
            lines=c.execute("SELECT * FROM lines").fetchall()
        except Exception:
            lines=[]
    finally:
        c.close()

    line_map={}
    for ln in lines:
        try: line_map[(ln["name"] or "").strip()]=ln
        except Exception: pass

    result=[]
    for r in rows:
        keys=r.keys()
        raw_date=next((r[col] for col in ("date","work_date","schedule_date","datum")
                       if col in keys and r[col]), None)
        if not raw_date: continue
        work_date=str(raw_date)[:10]
        if work_date < date_from or work_date > date_to: continue

        assigned=[]
        for col in ("driver","driver1","driver2","vozac","vozac1","vozac2"):
            if col in keys and r[col]:
                n=str(r[col]).strip()
                if n and n not in assigned: assigned.append(n)
        if name not in assigned: continue

        relation=next((str(r[col]).strip() for col in ("line","line_name","relation","route")
                       if col in keys and r[col]), "")
        ln=line_map.get(relation)

        minutes=0
        for col in ("duration_total_minutes","duration_minutes","duration_min","minutes"):
            if col in keys and r[col] not in (None,""):
                try: minutes=float(r[col]); break
                except Exception: pass
        if not minutes and ln:
            lk=ln.keys()
            try:
                if "duration_hours_int" in lk or "duration_minutes" in lk:
                    minutes=float(ln["duration_hours_int"] or 0)*60+float(ln["duration_minutes"] or 0)
            except Exception: pass
            if not minutes:
                for col in ("duration_hours","duration"):
                    if col in lk and ln[col] not in (None,""):
                        try: minutes=float(ln[col])*60; break
                        except Exception: pass

        share=minutes/len(assigned) if assigned else 0
        time_value=next((str(r[col] or "") for col in ("time","departure","departure_time")
                         if col in keys), "")
        vehicle=next((str(r[col] or "") for col in ("vehicle","bus","registration")
                      if col in keys), "")
        mins=round(share)

        result.append((work_date, relation, time_value, vehicle, mins))

    result.sort(key=lambda x:(x[0],x[2],x[1]))

    wb=Workbook()
    ws=wb.active
    ws.title="Detalj vozača"
    ws.append(["GLOBTOUR d.o.o. Međugorje"])
    ws.append(["DETALJNA STATISTIKA VOZAČA"])
    ws.append([f"Vozač: {name}"])
    ws.append([f"Period: {date_from} - {date_to}"])
    ws.append([])
    ws.append(["Datum","Linija / relacija","Vrijeme polaska","Vozilo","Odvezeni sati vozača"])

    for row_no in (1,2,3,4,6):
        for cell in ws[row_no]:
            cell.font=Font(bold=True)

    total=0
    for d,rel,tme,veh,mins in result:
        total+=mins
        ws.append([d,rel,tme,veh,f"{mins//60} h {mins%60} min"])

    ws.append([])
    ws.append(["UKUPNO","","","",f"{total//60} h {total%60} min"])
    for cell in ws[ws.max_row]:
        cell.font=Font(bold=True)

    for col,width in {"A":14,"B":40,"C":18,"D":20,"E":25}.items():
        ws.column_dimensions[col].width=width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment=Alignment(vertical="center")

    buf=io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"detalj_statistika_{name.replace(' ','_')}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@permission_required("schedule_view")
@app.route("/statistika/vozac")
def driver_statistics_detail():
    from datetime import datetime
    name=request.args.get("driver","").strip()
    date_from=request.args.get("date_from","").strip() or date.today().replace(day=1).isoformat()
    date_to=request.args.get("date_to","").strip() or date.today().isoformat() or date.today().isoformat()

    if not name:
        return redirect(url_for("driver_statistics", date_from=date_from, date_to=date_to))

    c=db()
    try:
        try:
            rows=c.execute("SELECT * FROM schedule ORDER BY date,id").fetchall()
        except Exception:
            rows=c.execute("SELECT * FROM schedules ORDER BY date,id").fetchall()

        try:
            lines=c.execute("SELECT * FROM lines").fetchall()
        except Exception:
            lines=[]
    finally:
        c.close()

    line_map={}
    for ln in lines:
        try:
            line_map[(ln["name"] or "").strip()]=ln
        except Exception:
            pass

    result=[]
    for r in rows:
        keys=r.keys()
        raw_date=next((r[col] for col in ("date","work_date","schedule_date","datum")
                       if col in keys and r[col]), None)
        if not raw_date:
            continue
        work_date=str(raw_date)[:10]
        if work_date < date_from or work_date > date_to:
            continue

        assigned=[]
        for col in ("driver","driver1","driver2","vozac","vozac1","vozac2"):
            if col in keys and r[col]:
                n=str(r[col]).strip()
                if n and n not in assigned:
                    assigned.append(n)

        if name not in assigned:
            continue

        relation=next((str(r[col]).strip() for col in ("line","line_name","relation","route")
                       if col in keys and r[col]), "")
        line_row=line_map.get(relation)

        minutes=0
        for col in ("duration_total_minutes","duration_minutes","duration_min","minutes"):
            if col in keys and r[col] not in (None,""):
                try:
                    minutes=float(r[col]); break
                except Exception:
                    pass
        if not minutes and line_row:
            lkeys=line_row.keys()
            if "duration_hours_int" in lkeys or "duration_minutes" in lkeys:
                try:
                    minutes=float(line_row["duration_hours_int"] or 0)*60 + float(line_row["duration_minutes"] or 0)
                except Exception:
                    minutes=0
            if not minutes:
                for col in ("duration_hours","duration"):
                    if col in lkeys and line_row[col] not in (None,""):
                        try:
                            minutes=float(line_row[col])*60; break
                        except Exception:
                            pass

        share=minutes/len(assigned) if assigned else 0
        time_value=next((str(r[col] or "") for col in ("time","departure","departure_time")
                         if col in keys), "")
        vehicle=next((str(r[col] or "") for col in ("vehicle","bus","registration")
                      if col in keys), "")

        result.append({
            "date": work_date,
            "line": relation,
            "time": time_value,
            "vehicle": vehicle,
            "driver_time_text": f"{int(round(share))//60} h {int(round(share))%60} min"
        })

    result.sort(key=lambda x:(x["date"], x["time"], x["line"]))
    return render_template("statistics_driver.html", driver=name, rows=result,
                           date_from=date_from, date_to=date_to)


def ensure_free_ride_items_table():
    c=db()
    c.execute("""CREATE TABLE IF NOT EXISTS free_ride_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        free_ride_id INTEGER NOT NULL,
        item_date TEXT,
        relation TEXT,
        km_total REAL DEFAULT 0,
        km_bih REAL DEFAULT 0,
        km_hr REAL DEFAULT 0,
        km_ino REAL DEFAULT 0,
        price_per_km REAL DEFAULT 0,
        amount REAL DEFAULT 0,
        sort_order INTEGER DEFAULT 0
    )""")
    # Add new kilometre columns when upgrading an existing database
    cols=[r["name"] for r in c.execute("PRAGMA table_info(free_ride_items)").fetchall()]
    for col in ("km_bih","km_hr","km_ino","price_per_km","amount"):
        if col not in cols:
            c.execute(f"ALTER TABLE free_ride_items ADD COLUMN {col} REAL DEFAULT 0")
    c.commit()
    c.close()

ensure_free_ride_items_table()

def save_free_ride_items(free_ride_id, form):
    rels=form.getlist("item_relation[]")
    dates=form.getlist("item_date[]")
    totals=form.getlist("item_km_total[]")
    bihs=form.getlist("item_km_bih[]")
    hrs=form.getlist("item_km_hr[]")
    inos=form.getlist("item_km_ino[]")
    prices=form.getlist("item_price[]")
    amounts=form.getlist("item_amount[]")
    if not rels:
        return
    c=db()
    c.execute("DELETE FROM free_ride_items WHERE free_ride_id=?",(free_ride_id,))
    for i, rel in enumerate(rels):
        rel=(rel or "").strip()
        kbih=_num(bihs[i] if i < len(bihs) else 0)
        khr=_num(hrs[i] if i < len(hrs) else 0)
        kino=_num(inos[i] if i < len(inos) else 0)
        # Ukupna kilometraža je uvijek automatski zbroj BiH + HR + INO.
        # Ne vjerujemo ručno poslanom polju KM ukupno.
        kt=kbih+khr+kino
        price=_num(prices[i] if i < len(prices) else 0)
        # I iznos se uvijek računa iz stvarne kilometraže i cijene po km.
        amount=kt*price
        if not rel and kt == 0 and price == 0 and amount == 0:
            continue
        c.execute("""INSERT INTO free_ride_items
                     (free_ride_id,item_date,relation,km_total,km_bih,km_hr,km_ino,price_per_km,amount,sort_order)
                     VALUES(?,?,?,?,?,?,?,?,?,?)""",
                  (free_ride_id, dates[i] if i < len(dates) else "", rel, kt, kbih, khr, kino, price, amount, i))
    # Glavni datum slobodne vožnje uvijek je datum prve unesene vožnje
    first=c.execute("""SELECT item_date FROM free_ride_items
                     WHERE free_ride_id=?
                       AND item_date IS NOT NULL
                       AND TRIM(item_date)<>''
                     ORDER BY sort_order ASC,id ASC LIMIT 1""",(free_ride_id,)).fetchone()
    if first and first["item_date"]:
        c.execute("UPDATE free_rides SET date_from=? WHERE id=?",(first["item_date"],free_ride_id))
    c.commit(); c.close()

def get_free_ride_items(free_ride_id):
    c=db()
    rows=c.execute("SELECT * FROM free_ride_items WHERE free_ride_id=? ORDER BY sort_order,id",
                   (free_ride_id,)).fetchall()
    c.close()
    return rows

@app.route("/api/klijenti")
def api_customers():
    try:
        c=db()
        rows=c.execute("SELECT id,name FROM customers WHERE COALESCE(active,'Da')!='Ne' ORDER BY name").fetchall()
        c.close()
        return jsonify([{"id":r["id"],"name":r["name"]} for r in rows])
    except Exception:
        return jsonify([])

@app.route("/klijenti")
@permission_required("free_rides_view")
def customers():
    q=request.args.get("q","").strip(); c=db()
    rows=c.execute("SELECT * FROM customers WHERE name LIKE ? OR city LIKE ? OR vat_number LIKE ? ORDER BY name",(f"%{q}%",f"%{q}%",f"%{q}%")).fetchall() if q else c.execute("SELECT * FROM customers ORDER BY name").fetchall()
    c.close(); return render_template("customers.html",rows=rows,q=q)
@app.route("/klijenti/novi",methods=["GET","POST"])
@permission_required("free_rides_edit")
def customer_new():
    if request.method=="POST":
        c=db()
        try:
            c.execute("INSERT INTO customers(name,id_number,vat_number,address,city,country,contact_person,phone,email,note,active) VALUES(?,?,?,?,?,?,?,?,?,?,?)",(request.form["name"],request.form.get("id_number"),request.form.get("vat_number"),request.form.get("address"),request.form.get("city"),request.form.get("country"),request.form.get("contact_person"),request.form.get("phone"),request.form.get("email"),request.form.get("note"),request.form.get("active","Da"))); c.commit(); c.close(); return redirect("/klijenti")
        except Exception: c.close(); flash("Klijent već postoji ili nije ispravno unesen.","error")
    return render_template("customer_form.html",row=None)
@app.route("/klijenti/<int:id>/uredi",methods=["GET","POST"])
@permission_required("free_rides_edit")
def customer_edit(id):
    c=db(); row=c.execute("SELECT * FROM customers WHERE id=?",(id,)).fetchone()
    if not row: c.close(); abort(404)
    if request.method=="POST":
        c.execute("UPDATE customers SET name=?,id_number=?,vat_number=?,address=?,city=?,country=?,contact_person=?,phone=?,email=?,note=?,active=? WHERE id=?",(request.form["name"],request.form.get("id_number"),request.form.get("vat_number"),request.form.get("address"),request.form.get("city"),request.form.get("country"),request.form.get("contact_person"),request.form.get("phone"),request.form.get("email"),request.form.get("note"),request.form.get("active","Da"),id)); c.commit(); c.close(); return redirect("/klijenti")
    c.close(); return render_template("customer_form.html",row=row)
@app.route("/slobodne-voznje")
@permission_required("free_rides_view")
def free_rides():
    kind=request.args.get("kind","reserved")
    upcoming=request.args.get("upcoming","") == "1"
    client=request.args.get("client","").strip()
    date_from=request.args.get("date_from","").strip()
    date_to=request.args.get("date_to","").strip() or date.today().isoformat()
    c=db()
    where=[]; params=[]
    if kind=="paid":
        where.append("payment_status='Plaćeno'")
    elif kind=="unpaid":
        where.append("COALESCE(payment_status,'') != 'Plaćeno'")
    else:
        where.append("kind=?"); params.append(kind)
    if date_from:
        where.append("COALESCE(date_from,'') >= ?"); params.append(date_from)
    if date_to:
        where.append("COALESCE(date_from,'') <= ?"); params.append(date_to)
    if upcoming:
        where.append("date(COALESCE(date_from,'')) >= date('now','localtime')")
        where.append("date(COALESCE(date_from,'')) <= date('now','localtime','+7 days')")
    sql="SELECT * FROM free_rides WHERE "+" AND ".join(where)+(" ORDER BY date_from ASC,id ASC" if upcoming else " ORDER BY date_from DESC,id DESC")
    rows=c.execute(sql,params).fetchall()

    # Pretraga klijenta radi nad stvarnim nazivom spremljenim na slobodnoj vožnji.
    # Normalizacija uklanja razlike u velikim/malim slovima, razmacima i kvačicama
    # (npr. OŠ Kiseljak, Oš Kiseljak ili naziv s dodatnim razmakom).
    if client:
        import unicodedata
        def _norm_client(value):
            value=str(value or '').strip().casefold()
            value=''.join(ch for ch in unicodedata.normalize('NFKD', value)
                          if not unicodedata.combining(ch))
            return ' '.join(value.split())
        wanted=_norm_client(client)
        rows=[r for r in rows if wanted in _norm_client(r["client"])]

    clients=c.execute("SELECT DISTINCT TRIM(client) AS client FROM free_rides WHERE TRIM(COALESCE(client,'')) != '' ORDER BY client COLLATE NOCASE").fetchall()
    try:
        customer_rows=c.execute("SELECT name FROM customers WHERE TRIM(COALESCE(name,'')) != '' ORDER BY name COLLATE NOCASE").fetchall()
    except Exception:
        customer_rows=[]

    # U padajućem izborniku uvijek prikaži i klijente koji stvarno postoje na vožnjama.
    _client_names=[]
    for rr in list(customer_rows)+list(clients):
        try: name=(rr["name"] if "name" in rr.keys() else rr["client"]).strip()
        except Exception: name=""
        if name and name.casefold() not in [x.casefold() for x in _client_names]:
            _client_names.append(name)
    customer_rows=[{"name":x} for x in sorted(_client_names,key=lambda x:x.casefold())]

    rows=apply_free_ride_first_dates(c, rows)
    rows=enrich_free_ride_rows(c, rows)
    c.close()
    return render_template("free_rides.html",rows=rows,kind=kind,client_filter=client,date_from_filter=date_from,date_to_filter=date_to,clients=clients,customer_rows=customer_rows,upcoming=upcoming)

@app.route("/slobodne-voznje/export")
@permission_required("free_rides_view")
def free_rides_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    kind=request.args.get("kind","reserved")
    client=request.args.get("client","").strip()
    date_from=request.args.get("date_from","").strip()
    date_to=request.args.get("date_to","").strip() or date.today().isoformat()
    c=db(); where=[]; params=[]
    if kind=="paid": where.append("payment_status='Plaćeno'")
    elif kind=="unpaid": where.append("COALESCE(payment_status,'') != 'Plaćeno'")
    else: where.append("kind=?"); params.append(kind)
    if date_from: where.append("COALESCE(date_from,'') >= ?"); params.append(date_from)
    if date_to: where.append("COALESCE(date_from,'') <= ?"); params.append(date_to)
    rows=c.execute("SELECT * FROM free_rides WHERE "+" AND ".join(where)+" ORDER BY date_from DESC,id DESC",params).fetchall()
    if client:
        import unicodedata
        def _norm_export_client(value):
            value=str(value or '').strip().casefold()
            value=''.join(ch for ch in unicodedata.normalize('NFKD', value)
                          if not unicodedata.combining(ch))
            return ' '.join(value.split())
        wanted=_norm_export_client(client)
        rows=[r for r in rows if wanted in _norm_export_client(r["client"])]
    c.close()
    wb=Workbook(); ws=wb.active
    titles={"reserved":"REZERVIRANE VOŽNJE","realized":"REALIZIRANE VOŽNJE","paid":"PLAĆENE VOŽNJE","unpaid":"NEPLAĆENE VOŽNJE"}
    ws.title="Slobodne vožnje"; ws.merge_cells("A1:J1"); ws["A1"]=titles.get(kind,"SLOBODNE VOŽNJE")
    ws["A1"].font=Font(bold=True,size=14); ws["A1"].alignment=Alignment(horizontal="center")
    headers=["Klijent","Relacija","Datum","Završetak","Vozilo","Vozač 1","Vozač 2","Status","Dokument","Iznos (€)"]
    ws.append(headers)
    for cell in ws[2]:
        cell.font=Font(bold=True)
        cell.alignment=Alignment(horizontal="center")
    for r in rows:
        ws.append([r["client"],r["relation"],r["date_from"],r["date_to"],r["vehicle"],r["driver1"],r["driver2"],r["status"] or r["payment_status"],r["document_no"],r["amount"]])
    widths=[28,32,14,14,16,24,24,16,18,14]
    for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width=w
    for row in ws.iter_rows(): 
        for cell in row: cell.alignment=Alignment(vertical="center")
    _sig_row = (total_row + 5) if "total_row" in locals() else (ws.max_row + 5)
    ws.cell(_sig_row,1,"Preuzeo:")
    ws.cell(_sig_row,3,f"Izdao: {_stored_document_issuer(r)}")
    ws.merge_cells(start_row=_sig_row+2,start_column=1,end_row=_sig_row+2,end_column=2)
    ws.merge_cells(start_row=_sig_row+2,start_column=3,end_row=_sig_row+2,end_column=4)
    ws.cell(_sig_row+2,1,"______________________________")
    ws.cell(_sig_row+2,3,"______________________________")
    ws.cell(_sig_row,3).alignment=Alignment(horizontal="right")
    ws.cell(_sig_row+2,3).alignment=Alignment(horizontal="right")
    # Uključi potpise u područje ispisa
    try:
        ws.print_area=f"A1:{get_column_letter(ws.max_column)}{_sig_row+2}"
    except Exception:
        pass
    bio=BytesIO(); wb.save(bio); bio.seek(0)
    filename="slobodne_voznje_"+kind+".xlsx"
    return send_file(bio,as_attachment=True,download_name=filename,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/c-fakture")
@permission_required("free_rides_view")
def c_invoices():
    client_filter=request.args.get("client","").strip()
    date_from=request.args.get("date_from","").strip()
    date_to=request.args.get("date_to","").strip() or date.today().isoformat()
    c=db()

    # Filter klijenata se puni isključivo iz stvarno izdanih C-faktura,
    # ne iz zasebnog modula "Klijenti".
    clients=c.execute("""
        SELECT client FROM (
            SELECT DISTINCT TRIM(client) AS client
            FROM free_rides
            WHERE LOWER(COALESCE(payment_method,'')) IN ('cash','gotovina','gotovina / cash')
              AND TRIM(COALESCE(client,'')) <> ''
            UNION
            SELECT DISTINCT TRIM(client) AS client
            FROM manual_c_invoices
            WHERE TRIM(COALESCE(client,'')) <> ''
        )
        ORDER BY client COLLATE NOCASE
    """).fetchall()

    ride_sql="""SELECT * FROM free_rides
                WHERE LOWER(COALESCE(payment_method,'')) IN ('cash','gotovina','gotovina / cash')
                  AND (?='' OR TRIM(COALESCE(client,''))=?)
                  AND (?='' OR COALESCE(issue_date,date_from,'')>=?)
                  AND (?='' OR COALESCE(issue_date,date_from,'')<=?)
                ORDER BY COALESCE(issue_date,date_from,'') DESC,id DESC"""
    rows=c.execute(ride_sql,(client_filter,client_filter,date_from,date_from,date_to,date_to)).fetchall()

    manual_sql="""SELECT * FROM manual_c_invoices
                  WHERE (?='' OR TRIM(COALESCE(client,''))=?)
                    AND (?='' OR COALESCE(issue_date,'')>=?)
                    AND (?='' OR COALESCE(issue_date,'')<=?)
                  ORDER BY COALESCE(issue_date,'') DESC,id DESC"""
    manual_rows=c.execute(manual_sql,(client_filter,client_filter,date_from,date_from,date_to,date_to)).fetchall()

    # C-fakture iz slobodnih vožnji: older documents may not have issued_by
    # stored, so provide the same safe fallback used by PDF/Excel documents.
    ride_rows=[]
    for r in rows:
        d=dict(r)
        if not str(d.get("issued_by") or "").strip():
            d["issued_by"]=_stored_document_issuer(r)
        ride_rows.append(d)
    rows=ride_rows
    c.close()

    return render_template(
        "c_invoices.html",
        rows=rows,
        manual_rows=manual_rows,
        clients=clients,
        client_filter=client_filter,
        date_from=date_from,
        date_to=date_to
    )


@app.route("/c-fakture/<int:id>")
@permission_required("free_rides_view")
def c_invoice_view(id):
    c=db()
    row=c.execute("SELECT * FROM free_rides WHERE id=?",(id,)).fetchone()
    if not row:
        c.close(); abort(404)
    if _payment_method(row)!="cash":
        c.close(); abort(404)
    changed=False
    if not str(row["document_no"] or "").strip():
        new_no=_next_document_number(c, "c_invoice")
        c.execute("UPDATE free_rides SET document_no=? WHERE id=?",(new_no,id)); changed=True
    if not str(row["issue_date"] or "").strip():
        c.execute("UPDATE free_rides SET issue_date=? WHERE id=?",(datetime.now().strftime("%Y-%m-%d"),id)); changed=True
    if changed:
        c.commit(); row=c.execute("SELECT * FROM free_rides WHERE id=?",(id,)).fetchone()
    items=c.execute("SELECT * FROM free_ride_items WHERE free_ride_id=? ORDER BY sort_order,id",(id,)).fetchall()
    c.close()
    # Bez PDV-a: ukupno je uvijek zbroj (KM × cijena po km) svih pojedinačnih vožnji.
    total=sum(_num(i["km_total"])*_num(i["price_per_km"]) for i in items)
    if not items:
        total=_num(row["km_total"])*_num(row["price_per_km"])
    return render_template("c_invoice_view.html",row=row,items=items,total=total,locked=_free_ride_paid(row))


def _c_invoice_data(id):
    c=db()
    row=c.execute("SELECT * FROM free_rides WHERE id=?",(id,)).fetchone()
    if not row:
        c.close()
        return None, [], None, 0.0
    if _payment_method(row)!="cash":
        c.close()
        return None, [], None, 0.0
    changed=False
    if not str(row["document_no"] or "").strip():
        new_no=_next_document_number(c, "c_invoice")
        c.execute("UPDATE free_rides SET document_no=? WHERE id=?",(new_no,id)); changed=True
    if not str(row["issue_date"] or "").strip():
        c.execute("UPDATE free_rides SET issue_date=? WHERE id=?",(datetime.now().strftime("%Y-%m-%d"),id)); changed=True
    if changed:
        c.commit(); row=c.execute("SELECT * FROM free_rides WHERE id=?",(id,)).fetchone()
    items=c.execute("SELECT * FROM free_ride_items WHERE free_ride_id=? ORDER BY sort_order,id",(id,)).fetchall()
    vehicle=c.execute("SELECT * FROM vehicles WHERE registration=?",(row["vehicle"],)).fetchone() if row["vehicle"] else None
    c.close()

    # C-faktura uvijek računa iz kilometara i cijene po km, bez PDV-a.
    out=[]
    total=0.0
    # Neke starije baze mogu imati redove bez novijih kolona.
    # Zato vrijednosti čitamo sigurno i koristimo glavnu cijenu vožnje kao rezervu.
    row_data=dict(row)
    row_price=_num(row_data.get("price_per_km", 0))
    for it in items:
        item_data=dict(it)
        km=_num(item_data.get("km_total", 0))
        item_price=_num(item_data.get("price_per_km", 0))
        price=item_price or row_price
        amount=km*price
        item_data["calc_amount"]=amount
        item_data["calc_price"]=price
        out.append(item_data)
        total += amount

    if not out:
        km=_num(row_data.get("km_total", 0))
        price=_num(row_data.get("price_per_km", 0))
        out=[{"item_date":row_data.get("date_from",""),"relation":row_data.get("relation",""),"km_total":km,
              "calc_price":price,"calc_amount":km*price}]
        total=km*price

    return row,out,vehicle,total


@app.route("/c-fakture/<int:id>/excel")
@permission_required("free_rides_view")
def c_invoice_excel(id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    row,items,vehicle,total=_c_invoice_data(id)
    if not row: abort(404)

    wb=Workbook()
    ws=wb.active
    ws.title="C-faktura"
    ws.sheet_view.showGridLines=False
    widths=[56,16,18,14]
    for n,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(n)].width=w

    dark_fill=PatternFill("solid", fgColor="1F4E78")
    light_fill=PatternFill("solid", fgColor="D9EAF7")
    total_fill=PatternFill("solid", fgColor="FFF2CC")
    thin=Side(style="thin", color="808080")
    medium=Side(style="medium", color="1F1F1F")
    border_all=Border(left=thin,right=thin,top=thin,bottom=thin)

    ws.merge_cells("A2:D2")
    c=ws["A2"]; c.value="GLOBTOUR MEĐUGORJE"; c.font=Font(bold=True,size=18,color="FFFFFF")
    c.fill=dark_fill; c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[2].height=30

    ws.merge_cells("A3:D3")
    c=ws["A3"]; c.value="d.o.o. za promet i turizam"; c.font=Font(italic=True,size=10)
    c.alignment=Alignment(horizontal="center")

    ws.merge_cells("A5:B5")
    c=ws["A5"]; c.value=f"KLIJENT: {row['client'] or ''}"; c.font=Font(bold=True,size=11)
    c.fill=light_fill; c.alignment=Alignment(vertical="center")

    ws.merge_cells("C5:D5")
    c=ws["C5"]; c.value=f"C-FAKTURA / RAČUN: {row['document_no'] or ''}"
    c.font=Font(bold=True,size=11); c.fill=light_fill
    c.alignment=Alignment(horizontal="right",vertical="center")
    for cc in range(1,5):
        ws.cell(5,cc).border=border_all
    ws.row_dimensions[5].height=24
    ws.merge_cells("A6:B6")
    ws["A6"]=f"DATUM IZDAVANJA: {row['issue_date'] or ''}"; ws["A6"].font=Font(bold=True); ws["A6"].fill=light_fill
    ws.merge_cells("C6:D6")
    ws["C6"]=f"VALUTA: {row['currency'] or 'BAM'}"; ws["C6"].font=Font(bold=True); ws["C6"].fill=light_fill; ws["C6"].alignment=Alignment(horizontal="right")
    for cc in range(1,5): ws.cell(6,cc).border=border_all

    headers=["OPIS USLUGE","BR. MJESTA","IZNOS","VALUTA"]
    for col,h in enumerate(headers,1):
        cell=ws.cell(7,col,h)
        cell.font=Font(bold=True,color="FFFFFF")
        cell.fill=dark_fill
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=border_all
    ws.row_dimensions[7].height=28

    seats=(vehicle["seats"] if vehicle and vehicle["seats"] not in (None,"") else "")
    currency=row["currency"] or "BAM"
    rno=8
    for it in items:
        date=str(it.get("item_date") or "")
        relation=it.get("relation") or ""
        values=[f"{date} – {relation}", seats, float(it["calc_amount"]), currency]
        for col,val in enumerate(values,1):
            cell=ws.cell(rno,col,val)
            cell.border=border_all
            cell.alignment=Alignment(horizontal="left" if col==1 else "center",vertical="center",wrap_text=True)
        ws.cell(rno,3).alignment=Alignment(horizontal="right",vertical="center")
        ws.cell(rno,3).number_format='#,##0.00'
        ws.row_dimensions[rno].height=28
        rno+=1

    # Keep the total directly under the last ride.
    total_row=rno
    for col in range(1,5):
        ws.cell(total_row,col).fill=total_fill
        ws.cell(total_row,col).border=Border(
            left=thin,right=thin,
            top=medium,bottom=medium
        )
    ws.cell(total_row,2,"UKUPNO:")
    ws.cell(total_row,2).font=Font(bold=True,size=11)
    ws.cell(total_row,2).alignment=Alignment(horizontal="right")
    ws.cell(total_row,3,float(total))
    ws.cell(total_row,3).font=Font(bold=True,size=11)
    ws.cell(total_row,3).number_format='#,##0.00'
    ws.cell(total_row,3).alignment=Alignment(horizontal="right")
    ws.cell(total_row,4,currency)
    ws.cell(total_row,4).font=Font(bold=True,size=11)
    ws.cell(total_row,4).alignment=Alignment(horizontal="center")
    ws.row_dimensions[total_row].height=26

    ws.merge_cells(start_row=total_row+3,start_column=1,end_row=total_row+3,end_column=4)
    c=ws.cell(total_row+3,1,".")
    c.font=Font(italic=True,size=9)
    c.alignment=Alignment(horizontal="left")

    ws.freeze_panes="A8"
    ws.page_setup.orientation="portrait"
    ws.page_setup.paperSize=ws.PAPERSIZE_A4
    ws.page_margins.left=0.4; ws.page_margins.right=0.4
    ws.page_margins.top=0.5; ws.page_margins.bottom=0.5
    ws.sheet_properties.pageSetUpPr.fitToPage=True
    ws.page_setup.fitToWidth=1
    ws.page_setup.fitToHeight=0
    ws.print_area=f"A1:D{total_row+3}"

    _sig_row = (total_row + 5) if "total_row" in locals() else (ws.max_row + 5)
    ws.cell(_sig_row,1,"Preuzeo:")
    ws.cell(_sig_row,3,f"Izdao: {_stored_document_issuer(row)}")
    ws.merge_cells(start_row=_sig_row+2,start_column=1,end_row=_sig_row+2,end_column=2)
    ws.merge_cells(start_row=_sig_row+2,start_column=3,end_row=_sig_row+2,end_column=4)
    ws.cell(_sig_row+2,1,"______________________________")
    ws.cell(_sig_row+2,3,"______________________________")
    ws.cell(_sig_row,3).alignment=Alignment(horizontal="right")
    ws.cell(_sig_row+2,3).alignment=Alignment(horizontal="right")
    # Uključi potpise u područje ispisa
    try:
        ws.print_area=f"A1:{get_column_letter(ws.max_column)}{_sig_row+2}"
    except Exception:
        pass
    bio=BytesIO(); wb.save(bio); bio.seek(0)
    no=re.sub(r"[^A-Za-z0-9_-]+","_",str(row["document_no"] or id))
    return send_file(bio,as_attachment=True,download_name=f"C-faktura_{no}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/c-fakture/<int:id>/pdf")
@permission_required("free_rides_view")
def c_invoice_pdf(id):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from io import BytesIO
    except ModuleNotFoundError:
        flash("PDF izvoz zahtijeva biblioteku reportlab.","error")
        return redirect(url_for("c_invoice_view",id=id))

    row,items,vehicle,total=_c_invoice_data(id)
    if not row: abort(404)

    # DejaVu Sans is included with the application so Croatian/Bosnian letters
    # (č, ć, ž, š, đ) render correctly on every computer.
    font_dir=os.path.join(os.path.dirname(__file__),"static","fonts")
    regular_font=os.path.join(font_dir,"DejaVuSans.ttf")
    bold_font=os.path.join(font_dir,"DejaVuSans-Bold.ttf")
    try:
        pdfmetrics.registerFont(TTFont("GlobtourDejaVu",regular_font))
        pdfmetrics.registerFont(TTFont("GlobtourDejaVu-Bold",bold_font))
        base_font="GlobtourDejaVu"
        base_bold="GlobtourDejaVu-Bold"
    except Exception:
        base_font="Helvetica"
        base_bold="Helvetica-Bold"

    bio=BytesIO()
    doc=SimpleDocTemplate(
        bio,pagesize=A4,leftMargin=15*mm,rightMargin=15*mm,
        topMargin=15*mm,bottomMargin=15*mm
    )
    styles=getSampleStyleSheet()
    normal=ParagraphStyle("cnormal",parent=styles["Normal"],fontName=base_font,fontSize=9,leading=13)
    small=ParagraphStyle("csmall",parent=normal,fontSize=8.5,leading=11)
    title=ParagraphStyle("ctitle",parent=styles["Title"],fontName=base_bold,fontSize=17,leading=22,alignment=TA_CENTER)
    subtitle=ParagraphStyle("csubtitle",parent=normal,alignment=TA_CENTER,fontSize=9)

    story=[
        Paragraph("GLOBTOUR MEĐUGORJE",title),
        Paragraph("d.o.o. za promet i turizam",subtitle),
        Spacer(1,8*mm)
    ]

    top=Table([[
        Paragraph(f"<b>KLIJENT:</b><br/>{row['client'] or ''}",normal),
        Paragraph(f"<b>C-FAKTURA / RAČUN:</b><br/>{row['document_no'] or ''}<br/><b>DATUM IZDAVANJA:</b> {row['issue_date'] or ''}",
                  ParagraphStyle("right",parent=normal,alignment=TA_RIGHT))
    ]],colWidths=[105*mm,75*mm])
    top.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#D9EAF7")),
        ("BOX",(0,0),(-1,-1),0.7,colors.HexColor("#808080")),
        ("INNERGRID",(0,0),(-1,-1),0.4,colors.HexColor("#B0B0B0")),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("LEFTPADDING",(0,0),(-1,-1),8),
        ("RIGHTPADDING",(0,0),(-1,-1),8),
        ("TOPPADDING",(0,0),(-1,-1),7),
        ("BOTTOMPADDING",(0,0),(-1,-1),7)
    ]))
    story += [top,Spacer(1,6*mm)]

    seats=(vehicle["seats"] if vehicle and vehicle["seats"] not in (None,"") else "")
    cur=row["currency"] or "BAM"
    data=[[
        Paragraph("<b>OPIS USLUGE</b>",small),
        Paragraph("<b>BR. MJESTA</b>",small),
        Paragraph("<b>IZNOS</b>",small),
        Paragraph("<b>VALUTA</b>",small)
    ]]
    for it in items:
        desc=f"{it.get('item_date') or ''} – {it.get('relation') or ''}"
        data.append([
            Paragraph(desc,small),
            str(seats),
            f"{it['calc_amount']:,.2f}".replace(","," "),
            cur
        ])
    data.append(["","UKUPNO:",f"{total:,.2f}".replace(","," "),cur])

    table=Table(data,colWidths=[115*mm,22*mm,25*mm,18*mm],repeatRows=1)
    last=len(data)-1
    table.setStyle(TableStyle([
        ("GRID",(0,0),(-1,-1),0.45,colors.HexColor("#808080")),
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E78")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),base_bold),
        ("FONTNAME",(1,last),(-1,last),base_bold),
        ("BACKGROUND",(0,last),(-1,last),colors.HexColor("#FFF2CC")),
        ("LINEABOVE",(0,last),(-1,last),1.0,colors.black),
        ("LINEBELOW",(0,last),(-1,last),1.0,colors.black),
        ("ALIGN",(1,1),(-1,-1),"CENTER"),
        ("ALIGN",(2,1),(2,-1),"RIGHT"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),7),
        ("BOTTOMPADDING",(0,0),(-1,-1),7),
        ("LEFTPADDING",(0,0),(-1,-1),6),
        ("RIGHTPADDING",(0,0),(-1,-1),6),
    ]))
    story += [table,Spacer(1,8*mm),Paragraph(".",small)]
    story.extend(_document_signature_footer(styles, _stored_document_issuer(row)))
    doc.build(story)
    bio.seek(0)
    no=re.sub(r"[^A-Za-z0-9_-]+","_",str(row["document_no"] or id))
    return send_file(bio,as_attachment=True,download_name=f"C-faktura_{no}.pdf",mimetype="application/pdf")


@app.route("/c-fakture/<int:id>/uredi",methods=["GET","POST"])
@permission_required("free_rides_edit")
def c_invoice_edit(id):
    c=db()
    row=c.execute("SELECT * FROM free_rides WHERE id=?",(id,)).fetchone()
    if not row:
        c.close(); abort(404)
    if _payment_method(row)!="cash":
        c.close(); abort(404)
    if _free_ride_paid(row):
        c.close()
        flash("C-faktura je zaključana jer je vožnja označena kao plaćena.","error")
        return redirect(url_for("c_invoice_view",id=id))
    if request.method=="POST":
        amount=_num(request.form.get("amount"))
        c.execute("""UPDATE free_rides SET client=?,relation=?,date_from=?,date_to=?,vehicle=?,driver1=?,driver2=?,
                     document_no=?,issue_date=?,amount=?,notes=?,currency=?,payment_method='cash' WHERE id=?""",
                  (request.form.get("client",""),request.form.get("relation",""),request.form.get("date_from") or None,
                   request.form.get("date_to") or None,request.form.get("vehicle",""),request.form.get("driver1",""),
                   request.form.get("driver2",""),request.form.get("document_no",""),request.form.get("issue_date") or None,amount,
                   request.form.get("notes",""),request.form.get("currency","BAM"),id))
        c.commit(); c.close()
        save_free_ride_items(id,request.form)
        flash("Sve izmjene C-fakture su spremljene.","success")
        return redirect(url_for("c_invoice_view",id=id))
    items=c.execute("SELECT * FROM free_ride_items WHERE free_ride_id=? ORDER BY sort_order,id",(id,)).fetchall()
    drivers=c.execute("SELECT name FROM drivers WHERE COALESCE(active,'Da')!='Ne' ORDER BY name").fetchall()
    vehicles=c.execute("SELECT registration,seats FROM vehicles ORDER BY registration").fetchall()
    c.close()
    return render_template("c_invoice_edit.html",row=row,items=items,drivers=drivers,vehicles=vehicles)



def _manual_c_data(id):
    c=db(); row=c.execute("SELECT * FROM manual_c_invoices WHERE id=?",(id,)).fetchone()
    if not row: c.close(); return None, [], 0.0
    items=c.execute("SELECT * FROM manual_c_invoice_items WHERE invoice_id=? ORDER BY sort_order,id",(id,)).fetchall(); c.close()
    return row,items,sum(_num(x["amount"]) for x in items)

@app.route("/c-fakture/obicna/novi",methods=["GET","POST"])
@permission_required("free_rides_edit")
def manual_c_invoice_new():
    c=db()
    if request.method=="POST":
        issuer_name=_document_issuer_name()
        # Broj obične C-fakture dodjeljuje se automatski.
        no=_next_document_number(c,"c_invoice")
        issue=request.form.get("issue_date") or datetime.now().strftime("%Y-%m-%d")
        try:
            cur=c.execute("INSERT INTO manual_c_invoices(document_no,client,issue_date,currency,note,issued_by,created_at) VALUES(?,?,?,?,?,?,?)",(no,request.form.get("client",""),issue,request.form.get("currency","BAM"),request.form.get("note",""),issuer_name,datetime.now().isoformat(timespec="seconds")))
        except sqlite3.IntegrityError:
            flash("Broj C-fakture već postoji.","error"); c.close(); return render_template("manual_c_invoice_form.html",row=None,items=[],default_no=no,today=issue)
        iid=cur.lastrowid
        for n,(d,a) in enumerate(zip(request.form.getlist("item_description[]"),request.form.getlist("item_amount[]"))):
            if str(d).strip() or str(a).strip(): c.execute("INSERT INTO manual_c_invoice_items(invoice_id,description,amount,sort_order) VALUES(?,?,?,?)",(iid,d,_num(a),n))
        c.commit(); c.close(); flash("C-faktura je spremljena.","success"); return redirect(url_for("manual_c_invoice_view",id=iid))
    no=_next_document_number(c,"c_invoice"); c.close()
    return render_template("manual_c_invoice_form.html",row=None,items=[],default_no=no,today=datetime.now().strftime("%Y-%m-%d"))

@app.route("/c-fakture/obicna/<int:id>")
@permission_required("free_rides_view")
def manual_c_invoice_view(id):
    row,items,total=_manual_c_data(id)
    if not row: abort(404)
    return render_template("manual_c_invoice_view.html",row=row,items=items,total=total)

@app.route("/c-fakture/obicna/<int:id>/plati",methods=["POST"])
@permission_required("free_rides_edit")
def manual_c_invoice_mark_paid(id):
    c=db(); row=c.execute("SELECT * FROM manual_c_invoices WHERE id=?",(id,)).fetchone()
    if not row: c.close(); abort(404)
    if _manual_invoice_paid(row) and not _is_admin_user():
        c.close(); flash("C-faktura je već plaćena i zaključana.","error")
        return redirect(url_for("manual_c_invoice_view",id=id))
    c.execute("UPDATE manual_c_invoices SET payment_status='Plaćeno' WHERE id=?",(id,))
    c.commit(); c.close()
    flash("C-faktura je označena kao plaćena i zaključana.","success")
    return redirect(url_for("manual_c_invoice_view",id=id))

@app.route("/c-fakture/obicna/<int:id>/uredi",methods=["GET","POST"])
@permission_required("free_rides_edit")
def manual_c_invoice_edit(id):
    c=db(); row=c.execute("SELECT * FROM manual_c_invoices WHERE id=?",(id,)).fetchone()
    if not row: c.close(); abort(404)
    if _manual_invoice_paid(row) and not _is_admin_user():
        c.close()
        flash("C-faktura je označena kao plaćena i više se ne može uređivati.","error")
        return redirect(url_for("manual_c_invoice_view",id=id))
    if request.method=="POST":
        no=request.form.get("document_no","").strip() or row["document_no"]
        c.execute("UPDATE manual_c_invoices SET document_no=?,client=?,issue_date=?,currency=?,note=? WHERE id=?",(no,request.form.get("client",""),request.form.get("issue_date") or None,request.form.get("currency","BAM"),request.form.get("note",""),id))
        c.execute("DELETE FROM manual_c_invoice_items WHERE invoice_id=?",(id,))
        for n,(d,a) in enumerate(zip(request.form.getlist("item_description[]"),request.form.getlist("item_amount[]"))):
            if str(d).strip() or str(a).strip(): c.execute("INSERT INTO manual_c_invoice_items(invoice_id,description,amount,sort_order) VALUES(?,?,?,?)",(id,d,_num(a),n))
        c.commit(); c.close(); flash("Izmjene su spremljene.","success"); return redirect(url_for("manual_c_invoice_view",id=id))
    items=c.execute("SELECT * FROM manual_c_invoice_items WHERE invoice_id=? ORDER BY sort_order,id",(id,)).fetchall(); c.close()
    return render_template("manual_c_invoice_form.html",row=row,items=items,default_no=row["document_no"],today=row["issue_date"] or "")

@app.route("/c-fakture/obicna/<int:id>/excel")
@permission_required("free_rides_view")
def manual_c_invoice_excel(id):
    from openpyxl import Workbook
    from openpyxl.styles import Font,Alignment,Border,Side,PatternFill
    from io import BytesIO
    row,items,total=_manual_c_data(id)
    if not row: abort(404)
    wb=Workbook(); ws=wb.active; ws.title="C-faktura"; ws.sheet_view.showGridLines=False
    ws.column_dimensions['A'].width=70; ws.column_dimensions['B'].width=20; ws.column_dimensions['C'].width=14
    dark=PatternFill('solid',fgColor='1F4E78'); light=PatternFill('solid',fgColor='D9EAF7'); yellow=PatternFill('solid',fgColor='FFF2CC'); side=Side(style='thin',color='808080'); border=Border(left=side,right=side,top=side,bottom=side)
    ws.merge_cells('A2:C2'); ws['A2']='GLOBTOUR MEĐUGORJE'; ws['A2'].font=Font(bold=True,size=18,color='FFFFFF'); ws['A2'].fill=dark; ws['A2'].alignment=Alignment(horizontal='center'); ws.row_dimensions[2].height=30
    ws.merge_cells('A3:C3'); ws['A3']='d.o.o. za promet i turizam'; ws['A3'].alignment=Alignment(horizontal='center')
    ws.merge_cells('A5:B5'); ws['A5']=f"KLIJENT: {row['client'] or ''}"; ws['A5'].font=Font(bold=True); ws['A5'].fill=light
    ws['C5']=f"C-FAKTURA: {row['document_no']}"; ws['C5'].font=Font(bold=True); ws['C5'].fill=light
    ws.merge_cells('A6:B6'); ws['A6']=f"DATUM IZDAVANJA: {row['issue_date'] or ''}"; ws['A6'].font=Font(bold=True); ws['C6']=f"VALUTA: {row['currency'] or 'BAM'}"; ws['C6'].font=Font(bold=True)
    for rr in (5,6):
        for cc in range(1,4): ws.cell(rr,cc).border=border
    for col,h in enumerate(['OPIS USLUGE','IZNOS','VALUTA'],1):
        x=ws.cell(8,col,h); x.font=Font(bold=True,color='FFFFFF'); x.fill=dark; x.alignment=Alignment(horizontal='center'); x.border=border
    r=9
    for it in items:
        vals=[it['description'] or '',_num(it['amount']),row['currency'] or 'BAM']
        for col,val in enumerate(vals,1): ws.cell(r,col,val).border=border
        ws.cell(r,2).number_format='#,##0.00'; ws.cell(r,2).alignment=Alignment(horizontal='right'); r+=1
    for col in range(1,4): ws.cell(r,col).fill=yellow; ws.cell(r,col).border=border
    ws.cell(r,1,'UKUPNO:').font=Font(bold=True); ws.cell(r,2,total).font=Font(bold=True); ws.cell(r,2).number_format='#,##0.00'; ws.cell(r,3,row['currency'] or 'BAM').font=Font(bold=True)
    ws.merge_cells(start_row=r+2,start_column=1,end_row=r+2,end_column=3); ws.cell(r+2,1,'.').font=Font(italic=True)
    _sig_row = (total_row + 5) if "total_row" in locals() else (ws.max_row + 5)
    ws.cell(_sig_row,1,"Preuzeo:")
    ws.cell(_sig_row,3,f"Izdao: {_stored_document_issuer(r)}")
    ws.merge_cells(start_row=_sig_row+2,start_column=1,end_row=_sig_row+2,end_column=2)
    ws.merge_cells(start_row=_sig_row+2,start_column=3,end_row=_sig_row+2,end_column=4)
    ws.cell(_sig_row+2,1,"______________________________")
    ws.cell(_sig_row+2,3,"______________________________")
    ws.cell(_sig_row,3).alignment=Alignment(horizontal="right")
    ws.cell(_sig_row+2,3).alignment=Alignment(horizontal="right")
    # Uključi potpise u područje ispisa
    try:
        ws.print_area=f"A1:{get_column_letter(ws.max_column)}{_sig_row+2}"
    except Exception:
        pass
    bio=BytesIO(); wb.save(bio); bio.seek(0); no=re.sub(r'[^A-Za-z0-9_-]+','_',str(row['document_no']))
    return send_file(bio,as_attachment=True,download_name=f'C-faktura_{no}.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route("/c-fakture/obicna/<int:id>/pdf")
@permission_required("free_rides_view")
def manual_c_invoice_pdf(id):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet,ParagraphStyle
        from reportlab.lib.enums import TA_CENTER,TA_RIGHT
        from reportlab.platypus import SimpleDocTemplate,Paragraph,Spacer,Table,TableStyle
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from io import BytesIO
    except ModuleNotFoundError: abort(500)
    row,items,total=_manual_c_data(id)
    if not row: abort(404)
    fd=os.path.join(os.path.dirname(__file__),'static','fonts'); pdfmetrics.registerFont(TTFont('GlobtourDejaVu',os.path.join(fd,'DejaVuSans.ttf'))); pdfmetrics.registerFont(TTFont('GlobtourDejaVu-Bold',os.path.join(fd,'DejaVuSans-Bold.ttf')))
    bio=BytesIO(); doc=SimpleDocTemplate(bio,pagesize=A4,leftMargin=15*mm,rightMargin=15*mm,topMargin=15*mm,bottomMargin=15*mm); styles=getSampleStyleSheet(); normal=ParagraphStyle('n',parent=styles['Normal'],fontName='GlobtourDejaVu',fontSize=9,leading=13); title=ParagraphStyle('t',parent=styles['Title'],fontName='GlobtourDejaVu-Bold',fontSize=17,alignment=TA_CENTER)
    story=[Paragraph('GLOBTOUR MEĐUGORJE',title),Paragraph('d.o.o. za promet i turizam',ParagraphStyle('s',parent=normal,alignment=TA_CENTER)),Spacer(1,8*mm)]
    top=Table([[Paragraph(f"<b>KLIJENT:</b><br/>{row['client'] or ''}",normal),Paragraph(f"<b>C-FAKTURA / RAČUN:</b><br/>{row['document_no'] or ''}<br/><b>DATUM IZDAVANJA:</b> {row['issue_date'] or ''}",ParagraphStyle('r',parent=normal,alignment=TA_RIGHT))]],colWidths=[105*mm,75*mm]); top.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#D9EAF7')),('BOX',(0,0),(-1,-1),.7,colors.grey),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('PADDING',(0,0),(-1,-1),8)])); story += [top,Spacer(1,6*mm)]
    data=[[Paragraph('<b>OPIS USLUGE</b>',normal),Paragraph('<b>IZNOS</b>',normal),Paragraph('<b>VALUTA</b>',normal)]] + [[Paragraph(str(x['description'] or ''),normal),f"{_num(x['amount']):,.2f}".replace(',',' '),row['currency'] or 'BAM'] for x in items] + [['UKUPNO:',f"{total:,.2f}".replace(',',' '),row['currency'] or 'BAM']]
    tb=Table(data,colWidths=[130*mm,30*mm,20*mm],repeatRows=1); last=len(data)-1; tb.setStyle(TableStyle([('GRID',(0,0),(-1,-1),.45,colors.grey),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1F4E78')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'GlobtourDejaVu-Bold'),('FONTNAME',(0,last),(-1,last),'GlobtourDejaVu-Bold'),('BACKGROUND',(0,last),(-1,last),colors.HexColor('#FFF2CC')),('ALIGN',(1,1),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('PADDING',(0,0),(-1,-1),7)])); story += [tb,Spacer(1,8*mm),Paragraph('.',normal)]
    story.extend(_document_signature_footer(styles, _stored_document_issuer(row)))
    doc.build(story); bio.seek(0); no=re.sub(r'[^A-Za-z0-9_-]+','_',str(row['document_no']))
    return send_file(bio,as_attachment=True,download_name=f'C-faktura_{no}.pdf',mimetype='application/pdf')

def _customer_data(client_name):
    if not client_name:
        return None
    c=db()
    try:
        row=c.execute("SELECT * FROM customers WHERE TRIM(name)=TRIM(?)",(client_name,)).fetchone()
    except Exception:
        row=None
    c.close()
    return row

def _customer_context(client_name):
    row=_customer_data(client_name)
    if not row:
        return {"customer":None}
    return {"customer":dict(row)}

@app.route("/predracuni")
@permission_required("free_rides_view")
def proformas():
    client_filter=request.args.get("client","").strip()
    date_from=request.args.get("date_from","").strip()
    date_to=request.args.get("date_to","").strip() or date.today().isoformat()
    c=db()
    clients=c.execute("""
        SELECT DISTINCT TRIM(COALESCE(NULLIF(p.client,''), f.client)) AS client
        FROM proformas p
        LEFT JOIN free_rides f ON f.id=p.free_ride_id
        WHERE (p.free_ride_id IS NULL
               OR LOWER(COALESCE(f.payment_method,'account')) NOT IN ('cash','gotovina','gotovina / cash'))
          AND TRIM(COALESCE(NULLIF(p.client,''), f.client,'')) <> ''
        ORDER BY client
    """).fetchall()
    rows=c.execute("""
        SELECT p.* FROM proformas p
        LEFT JOIN free_rides f ON f.id=p.free_ride_id
        WHERE (p.free_ride_id IS NULL
               OR LOWER(COALESCE(f.payment_method,'account')) NOT IN ('cash','gotovina','gotovina / cash'))
          AND (?='' OR TRIM(COALESCE(NULLIF(p.client,''), f.client))=?)
          AND (?='' OR date(COALESCE(p.issue_date,'')) >= date(?))
          AND (?='' OR date(COALESCE(p.issue_date,'')) <= date(?))
        ORDER BY p.issue_date DESC,p.id DESC
    """,(client_filter,client_filter,date_from,date_from,date_to,date_to)).fetchall()
    c.close()
    return render_template(
        "proformas.html",
        rows=rows,
        clients=clients,
        client_filter=client_filter,
        date_from=date_from,
        date_to=date_to
    )


@app.route("/predracuni/novi/<int:ride_id>",methods=["GET","POST"])
@permission_required("free_rides_edit")
def proforma_new(ride_id):
    c=db()
    ride=c.execute("SELECT * FROM free_rides WHERE id=?",(ride_id,)).fetchone()
    if not ride:
        c.close(); abort(404)
    ride_items=c.execute("SELECT * FROM free_ride_items WHERE free_ride_id=? ORDER BY sort_order,id",(ride_id,)).fetchall()
    if request.method=="POST":
        issuer_name=_document_issuer_name()
        number=request.form.get("number","").strip() or _next_document_number(c, "proforma")
        relation=request.form.get("relation") or (ride_items[0]["relation"] if ride_items else ride["relation"])
        km_total=float(request.form.get("km_total") or 0)
        c.execute("""INSERT INTO proformas(number,free_ride_id,client,relation,issue_date,due_date,amount,vat_text,currency,client_address,client_city,client_id,client_vat_number,km_total,km_bih,km_hr,km_ino,price_per_km,note,status,issued_by)
                     VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                  (number,ride_id,request.form.get("client"),relation,request.form.get("issue_date"),request.form.get("due_date"),
                   float(request.form.get("amount") or 0),request.form.get("vat_text"),request.form.get("currency") or "BAM",
                   request.form.get("client_address"),request.form.get("client_city"),request.form.get("client_id"),request.form.get("client_vat_number"),km_total,
                   0,0,0,float(request.form.get("price_per_km") or 0),
                   ("OBRAČUNAO: " + request.form.get("prepared_by","").strip() + "\n" + (request.form.get("note") or "")),"Otvoren",issuer_name))
        c.commit()
        pid=c.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        c.close()
        _save_items(pid, request.form)
        return redirect(url_for("proforma_view",id=pid))
    default_conn=db()
    default_no=_next_document_number(default_conn, "proforma")
    default_conn.close()
    c.close()
    customer=_customer_data(ride["client"] if ride and "client" in ride.keys() else "")
    return render_template("proforma_form.html", customer=customer,ride=ride,ride_items=ride_items,
                           default_no=default_no,
                           today=datetime.now().strftime("%Y-%m-%d"))

@app.route("/predracuni/<int:id>")
@permission_required("free_rides_view")
def proforma_view(id):
    c=db()
    row=c.execute("SELECT * FROM proformas WHERE id=?",(id,)).fetchone()
    if not row:
        c.close()
        abort(404)
    items=c.execute("SELECT * FROM proforma_items WHERE proforma_id=? ORDER BY sort_order,id",(id,)).fetchall()
    locked=_proforma_locked(c,row)
    c.close()

    calc_total=sum(
        _num(i["amount"]) or (_num(i["km_total"])*_num(i["price_per_km"]))
        for i in items
    )
    if not items:
        calc_total=_num(row["amount"])

    return render_template(
        "proforma_view.html",
        row=row, items=items, calc_total=calc_total,
        locked=locked, vat_text=(row["vat_text"] or "")
    )

@app.route("/predracuni/<int:id>/pdf")
@permission_required("free_rides_view")
def proforma_pdf(id):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from io import BytesIO
    except ModuleNotFoundError:
        flash("PDF izvoz zahtijeva biblioteku reportlab. Pokreni: pip install reportlab", "error")
        return redirect(url_for("proforma_view", id=id))

    c=db()
    r=c.execute("SELECT * FROM proformas WHERE id=?",(id,)).fetchone()
    items=c.execute("SELECT * FROM proforma_items WHERE proforma_id=? ORDER BY sort_order,id",(id,)).fetchall()
    free_items=[]
    if r and r["free_ride_id"]:
        free_items=c.execute("SELECT * FROM free_ride_items WHERE free_ride_id=? ORDER BY sort_order,id",(r["free_ride_id"],)).fetchall()
    c.close()
    if not r: abort(404)

    if not items:
        items=[{"item_date":r["issue_date"],"relation":r["relation"],"km_total":r["km_total"],"price_per_km":r["price_per_km"],"amount":r["amount"]}]

    bio=BytesIO()
    doc=SimpleDocTemplate(bio,pagesize=A4,rightMargin=10*mm,leftMargin=10*mm,topMargin=10*mm,bottomMargin=10*mm)
    font_dir=os.path.join(os.path.dirname(__file__),"static","fonts")
    regular_font=os.path.join(font_dir,"DejaVuSans.ttf")
    bold_font=os.path.join(font_dir,"DejaVuSans-Bold.ttf")
    try:
        pdfmetrics.registerFont(TTFont("GlobtourDejaVu",regular_font))
        pdfmetrics.registerFont(TTFont("GlobtourDejaVu-Bold",bold_font))
        base_font="GlobtourDejaVu"
        base_bold="GlobtourDejaVu-Bold"
    except Exception:
        base_font="Helvetica"
        base_bold="Helvetica-Bold"

    styles=getSampleStyleSheet()
    small=ParagraphStyle("small",parent=styles["Normal"],fontName=base_font,fontSize=7.3,leading=9)
    normal=ParagraphStyle("normal",parent=styles["Normal"],fontName=base_font,fontSize=8,leading=10)
    center=ParagraphStyle("center",parent=normal,alignment=TA_CENTER)
    right=ParagraphStyle("right",parent=normal,alignment=TA_RIGHT)
    title=ParagraphStyle("title",parent=styles["Title"],fontName=base_bold,fontSize=15,leading=18,alignment=TA_CENTER)
    story=[]

    left_header='''<b>globtour    međugorje</b><br/>
d.o.o. za promet i turizam - Touring Company<br/>
BIH - 88 260 Čitluk; www.globtour.com<br/>
Tel / fax: +387 36 / 653 253, 653 251<br/>
E-mail: globtour@globtour.com<br/>
JIB: 4227010020009; PDV: 227010020009<br/>
P.B. 17070363<br/>
Županijski sud Mostar: U/I-701/92'''
    bank_header='''<b>Žiro račun:</b><br/>
154 922 2000 717 581 - Intesa SanPaolo banka d.d. Sarajevo<br/><br/>
<b>Devizni račun / Bank account:</b><br/>
SWIFT: UPBKBA22<br/>
IBAN: BA39 154922 2000 717581'''
    ht=Table([[Paragraph(left_header,small),Paragraph(bank_header,small)]],colWidths=[92*mm,98*mm])
    ht.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    story += [ht, Spacer(1,6*mm)]

    client=f"<b>{r['client'] or ''}</b><br/>{r['client_address'] or ''}<br/>{r['client_city'] or ''}<br/>ID: {r['client_id'] or ''}"
    pr_data=[[Paragraph(client,normal),Paragraph("<b>PREDRAČUN</b>",title),Paragraph(f"Valuta / Currency:<br/><b>{r['currency'] or 'BAM'}</b>",center)],
             ["",Paragraph(f"Datum / Date: {r['issue_date'] or ''}",right),Paragraph(f"Br: {r['number'] or ''}",right)]]
    pt=Table(pr_data,colWidths=[70*mm,70*mm,50*mm])
    pt.setStyle(TableStyle([("BOX",(0,0),(-1,-1),0.6,colors.black),("LINEBELOW",(0,0),(-1,0),0.4,colors.grey),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("SPAN",(0,0),(0,1)),("PADDING",(0,0),(-1,-1),6)]))
    story += [pt, Spacer(1,6*mm)]

    headers=["Datum","Opis / Description","KM UK.","BIH","HR","INO","CIJENA / km","OSNOV PDV BIH","OSNOV PDV HR","PDV BIH","PDV HR","NEOPOREZ.","UKUPNO"]
    data=[headers]
    sums={"ob":0.0,"oh":0.0,"pb":0.0,"ph":0.0,"neo":0.0,"total":0.0}
    cur=r["currency"] or "BAM"
    for n,it in enumerate(items):
        fi=free_items[n] if n < len(free_items) else None
        kb=_num(fi["km_bih"]) if fi else (_num(r["km_bih"]) if len(items)==1 else 0)
        kh=_num(fi["km_hr"]) if fi else (_num(r["km_hr"]) if len(items)==1 else 0)
        ki=_num(fi["km_ino"]) if fi else (_num(r["km_ino"]) if len(items)==1 else 0)
        kt=_num(it["km_total"]) or (kb+kh+ki)
        p=_num(it["price_per_km"]) or _num(r["price_per_km"])
        ob=kb*p; oh=kh*p; pb=ob*0.17; ph=oh*0.25; neo=ki*p
        total=ob+oh+pb+ph+neo
        for k,v in [("ob",ob),("oh",oh),("pb",pb),("ph",ph),("neo",neo),("total",total)]: sums[k]+=v
        data.append([str(it["item_date"] or ""),str(it["relation"] or ""),f"{kt:.0f}",f"{kb:.0f}",f"{kh:.0f}",f"{ki:.0f}",f"{p:.2f} {cur}",f"{ob:.2f}",f"{oh:.2f}",f"{pb:.2f}",f"{ph:.2f}",f"{neo:.2f}",f"{total:.2f}"])
    data.append(["","UKUPNO","","","","","",f"{sums['ob']:.2f}",f"{sums['oh']:.2f}",f"{sums['pb']:.2f}",f"{sums['ph']:.2f}",f"{sums['neo']:.2f}",f"{sums['total']:.2f} {cur}"])
    widths=[16*mm,34*mm,11*mm,8*mm,8*mm,8*mm,16*mm,16*mm,16*mm,12*mm,12*mm,15*mm,18*mm]
    tb=Table(data,colWidths=widths,repeatRows=1)
    tb.setStyle(TableStyle([("GRID",(0,0),(-1,-1),0.35,colors.black),("BACKGROUND",(0,0),(-1,0),colors.HexColor("#EDEDED")),("FONTNAME",(0,0),(-1,0),base_bold),("FONTNAME",(0,-1),(-1,-1),base_bold),("ALIGN",(0,0),(-1,0),"CENTER"),("ALIGN",(2,1),(-1,-1),"RIGHT"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("FONTSIZE",(0,0),(-1,-1),6.2),("LEADING",(0,0),(-1,-1),7.5),("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),("LEFTPADDING",(0,0),(-1,-1),2),("RIGHTPADDING",(0,0),(-1,-1),2)]))
    story += [tb, Spacer(1,6*mm)]

    note_text=r["note"] or ""; prepared_by=""
    if note_text.startswith("OBRAČUNAO: "):
        first,sep,rest=note_text.partition("\n"); prepared_by=first.replace("OBRAČUNAO: ","").strip(); note_text=rest if sep else ""
    if r["vat_text"]: story.append(Paragraph(f"<b>PDV:</b> {r['vat_text']}",normal))
    if note_text: story.append(Paragraph(f"<b>Napomena:</b> {note_text}",normal))
    story += [Spacer(1,8*mm),Paragraph(f"Obračunao: {prepared_by or '____________________'}",normal)]
    story.extend(_document_signature_footer(styles, _stored_document_issuer(r)))
    doc.build(story); bio.seek(0)
    return send_file(bio,as_attachment=True,download_name=f"Predracun_{str(r['number']).replace('/','-')}.pdf",mimetype="application/pdf")

@app.route("/predracuni/<int:id>/excel")
@permission_required("free_rides_view")
def proforma_excel(id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    c=db()
    r=c.execute("SELECT * FROM proformas WHERE id=?",(id,)).fetchone()
    items=c.execute("SELECT * FROM proforma_items WHERE proforma_id=? ORDER BY sort_order,id",(id,)).fetchall()
    free_items=[]
    if r and r["free_ride_id"]:
        free_items=c.execute("SELECT * FROM free_ride_items WHERE free_ride_id=? ORDER BY sort_order,id",(r["free_ride_id"],)).fetchall()
    c.close()
    if not r: abort(404)

    wb=Workbook(); ws=wb.active; ws.title="Predračun"
    widths=[14,34,11,10,10,10,13,14,14,11,11,14,15]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    left_header=[
        (3,"globtour    međugorje"),
        (4,"d.o.o. za promet i turizam - Touring Company"),
        (5,"BIH - 88 260 Čitluk; www. globtour.com"),
        (6,"Tel / fax:  +387 36 / 653 253, 653 251"),
        (7,"E-mail: globtour@globtour.com"),
        (8,"JIB : 4227010020009; PDV: 227010020009"),
        (9,"P.B. 17070363"),
        (10,"Županijski sud Mostar: U/I-701/92"),
    ]
    for row,text in left_header:
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
        ws.cell(row,1,text)
    bank_header=[
        (4,"Žiro račun:"),
        (5,"154 922 2000 717 581 - Intesa SanPaolo banka d.d. Sarajevo"),
        (6,"Devizni račun / Bank account:"),
        (7,"SWIFT: UPBKBA22"),
        (8,"IBAN: BA39 154922 2000 717581"),
    ]
    for row,text in bank_header:
        ws.merge_cells(start_row=row,start_column=3,end_row=row,end_column=13)
        ws.cell(row,3,text)
    ws["A3"].font=Font(bold=True,size=16)
    ws["C4"].font=Font(bold=True); ws["C6"].font=Font(bold=True)
    ws.merge_cells("A13:C16"); ws["A13"]=f"{r['client'] or ''}\n{r['client_address'] or ''}\n{r['client_city'] or ''}\nID: {r['client_id'] or ''}"; ws["A13"].alignment=Alignment(wrap_text=True,vertical="top")
    ws.merge_cells("D13:J15"); ws["D13"]="PREDRAČUN"; ws["D13"].font=Font(bold=True,size=16); ws["D13"].alignment=Alignment(horizontal="center",vertical="center")
    ws["K13"]="Valuta /"; ws["K14"]="Currency:"; ws["L14"]=r["currency"] or "BAM"
    ws["I16"]=f"Datum/Date: {r['issue_date'] or ''}"; ws["K16"]=f"Br: {r['number']}"
    headers=["Datum","Opis / Description","KM UKUPNO","BIH","HR","INO","CIJENA po 1 km","OSNOV PDV - BIH","OSNOV PDV-HR","PDV BIH","PDV HR","NEOPOREZ.","UKUPNO"]
    for i,h in enumerate(headers,1): ws.cell(18,i,h)

    # Fallback for older proformas with no item rows.
    if not items:
        items=[{"item_date":r["issue_date"],"relation":r["relation"],"km_total":r["km_total"],"price_per_km":r["price_per_km"],"amount":r["amount"]}]
    sums={"ob":0.0,"oh":0.0,"pb":0.0,"ph":0.0,"neo":0.0,"total":0.0}
    rowno=20
    for n,it in enumerate(items):
        fi=free_items[n] if n < len(free_items) else None
        kb=float(fi["km_bih"] or 0) if fi else (float(r["km_bih"] or 0) if len(items)==1 else 0)
        kh=float(fi["km_hr"] or 0) if fi else (float(r["km_hr"] or 0) if len(items)==1 else 0)
        ki=float(fi["km_ino"] or 0) if fi else (float(r["km_ino"] or 0) if len(items)==1 else 0)
        kt=float(it["km_total"] or kb+kh+ki)
        p=float(it["price_per_km"] or r["price_per_km"] or 0)
        # PDV: BiH 17%, HR 25%, INO neoporezivi dio.
        ob=kb*p; oh=kh*p; pb=ob*0.17; ph=oh*0.25; neo=ki*p
        total=ob+oh+pb+ph+neo
        vals=[it["item_date"],it["relation"],kt,kb,kh,ki,p,ob,oh,pb,ph,neo,total]
        for col,v in enumerate(vals,1): ws.cell(rowno,col,v)
        for key,val in [("ob",ob),("oh",oh),("pb",pb),("ph",ph),("neo",neo),("total",total)]: sums[key]+=val
        rowno+=1

    total_row=max(31,rowno+1)
    ws.cell(total_row,2,"UKUPNO:")
    ws.cell(total_row,8,sums["ob"]); ws.cell(total_row,9,sums["oh"])
    ws.cell(total_row,10,sums["pb"]); ws.cell(total_row,11,sums["ph"])
    ws.cell(total_row,12,sums["neo"]); ws.cell(total_row,13,sums["total"])
    prepared_by=""
    note_text=r["note"] or ""
    if note_text.startswith("OBRAČUNAO: "):
        prepared_by=note_text.split("\n",1)[0].replace("OBRAČUNAO: ","").strip()
    ws.cell(total_row+1,1,f"Obračunao: {prepared_by or '____________________'}")

    thin=Side(style="thin")
    for rr in ws.iter_rows(min_row=18,max_row=total_row,min_col=1,max_col=13):
        for cell in rr:
            cell.border=Border(left=thin,right=thin,top=thin,bottom=thin)
    for c0 in ws[18]:
        c0.font=Font(bold=True); c0.alignment=Alignment(horizontal="center",wrap_text=True)
    for rr in range(20,rowno):
        for cc in range(3,14):
            ws.cell(rr,cc).number_format='#,##0.00'
    for cc in range(8,14): ws.cell(total_row,cc).number_format='#,##0.00'
    ws.cell(total_row,2).font=Font(bold=True)
    _sig_row = (total_row + 5) if "total_row" in locals() else (ws.max_row + 5)
    ws.cell(_sig_row,1,"Preuzeo:")
    ws.cell(_sig_row,3,f"Izdao: {_stored_document_issuer(r)}")
    ws.merge_cells(start_row=_sig_row+2,start_column=1,end_row=_sig_row+2,end_column=2)
    ws.merge_cells(start_row=_sig_row+2,start_column=3,end_row=_sig_row+2,end_column=4)
    ws.cell(_sig_row+2,1,"______________________________")
    ws.cell(_sig_row+2,3,"______________________________")
    ws.cell(_sig_row,3).alignment=Alignment(horizontal="right")
    ws.cell(_sig_row+2,3).alignment=Alignment(horizontal="right")
    # Uključi potpise u područje ispisa
    try:
        ws.print_area=f"A1:{get_column_letter(ws.max_column)}{_sig_row+2}"
    except Exception:
        pass
    # Ispis: A4 landscape, cijeli predračun na jednu stranicu po širini.
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.18
    ws.page_margins.right = 0.18
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.30
    ws.page_margins.header = 0.10
    ws.page_margins.footer = 0.10
    ws.print_options.horizontalCentered = True
    ws.sheet_properties.outlinePr.summaryBelow = True
    ws.freeze_panes = None
    ws.sheet_properties.pageSetUpPr.autoPageBreaks = False
    ws.print_area = f"A1:{get_column_letter(ws.max_column)}{_sig_row+2}"
    ws.print_title_rows = "1:18"
    bio=BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio,as_attachment=True,download_name=f"Predracun_{r['number'].replace('/','-')}.xlsx",mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def _payment_method(row):
    try:
        value=str(row["payment_method"] or "").strip().lower()
    except Exception:
        value=""
    if value in ("cash","gotovina","gotovina / cash"):
        return "cash"
    return "account"

def _is_admin_user():
    u=auth_user()
    try:
        return u is not None and (str(u["role"] or "").strip().lower()=="admin" or str(u["permissions"] or "").strip()=="*")
    except Exception:
        return False

def _manual_invoice_paid(row):
    try:
        return str(row["payment_status"] or "").strip().lower() in ("plaćeno","placeno","paid")
    except Exception:
        return False

def _free_ride_paid(row):
    try:
        return str(row["payment_status"] or "").strip().lower() in ("plaćeno","placeno","paid")
    except Exception:
        return False

def _proforma_locked(c, proforma_row):
    """A proforma is read-only only when the linked free ride is explicitly paid."""
    if not proforma_row or "free_ride_id" not in proforma_row.keys():
        return False
    free_ride_id = proforma_row["free_ride_id"]
    if not free_ride_id:
        return False

    fr = c.execute("SELECT * FROM free_rides WHERE id=?", (free_ride_id,)).fetchone()
    if not fr:
        return False

    # In this project payment_status is the authoritative payment field.
    if "payment_status" in fr.keys():
        value = str(fr["payment_status"] or "").strip().lower()
        return value in ("plaćeno", "placeno", "paid")

    # Compatibility with older database versions.
    if "paid" in fr.keys():
        return str(fr["paid"]).strip().lower() in ("1", "true", "da", "yes")
    if "is_paid" in fr.keys():
        return str(fr["is_paid"]).strip().lower() in ("1", "true", "da", "yes")
    return False

@app.route("/predracuni/<int:id>/uredi",methods=["GET","POST"])
@permission_required("free_rides_edit")
def proforma_edit(id):
    c=db()
    r=c.execute("SELECT * FROM proformas WHERE id=?",(id,)).fetchone()
    if not r:
        c.close()
        abort(404)

    if _proforma_locked(c, r) and not _is_admin_user():
        c.close()
        flash("Predračun je zaključan jer je vožnja označena kao plaćena.", "error")
        return redirect(url_for("proforma_view", id=id))

    if request.method=="POST":
        # Save ALL editable fields in one database transaction.
        prepared_by=request.form.get("prepared_by","").strip()
        note=request.form.get("note","").strip()
        stored_note=(f"OBRAČUNAO: {prepared_by}\\n{note}" if prepared_by else note)

        km_bih=_num(request.form.get("km_bih"))
        km_hr=_num(request.form.get("km_hr"))
        km_ino=_num(request.form.get("km_ino"))
        # Ukupna kilometraža se ne unosi ručno.
        km_total=km_bih+km_hr+km_ino
        price_per_km=_num(request.form.get("price_per_km"))
        amount=_num(request.form.get("amount"))

        # Spremi stavke i uvijek ponovno izračunaj iznose iz aktualne cijene po km.
        if request.form.getlist("item_relation[]") or request.form.getlist("item_date[]"):
            manual_price=request.form.get("price_per_km","").strip()
            override_price=_num(manual_price) if manual_price != "" else None
            item_total=_save_items(id,request.form,c,price_override=override_price)
            amount=item_total
            # Glavni podaci se također usklađuju sa spremljenim stavkama.
            agg=c.execute("""SELECT COALESCE(SUM(km_total),0) AS km_total,
                                    COALESCE(MAX(price_per_km),0) AS price_per_km
                             FROM proforma_items WHERE proforma_id=?""",(id,)).fetchone()
            if agg:
                km_total=_num(agg["km_total"])
                if agg["price_per_km"]:
                    price_per_km=_num(agg["price_per_km"])

        # Ako nema pojedinačnih stavki, ukupni iznos je kilometraža × cijena/km.
        if not (request.form.getlist("item_relation[]") or request.form.getlist("item_date[]")):
            amount=km_total*price_per_km

        c.execute("""UPDATE proformas SET
            number=?, client=?, relation=?, issue_date=?, due_date=?,
            amount=?, vat_text=?, currency=?, client_address=?, client_city=?,
            client_id=?, client_vat_number=?, km_total=?, km_bih=?, km_hr=?, km_ino=?,
            price_per_km=?, note=?
            WHERE id=?""",
            (request.form.get("number","").strip(),
             request.form.get("client","").strip(),
             request.form.get("relation","").strip(),
             request.form.get("issue_date") or None,
             request.form.get("due_date") or None,
             amount,
             request.form.get("vat_text","").strip(),
             request.form.get("currency","BAM"),
             request.form.get("client_address","").strip(),
             request.form.get("client_city","").strip(),
             request.form.get("client_id","").strip(),
             request.form.get("client_vat_number","").strip(),
             km_total, km_bih, km_hr, km_ino, price_per_km,
             stored_note,id))
        c.commit()
        c.close()
        flash("Sve izmjene predračuna su spremljene.","success")
        return redirect(url_for("proforma_view",id=id))

    items=c.execute("SELECT * FROM proforma_items WHERE proforma_id=? ORDER BY sort_order,id",(id,)).fetchall()
    c.close()

    # Older predračuni without individual stavke still get one editable row.
    if not items:
        items=[{"item_date":r["issue_date"] or "","relation":r["relation"] or "",
                "km_total":r["km_total"] or 0,"price_per_km":r["price_per_km"] or 0,
                "amount":r["amount"] or 0}]

    prepared_by=""
    plain_note=r["note"] or ""
    if plain_note.startswith("OBRAČUNAO: "):
        first,sep,rest=plain_note.partition("\\n")
        prepared_by=first.replace("OBRAČUNAO: ","").strip()
        plain_note=rest if sep else ""

    return render_template("proforma_edit.html", customer=_customer_data(r["client"] if r and "client" in r.keys() else ""),row=r,items=items,
                           prepared_by=prepared_by,plain_note=plain_note)

@app.route("/slobodne-voznje/dodaj",methods=["GET","POST"])
@permission_required("free_rides_edit")
def free_ride_add():
    c=db()
    if request.method=="POST":
        issuer_name=_document_issuer_name()
        kind=request.form.get("kind","reserved")
        payment_method=request.form.get("payment_method","account")
        document_no=(request.form.get("document_no") or "").strip()
        if payment_method=="cash" and not document_no:
            document_no=_next_document_number(c, "c_invoice")
        c.execute("""INSERT INTO free_rides(kind,client,relation,date_from,date_to,passengers,vehicle,driver1,driver2,status,payment_status,document_no,amount,notes,payment_method,issued_by) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",(kind,request.form.get("client"),request.form.get("relation"),request.form.get("date_from") or None,request.form.get("date_to") or None,request.form.get("passengers"),request.form.get("vehicle"),request.form.get("driver1"),request.form.get("driver2"),request.form.get("status"),request.form.get("payment_status"),document_no,float(request.form.get("amount") or 0),request.form.get("notes"),payment_method,issuer_name))
        c.commit()
        free_ride_id=c.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        c.close()
        save_free_ride_items(free_ride_id, request.form)
        flash("Vožnja je spremljena.","success")
        return redirect(url_for("free_rides",kind=kind))
    drivers=c.execute("SELECT name FROM drivers WHERE COALESCE(active,'Da')!='Ne' ORDER BY name").fetchall()
    vehicles=c.execute("SELECT registration, seats FROM vehicles ORDER BY registration").fetchall()
    c.close()
    return render_template("free_ride_form.html",row=None,items=[],drivers=drivers,vehicles=vehicles)

@app.route("/slobodne-voznje/uredi/<int:id>",methods=["GET","POST"])
@permission_required("free_rides_edit")
def free_ride_edit(id):
    c=db(); row=c.execute("SELECT * FROM free_rides WHERE id=?",(id,)).fetchone()
    if not row:
        c.close()
        abort(404)
    payment_method=request.form.get("payment_method","account") if request.method=="POST" else ((row["payment_method"] if "payment_method" in row.keys() else "account") or "account")
    if _free_ride_paid(row) and not _is_admin_user():
        c.close()
        flash("Vožnja je označena kao plaćena i više se ne može uređivati.","error")
        return redirect(url_for("free_rides",kind="paid"))
    if request.method=="POST":
        c.execute("""UPDATE free_rides SET kind=?,client=?,relation=?,date_from=?,date_to=?,passengers=?,vehicle=?,driver1=?,driver2=?,status=?,payment_status=?,document_no=?,amount=?,notes=? WHERE id=?""",
        (request.form.get("kind"),request.form.get("client"),request.form.get("relation"),request.form.get("date_from") or None,request.form.get("date_to") or None,request.form.get("passengers"),request.form.get("vehicle"),request.form.get("driver1"),request.form.get("driver2"),request.form.get("status"),row["payment_status"],request.form.get("document_no"),float(request.form.get("amount") or 0),request.form.get("notes"),id))
        try:
            c.execute("UPDATE free_rides SET payment_method=? WHERE id=?", (payment_method, id))
            if payment_method=="cash":
                current_no=(request.form.get("document_no") or row["document_no"] or "").strip()
                if not current_no:
                    current_no=_next_document_number(c, "c_invoice")
                    c.execute("UPDATE free_rides SET document_no=? WHERE id=?", (current_no,id))
        except Exception:
            pass
        c.commit()
        kind=request.form.get("kind")
        c.close()
        save_free_ride_items(id, request.form)
        return redirect(url_for("free_rides",kind=kind))
    drivers=c.execute("SELECT name FROM drivers ORDER BY name").fetchall(); vehicles=c.execute("SELECT registration, seats FROM vehicles ORDER BY registration").fetchall(); c.close()
    items=get_free_ride_items(id)
    return render_template("free_ride_form.html",row=row,items=items,drivers=drivers,vehicles=vehicles)

@app.route("/slobodne-voznje/obrisi/<int:id>",methods=["POST"])
@permission_required("free_rides_delete")
def free_ride_delete(id):
    c=db(); kind=(c.execute("SELECT kind FROM free_rides WHERE id=?",(id,)).fetchone() or {"kind":"reserved"})["kind"]; c.execute("DELETE FROM free_rides WHERE id=?",(id,)); c.commit(); c.close(); return redirect(url_for("free_rides",kind=kind))

@app.route("/slobodne-voznje/realiziraj/<int:id>",methods=["POST"])
@permission_required("free_rides_edit")
def free_ride_realize(id):
    c=db(); c.execute("UPDATE free_rides SET kind='realized',status='Realizirano' WHERE id=?",(id,)); c.commit(); c.close(); return redirect(url_for("free_rides",kind="realized"))

@app.route("/slobodne-voznje/plati/<int:id>",methods=["POST"])
@permission_required("free_rides_edit")
def free_ride_mark_paid(id):
    c=db()
    c.execute("UPDATE free_rides SET payment_status='Plaćeno' WHERE id=?",(id,))
    c.commit()
    c.close()
    flash("Vožnja je označena kao plaćena i prebačena među plaćene vožnje.","success")
    return redirect(url_for("free_rides",kind="paid"))
@app.route("/statistika/slobodne-voznje")
@permission_required("free_rides_view")
def free_ride_statistics():
    from datetime import datetime, date

    date_from=request.args.get("date_from","").strip()
    date_to=request.args.get("date_to","").strip() or date.today().isoformat()
    driver_filter=request.args.get("driver","").strip()
    if not date_from: date_from=date.today().replace(day=1).isoformat()
    if not date_to: date_to=date.today().isoformat()

    c=db()
    drivers=c.execute("SELECT name FROM drivers WHERE active='Da' ORDER BY name").fetchall()
    # Datum slobodne vožnje sada je datum svake pojedinačne stavke/vožnje.
    rows=c.execute("""
        SELECT fr.id,fr.driver1,fr.driver2,fr.vehicle,fr.client,
               fi.item_date,fi.relation,fi.km_total
        FROM free_rides fr
        JOIN free_ride_items fi ON fi.free_ride_id=fr.id
        WHERE fr.kind='realized'
          AND TRIM(COALESCE(fi.item_date,''))<>''
          AND fi.item_date>=? AND fi.item_date<=?
        ORDER BY fi.item_date,fr.id,fi.sort_order,fi.id
    """,(date_from,date_to)).fetchall()
    c.close()

    stats={}; detail=[]
    for r in rows:
        name_date=r['item_date']
        assigned=[]
        for slot in ('driver1','driver2'):
            name=(r[slot] or '').strip()
            if name and name not in assigned: assigned.append(name)
        for name in assigned:
            if driver_filter and name!=driver_filter: continue
            x=stats.setdefault(name,{'driver':name,'days':set(),'rides':0})
            x['days'].add(name_date)
            x['rides']+=1
            detail.append({
                'driver':name,
                'date_from':name_date,
                'date_to':name_date,
                'relation':r['relation'] or '',
                'vehicle':r['vehicle'] or '',
                'client':r['client'] or '',
                'days_count':1,
                'km_total':float(r['km_total'] or 0)
            })

    summary=[{'driver':x['driver'],'days_count':len(x['days']),'rides_count':x['rides']} for x in stats.values()]
    summary.sort(key=lambda x:(-x['days_count'],x['driver'].lower()))
    detail.sort(key=lambda x:(x['driver'].lower(),x['date_from'],x['relation'].lower()))
    return render_template('free_ride_statistics.html',summary=summary,detail=detail,
                           top_days=summary[0] if summary else None,drivers=drivers,
                           date_from=date_from,date_to=date_to,driver_filter=driver_filter)

@app.route("/statistika")
@permission_required("schedule_view")
def driver_statistics():
    from datetime import datetime, timedelta

    date_from=request.args.get("date_from","").strip()
    date_to=request.args.get("date_to","").strip() or date.today().isoformat()
    driver_filter=request.args.get("driver","").strip()

    if not date_from:
        date_from=date.today().replace(day=1).isoformat()
    if not date_to:
        date_to=date.today().isoformat()

    c=db()

    # Drivers are loaded without relying on an active flag format.
    try:
        drivers=c.execute("SELECT * FROM drivers ORDER BY name").fetchall()
    except Exception:
        drivers=[]

    # Read all schedule rows and apply the date range in Python so older/newer DB schemas work.
    try:
        schedule_rows=c.execute("SELECT * FROM schedule ORDER BY date,id").fetchall()
    except Exception:
        try:
            schedule_rows=c.execute("SELECT * FROM schedules ORDER BY date,id").fetchall()
        except Exception:
            schedule_rows=[]

    # Lines for duration/return calculations.
    try:
        lines=c.execute("SELECT * FROM lines").fetchall()
    except Exception:
        lines=[]

    line_map={}
    for ln in lines:
        try:
            line_map[str(ln["name"]).strip()]=ln
        except Exception:
            pass

    # Return/occupancy records if the project stores them separately.
    try:
        returns=c.execute("SELECT * FROM schedule_returns").fetchall()
    except Exception:
        returns=[]

    c.close()

    try:
        f_from=datetime.strptime(date_from,"%Y-%m-%d").date()
        f_to=datetime.strptime(date_to,"%Y-%m-%d").date()
    except Exception:
        f_from=date.today().replace(day=1)
        f_to=date.today()

    stats={}
    details={}

    def add_driver(name, occupied_days, hours, relation, work_date):
        name=(name or "").strip()
        if not name or name.lower() in ("none","-","n/a"):
            return
        if driver_filter and name!=driver_filter:
            return
        if name not in stats:
            stats[name]={"driver":name,"rides":0,"days":set(),"hours":0}
            details[name]=[]
        stats[name]["rides"]+=1
        stats[name]["days"].update(occupied_days)
        stats[name]["hours"]+=hours
        details[name].append({
            "date":work_date.isoformat() if hasattr(work_date,"isoformat") else str(work_date),
            "relation":relation or "",
            "days":len(occupied_days),
            "hours":hours
        })

    for r in schedule_rows:
        keys=r.keys()
        # Flexible date column names.
        raw_date=None
        for col in ("date","work_date","schedule_date","datum"):
            if col in keys and r[col]:
                raw_date=r[col]; break
        if not raw_date: continue
        try:
            work_date=datetime.strptime(str(raw_date)[:10],"%Y-%m-%d").date()
        except Exception:
            continue
        if work_date>f_to or work_date<f_from:
            continue

        relation=""
        for col in ("line","line_name","relation","route"):
            if col in keys and r[col]:
                relation=str(r[col]).strip(); break

        # Determine duration in minutes/hours from line data or row.
        minutes=0
        for col in ("duration_minutes","duration_min","minutes"):
            if col in keys and r[col] not in (None,""):
                try: minutes=float(r[col]); break
                except Exception: pass
        if not minutes and relation in line_map:
            ln=line_map[relation]
            lkeys=ln.keys()
            # Linija sprema sate i minute odvojeno: npr. 12 h + 0 min.
            # duration_minutes sam po sebi nije ukupno trajanje.
            if "duration_hours_int" in lkeys or "duration_minutes" in lkeys:
                try:
                    minutes=float(ln["duration_hours_int"] or 0)*60 + float(ln["duration_minutes"] or 0)
                except Exception:
                    minutes=0
            if not minutes:
                for col in ("duration","trajanje","duration_hours"):
                    if col in lkeys and ln[col]:
                        val=str(ln[col])
                        m=re.match(r'^\s*(\d+)\s*:\s*(\d+)',val)
                        if m: minutes=int(m.group(1))*60+int(m.group(2))
                        else:
                            try: minutes=float(val)*60
                            except Exception: pass
        if not minutes:
            for col in ("duration","trajanje","hours"):
                if col in keys and r[col] not in (None,""):
                    try: minutes=float(r[col])*60; break
                    except Exception: pass

        hours=minutes/60.0

        # Occupancy: departure day plus any additional days implied by a return record;
        # otherwise duration determines occupied calendar days.
        end_dt=work_date + timedelta(days=max(0, int((minutes-1)//1440)))
        occupied=set()
        d=work_date
        while d<=end_dt:
            if f_from<=d<=f_to: occupied.add(d.isoformat())
            d+=timedelta(days=1)

        # Driver columns vary between versions.
        assigned=[]
        for col in ("driver","driver1","driver2","vozac","vozac1","vozac2"):
            if col in keys and r[col]:
                n=str(r[col]).strip()
                if n and n not in assigned: assigned.append(n)

        # Some versions store a comma-separated driver list.
        if not assigned:
            for col in ("drivers","vozaci"):
                if col in keys and r[col]:
                    for n in str(r[col]).replace("; ",",").split(","):
                        n=n.strip()
                        if n and n not in assigned: assigned.append(n)

        # Divide driven hours equally among assigned drivers.
        per_driver=hours/len(assigned) if assigned else 0
        for name in assigned:
            add_driver(name,occupied,per_driver,relation,work_date)

    summary=[]
    for x in stats.values():
        summary.append({
            "driver":x["driver"],
            "rides":x["rides"],
            "days_count":len(x["days"]),
            "hours":x["hours"],
            "hours_text": hours_hm(x["hours"])
        })
    summary.sort(key=lambda x:(-x["rides"],x["driver"].lower()))

    top_rides=max(summary,key=lambda x:x["rides"]) if summary else None
    top_days=max(summary,key=lambda x:x["days_count"]) if summary else None
    top_hours=max(summary,key=lambda x:x["hours"]) if summary else None

    detail=[]
    if driver_filter:
        detail=details.get(driver_filter,[])

    return render_template(
        "driver_statistics.html",
        summary=summary, detail=detail,
        top_rides=top_rides, top_days=top_days, top_hours=top_hours,
        drivers=drivers, date_from=date_from, date_to=date_to,
        driver_filter=driver_filter
    )



@app.route("/statistika/autobusi")
@permission_required("schedule_view")
def vehicle_statistics():
    from datetime import datetime, timedelta, date
    date_from=request.args.get("date_from","").strip() or date.today().replace(day=1).isoformat()
    date_to=request.args.get("date_to","").strip() or date.today().isoformat()
    vehicle_filter=request.args.get("vehicle","").strip()

    c=db()
    try:
        vehicles=c.execute("SELECT registration FROM vehicles WHERE TRIM(COALESCE(registration,''))<>'' ORDER BY registration").fetchall()
    except Exception:
        vehicles=[]
    try:
        schedule_rows=c.execute("SELECT * FROM schedules ORDER BY date,id").fetchall()
    except Exception:
        schedule_rows=[]
    try: lines=c.execute("SELECT * FROM lines").fetchall()
    except Exception: lines=[]
    c.close()

    line_map={}
    for ln in lines:
        try:
            name=str(ln["name"] or "").strip()
            if name: line_map[name]=ln
        except Exception: pass

    try:
        f_from=datetime.strptime(date_from,"%Y-%m-%d").date()
        f_to=datetime.strptime(date_to,"%Y-%m-%d").date()
    except Exception:
        f_from=date.today().replace(day=1); f_to=date.today()

    stats={}; details={}
    for r in schedule_rows:
        keys=r.keys()
        raw_date=next((r[col] for col in ("date","work_date","schedule_date","datum") if col in keys and r[col]),None)
        if not raw_date: continue
        try: work_date=datetime.strptime(str(raw_date)[:10],"%Y-%m-%d").date()
        except Exception: continue
        if work_date<f_from or work_date>f_to: continue

        vehicle=next((str(r[col] or "").strip() for col in ("vehicle","bus","registration") if col in keys and r[col]),"")
        if not vehicle or vehicle.lower() in ("none","-","n/a"): continue
        if vehicle_filter and vehicle!=vehicle_filter: continue

        relation=next((str(r[col] or "").strip() for col in ("line","line_name","relation","route") if col in keys and r[col]),"")
        ln=line_map.get(relation)
        minutes=0.0
        km=0.0
        days_count=1

        # Sati i kilometraža prvenstveno se povlače s odabrane linije.
        if ln:
            lkeys=ln.keys()
            if "duration_hours_int" in lkeys or "duration_minutes" in lkeys:
                try: minutes=float(ln["duration_hours_int"] or 0)*60+float(ln["duration_minutes"] or 0)
                except Exception: pass
            if not minutes:
                for col in ("duration_hours","duration"):
                    if col in lkeys and ln[col] not in (None,""):
                        try: minutes=float(ln[col])*60; break
                        except Exception: pass
            for col in ("distance_km","kilometers","km_total","distance"):
                if col in lkeys and ln[col] not in (None,""):
                    try: km=float(ln[col] or 0); break
                    except Exception: pass
            for col in ("duration_days","days"):
                if col in lkeys and ln[col] not in (None,""):
                    try: days_count=max(1,int(float(ln[col]))); break
                    except Exception: pass

        if not minutes:
            for col in ("duration_total_minutes","duration_minutes","duration_min","minutes"):
                if col in keys and r[col] not in (None,""):
                    try: minutes=float(r[col]); break
                    except Exception: pass
        if not km:
            for col in ("distance_km","km_total","kilometers","distance"):
                if col in keys and r[col] not in (None,""):
                    try: km=float(r[col]); break
                    except Exception: pass

        occupied=set()
        for offset in range(max(1,days_count)):
            d=work_date+timedelta(days=offset)
            if f_from<=d<=f_to: occupied.add(d.isoformat())

        x=stats.setdefault(vehicle,{"vehicle":vehicle,"rides":0,"lines":set(),"days":set(),"hours":0.0,"km":0.0})
        x["rides"]+=1; x["lines"].add(relation); x["days"].update(occupied); x["hours"]+=minutes/60.0; x["km"]+=km
        details.setdefault(vehicle,[]).append({"date":work_date.isoformat(),"line":relation,"days":len(occupied),"hours":minutes/60.0,"km":km})

    summary=[]
    for x in stats.values():
        summary.append({"vehicle":x["vehicle"],"rides":x["rides"],"different_lines":len([v for v in x["lines"] if v]),
                        "days_count":len(x["days"]),"hours":x["hours"],"hours_text":f"{int(round(x['hours']*60))//60} h {int(round(x['hours']*60))%60} min","km":x["km"]})
    summary.sort(key=lambda x:(-x["days_count"],x["vehicle"].lower()))
    top_days=max(summary,key=lambda x:x["days_count"]) if summary else None
    top_hours=max(summary,key=lambda x:x["hours"]) if summary else None
    top_km=max(summary,key=lambda x:x["km"]) if summary else None
    return render_template("statistics_vehicle.html",summary=summary,vehicles=vehicles,
                           date_from=date_from,date_to=date_to,vehicle_filter=vehicle_filter,
                           top_days=top_days,top_hours=top_hours,top_km=top_km)


@app.route("/statistika/autobusi/detaljno/<path:registration>")
@permission_required("schedule_view")
def vehicle_statistics_detail(registration):
    from datetime import datetime, timedelta, date
    date_from=request.args.get("date_from","").strip() or date.today().replace(day=1).isoformat()
    date_to=request.args.get("date_to","").strip() or date.today().isoformat()
    c=db()
    try: rows=c.execute("SELECT * FROM schedules ORDER BY date,id").fetchall()
    except Exception: rows=[]
    try: lines=c.execute("SELECT * FROM lines").fetchall()
    except Exception: lines=[]
    c.close()
    line_map={str(x["name"] or "").strip():x for x in lines if str(x["name"] or "").strip()}
    try: f_from=datetime.strptime(date_from,"%Y-%m-%d").date(); f_to=datetime.strptime(date_to,"%Y-%m-%d").date()
    except Exception: f_from=date.today().replace(day=1); f_to=date.today()
    detail=[]; total_minutes=0; total_km=0; occupied=set()
    for r in rows:
        if str(r["date"] or "")[:10] < date_from or str(r["date"] or "")[:10] > date_to: continue
        vehicle=str(r["vehicle"] or "").strip()
        if vehicle != registration: continue
        wd=datetime.strptime(str(r["date"])[:10],"%Y-%m-%d").date()
        relation=str(r["line"] or "").strip(); ln=line_map.get(relation)
        mins=0; km=0; days=1
        if ln:
            try: mins=float(ln["duration_hours_int"] or 0)*60+float(ln["duration_minutes"] or 0)
            except Exception: mins=0
            if not mins:
                try: mins=float(ln["duration_hours"] or 0)*60
                except Exception: pass
            try: km=float(ln["distance_km"] or 0)
            except Exception: pass
            try: days=max(1,int(ln["duration_days"] or 1))
            except Exception: pass
        occ=[]
        for off in range(days):
            dd=wd+timedelta(days=off)
            if f_from<=dd<=f_to: occupied.add(dd.isoformat()); occ.append(dd.isoformat())
        total_minutes+=mins; total_km+=km
        drivers=[str(r[k] or "").strip() for k in ("driver1","driver2") if k in r.keys() and str(r[k] or "").strip()]
        detail.append({"date":wd.isoformat(),"line":relation,"time":r["time"] or "","drivers":", ".join(drivers) or "—","days":len(occ),"minutes":mins,"hours_text":hours_hm(mins/60),"km":km})
    return render_template("statistics_vehicle_detail.html",registration=registration,detail=detail,date_from=date_from,date_to=date_to,total_hours=hours_hm(total_minutes/60),total_km=total_km,total_days=len(occupied))


@app.route("/statistika/izvoz")
def driver_statistics_export():
    u=auth_user()
    if not u or not has_permission(u,"schedule_view"):
        return render_template("403.html"),403
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    date_from=request.args.get("date_from","").strip() or date.today().replace(day=1).isoformat()
    date_to=request.args.get("date_to","").strip() or date.today().isoformat() or date.today().isoformat()
    driver_filter=request.args.get("driver","").strip()
    c=db()
    rows=c.execute("""SELECT s.date,s.line,s.driver1,s.driver2,s.time,s.vehicle,
                             COALESCE(NULLIF(l.duration_hours_int*60 + l.duration_minutes,0),
                           ROUND(COALESCE(l.duration_hours,0)*60)) AS duration_total_minutes
                      FROM schedules s LEFT JOIN lines l ON TRIM(l.name)=TRIM(s.line)
                      WHERE s.date>=? AND s.date<=?
                      ORDER BY date,line,time""",(date_from,date_to)).fetchall()
    c.close()
    stats={}; detail=[]
    for r in rows:
        assigned=[]
        for slot in ("driver1","driver2"):
            name=(r[slot] or "").strip()
            if name and name not in assigned: assigned.append(name)
        share=float(r["duration_total_minutes"] or 0)/len(assigned) if assigned else 0
        for name in assigned:
            if driver_filter and name!=driver_filter: continue
            x=stats.setdefault(name,{"trips":0,"lines":set(),"days":set(),"minutes":0.0})
            x["trips"]+=1; x["minutes"]+=share; x["lines"].add((r["line"] or "").strip()); x["days"].add(r["date"])
            detail.append([r["date"],name,r["line"] or "",r["time"] or "",r["vehicle"] or "",f"{int(round(float(r['duration_total_minutes'] or 0)))//60} h {int(round(float(r['duration_total_minutes'] or 0)))%60} min",
                           f"{int(round(share))//60} h {int(round(share))%60} min"])
    wb=Workbook(); ws=wb.active; ws.title="Sažetak"
    ws.append(["Vozač","Broj vožnji","Različitih linija","Broj dana","Ukupno sati"])
    for cell in ws[1]:
        cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="2F75B5")
    for name,x in sorted(stats.items(),key=lambda z:(-z[1]["trips"],z[0].lower())):
        ws.append([name,x["trips"],len(x["lines"]),len(x["days"]),f"{int(round(x.get('minutes',0)))//60} h {int(round(x.get('minutes',0)))%60} min"])
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    for col,w in zip("ABCDE",[28,18,20,16,16]): ws.column_dimensions[col].width=w
    wd=wb.create_sheet("Detaljno")
    wd.append(["Datum","Vozač","Linija","Vrijeme","Vozilo","Trajanje linije (h)","Sati vozača"])
    for cell in wd[1]:
        cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="17365D")
    for r in detail: wd.append(r)
    wd.freeze_panes="A2"; wd.auto_filter.ref=wd.dimensions
    for col,w in zip("ABCDEFG",[16,28,35,12,18,20,16]): wd.column_dimensions[col].width=w
    _sig_row = (total_row + 5) if "total_row" in locals() else (ws.max_row + 5)
    ws.cell(_sig_row,1,"Preuzeo:")
    ws.cell(_sig_row,3,f"Izdao: {_stored_document_issuer(r)}")
    ws.merge_cells(start_row=_sig_row+2,start_column=1,end_row=_sig_row+2,end_column=2)
    ws.merge_cells(start_row=_sig_row+2,start_column=3,end_row=_sig_row+2,end_column=4)
    ws.cell(_sig_row+2,1,"______________________________")
    ws.cell(_sig_row+2,3,"______________________________")
    ws.cell(_sig_row,3).alignment=Alignment(horizontal="right")
    ws.cell(_sig_row+2,3).alignment=Alignment(horizontal="right")
    # Uključi potpise u područje ispisa
    try:
        ws.print_area=f"A1:{get_column_letter(ws.max_column)}{_sig_row+2}"
    except Exception:
        pass
    bio=BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio,as_attachment=True,
                     download_name=f"Statistika_vozaca_{date_from}_{date_to}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@permission_required("schedule_edit")
@app.route("/raspored/obrisi-liniju", methods=["POST"])
def clear_schedule_line():
    schedule_id=request.form.get("id", type=int)
    line=(request.form.get("line") or "").strip()
    selected_date=request.form.get("date") or date.today().isoformat()
    return_date=request.form.get("return_date") or selected_date
    c=db()
    if schedule_id:
        c.execute("UPDATE schedules SET driver1='', driver2='', vehicle='' WHERE id=?", (schedule_id,))
    elif line:
        # For the 'popuni' screen, clear only this exact saved line if it exists.
        c.execute("UPDATE schedules SET driver1='', driver2='', vehicle='' WHERE date=? AND TRIM(line)=?", (selected_date,line))
    c.commit(); c.close()
    flash("Podaci su obrisani samo s ove linije. Ostale linije nisu promijenjene.","success")
    return redirect(url_for("schedule", **{"from": return_date}))

@permission_required("schedule_edit")
@app.route("/raspored/<int:id>/obrisi-popunu", methods=["POST"])
def clear_schedule_row(id):
    c=db()
    row=c.execute("SELECT * FROM schedules WHERE id=?", (id,)).fetchone()
    if not row:
        c.close()
        abort(404)
    return_date=request.form.get("return_date") or row["date"] or date.today().isoformat()
    c.execute("UPDATE schedules SET driver1='', driver2='', vehicle='' WHERE id=?", (id,))
    c.commit()
    c.close()
    flash("Podaci su obrisani samo s ove linije. Linija je ostala u rasporedu kao 'Nije popunjeno'.","success")
    return redirect(url_for("schedule", **{"from": return_date}))

@permission_required("schedule_edit")
@app.route("/raspored/obrisi-sve", methods=["POST"])
def clear_schedule():
    from datetime import datetime, timedelta
    selected=request.form.get("date") or request.args.get("date") or date.today().isoformat()
    try:
        d0=datetime.strptime(selected,"%Y-%m-%d").date()
    except ValueError:
        d0=date.today()
    d1=d0+timedelta(days=1)
    c=db()
    # Brišu se sve konkretno spremljene vožnje za dva dana koja čine trenutni prikaz rasporeda.
    # Linije se ne brišu iz baze, pa će nakon toga opet biti prikazane kao "Nije popunjeno".
    c.execute("DELETE FROM schedules WHERE date IN (?,?)", (d0.isoformat(), d1.isoformat()))
    c.commit()
    c.close()
    flash("Raspored je očišćen. Linije su ostale u pregledu kao 'Nije popunjeno'.","success")
    return redirect(url_for("schedule", **{"from": d0.isoformat()}))

@permission_required("schedule_view")
@app.route("/raspored")
def schedule():
 from datetime import datetime
 d=request.args.get("from") or date.today().isoformat()
 try: selected=datetime.strptime(d,"%Y-%m-%d").date()
 except ValueError: selected=date.today()
 c=db(); rows,next_day=planned_rows_for_dates(c,selected); c.close()
 return render_template("schedule.html",rows=rows,d1=selected.isoformat(),d2=next_day.isoformat(),
                        selected_date=selected.isoformat(),next_date=next_day.isoformat())

def form_data(c):
 return (c.execute("SELECT * FROM drivers WHERE active='Da' ORDER BY name").fetchall(),
         c.execute("SELECT * FROM vehicles WHERE active='Da' ORDER BY registration").fetchall(),
         c.execute("SELECT * FROM lines WHERE active='Da' AND COALESCE(internal_return,0)=0 ORDER BY name").fetchall())
def normalize_vehicle_value(value):
    if value is None:
        return ""
    parts=[]
    for x in str(value).replace("–","/").replace("—","/").split("/"):
        x=x.strip()
        if x and x.lower() not in ("none","null"):
            if x not in parts:
                parts.append(x)
    return " / ".join(parts)

def validate(c,id,d,d1,d2,vehicle):
 e=[]
 if d1 and d2 and d1==d2:
  e.append("Isti vozač je upisan dva puta u istoj vožnji.")
 return e

def driver_conflicts(c,id,d,d1,d2):
 names=[x for x in (d1,d2) if x]
 if not names:
  return []
 ph=",".join(["?"]*len(names))
 return c.execute(
  f"""SELECT * FROM schedules WHERE id<>? AND date=?
      AND (TRIM(driver1) IN ({ph}) OR TRIM(driver2) IN ({ph}))
      ORDER BY time,line""",(id,d,*names,*names)).fetchall()

def vehicle_conflicts(c,id,d,vehicle):
 if not vehicle:
  return []
 return c.execute("""SELECT * FROM schedules WHERE id<>? AND date=? AND TRIM(vehicle)=TRIM(?) ORDER BY time,line""",
                  (id,d,vehicle)).fetchall()


def driver_conflict_warning(conflicts, d1, d2):
    names={x for x in (d1,d2) if x}
    if not conflicts or not names:
        return ""
    parts=[]
    seen=set()
    for r in conflicts:
        for slot in ("driver1","driver2"):
            nm=(r[slot] or "").strip()
            if nm in names:
                key=(nm,r["line"],r["time"])
                if key not in seen:
                    seen.add(key)
                    parts.append(f"{nm} je već na liniji {r['line']} u {r['time'] or ''}")
    return "UPOZORENJE: " + "; ".join(parts) + ". Isti vozač može voziti dvije povezane linije isti dan. Ako je to namjerno, klikni „Prihvati i spremi“."

@permission_required("schedule_edit")
@app.route("/raspored/popuni",methods=["GET","POST"])
def fill_planned():
 from datetime import datetime
 d=request.args.get("date") or request.form.get("date") or date.today().isoformat()
 source_date=request.args.get("source_date") or request.form.get("source_date") or return_date if False else None
 return_date=request.args.get("return_date") or request.form.get("return_date") or d
 source_date=request.args.get("source_date") or request.form.get("source_date") or return_date
 line=request.args.get("line","").strip() or request.form.get("line","").strip()
 c=db()
 row=c.execute("SELECT * FROM lines WHERE TRIM(name)=TRIM(?) AND active='Da' LIMIT 1",(line,)).fetchone()
 drivers,vehicles,lines=form_data(c)
 internal_lines=c.execute("""
     SELECT * FROM lines
     WHERE active='Da'
       AND LOWER(COALESCE(CAST(internal_return AS TEXT),'0')) IN ('1','true','da','yes')
     ORDER BY name
    """).fetchall()
 if not row:
  c.close()
  return "Linija nije pronađena",404

 if request.method=="POST":
  d=request.form.get("date",d).strip()
  d1=request.form.get("driver1","").strip()
  d2=request.form.get("driver2","").strip()
  v=normalize_vehicle_value(request.form.get("vehicle",""))
  note=request.form.get("note","").strip()
  # Group and time are always taken from the line definition.
  t=row["departure"] or ""
  sd=(row["group_type"] or "D0").strip()
  group="Danas" if sd=="D0" else "Sutra"
  errors=validate(c,0,d,d1,d2,v)
  conflict=driver_conflicts(c,0,d,d1,d2)
  warning=driver_conflict_warning(conflict,d1,d2)
  confirmed=(request.form.get("confirm_driver_conflict")=="1" or request.form.get("force_conflict")=="1")
  if not t:
   errors.append("Odabrana linija nema uneseno vrijeme polaska.")
  if errors:
   for e in errors: flash(e,"danger")
  elif warning and not confirmed:
   c.close()
   planned={"date":d,"line":row["name"],"time":t,"driver1":d1,"driver2":d2,
            "vehicle":v,"note":note,"group_name":row["schedule_day"],"source_date":source_date}
   return render_template("schedule_form.html",mode="add",row=planned,
                          drivers=drivers,vehicles=vehicles,lines=lines,internal_lines=internal_lines,
                          return_date=return_date,driver_conflict_warning=warning)
  else:
   c.execute("""INSERT INTO schedules
                (date,line,time,driver1,driver2,vehicle,note,group_name,source_date)
                VALUES(?,?,?,?,?,?,?,?,?)""",
             (d,row["name"],t,d1,d2,v,note,group,source_date))
   parent_id=c.execute("SELECT last_insert_rowid()").fetchone()[0]
   if request.form.get("add_return"):
    rdate=request.form.get("return_trip_date","").strip()
    rline=request.form.get("return_line","").strip()
    rd1=request.form.get("return_driver1","").strip() or d1
    rd2=request.form.get("return_driver2","").strip() or d2
    rv=normalize_vehicle_value(request.form.get("return_vehicle","")) or v
    rr=c.execute("SELECT * FROM lines WHERE TRIM(name)=TRIM(?) AND COALESCE(internal_return,0)=1",(rline,)).fetchone()
    if rr and rdate:
     c.execute("""INSERT INTO schedules(date,line,time,driver1,driver2,vehicle,note,group_name,source_date,is_return,return_of,hidden_from_schedule)
                  VALUES(?,?,?,?,?,?,?,?,?,?,?,1)""",
               (rdate,rr["name"],rr["departure"] or "",rd1,rd2,rv,"Povratak","Povratak",source_date,1,parent_id))
   c.commit()
   c.close()
   flash("Raspored je dodan.","success")
   return redirect(url_for("schedule",**{"from":return_date,"to":return_date}))
 c.close()
 planned={"date":d,"line":row["name"],"time":row["departure"],
          "driver1":"","driver2":"","vehicle":"","note":"",
           "group_name":row["schedule_day"],"source_date":source_date}
 return render_template("schedule_form.html",mode="add",row=planned,
                        drivers=drivers,vehicles=vehicles,lines=lines,
                        internal_lines=internal_lines,return_date=return_date)

@permission_required("schedule_edit")
@app.route("/raspored/dodaj",methods=["GET","POST"])
def add_schedule():
 c=db()
 return_date=request.args.get("return_date") or request.form.get("return_date") or date.today().isoformat()
 drivers,vehicles,lines=form_data(c)
 if request.method=="POST":
  d=request.form["date"]; line=request.form["line"]; d1=request.form.get("driver1","").strip(); d2=request.form.get("driver2","").strip(); v=normalize_vehicle_value(request.form.get("vehicle","")); note=request.form.get("note","").strip(); line_row=c.execute("SELECT departure,schedule_day FROM lines WHERE TRIM(name)=TRIM(?) AND active='Da' LIMIT 1",(line,)).fetchone()
  t=line_row["departure"] if line_row else ""
  day_group=(line_row["schedule_day"] or "").strip() if line_row else ""
  group="Danas" if day_group=="D0" else ("Sutra" if day_group=="D+1" else day_group)
  e=validate(c,0,d,d1,d2,v)
  conflict=driver_conflicts(c,0,d,d1,d2)
  warning=driver_conflict_warning(conflict,d1,d2)
  confirmed=(request.form.get("confirm_driver_conflict")=="1" or request.form.get("force_conflict")=="1")
  if not t: e.append("Odabrana linija nema uneseno vrijeme polaska.")
  if e:
   for x in e:flash(x,"danger")
  elif warning and not confirmed:
   c.close()
   planned={"date":d,"line":line,"time":t,"driver1":d1,"driver2":d2,"vehicle":v,
            "note":note,"group_name":group}
   return render_template("schedule_form.html",mode="add",row=planned,
                          drivers=drivers,vehicles=vehicles,lines=lines,
                          return_date=request.form.get("return_date") or d,
                          driver_conflict_warning=warning)
  else:
   c.execute("INSERT INTO schedules(date,line,time,driver1,driver2,vehicle,note,group_name,source_date) VALUES(?,?,?,?,?,?,?,?,?)",(d,line,t,d1,d2,v,note,group,d))
   c.commit();c.close();flash("Raspored je dodan.","success");return redirect(url_for("schedule",**{"from":d}))
 c.close();return render_template("schedule_form.html",mode="add",row={},drivers=drivers,vehicles=vehicles,lines=lines,return_date=return_date)

@permission_required("schedule_view", "schedule_edit")
@app.route("/raspored/<int:id>/uredi",methods=["GET","POST"])
def edit_schedule(id):
 c=db(); row=c.execute("SELECT * FROM schedules WHERE id=?",(id,)).fetchone()
 if not row:c.close();return "Nije pronađeno",404
 return_date=request.args.get("return_date") or request.form.get("return_date") or row["date"]
 drivers,vehicles,lines=form_data(c)
 if request.method=="POST":
  d=request.form["date"];line=request.form["line"];d1=request.form.get("driver1","").strip();d2=request.form.get("driver2","").strip();v=normalize_vehicle_value(request.form.get("vehicle",""));note=request.form.get("note","").strip();line_row=c.execute("SELECT departure,schedule_day FROM lines WHERE TRIM(name)=TRIM(?) AND active='Da' LIMIT 1",(line,)).fetchone()
  t=line_row["departure"] if line_row else ""
  day_group=(line_row["schedule_day"] or "").strip() if line_row else ""
  group="Danas" if day_group=="D0" else ("Sutra" if day_group=="D+1" else day_group)
  e=validate(c,id,d,d1,d2,v)
  conflict=driver_conflicts(c,id,d,d1,d2)
  warning=driver_conflict_warning(conflict,d1,d2)
  confirmed=(request.form.get("confirm_driver_conflict")=="1" or request.form.get("force_conflict")=="1")
  if not t:e.append("Odabrana linija nema uneseno vrijeme polaska.")
  if e:
   for x in e:flash(x,"danger")
  elif warning and not confirmed:
   c.close()
   planned={"date":d,"line":line,"time":t,"driver1":d1,"driver2":d2,"vehicle":v,
            "note":note,"group_name":group}
   return render_template("schedule_form.html",mode="edit",row=planned,
                          drivers=drivers,vehicles=vehicles,lines=lines,
                          return_date=return_date,driver_conflict_warning=warning)
  else:
   c.execute("UPDATE schedules SET date=?,line=?,time=?,driver1=?,driver2=?,vehicle=?,note=?,group_name=?,source_date=? WHERE id=?",(d,line,t,d1,d2,v,note,group,d,id))
   c.commit();c.close();flash("Raspored je izmijenjen.","success");return redirect(url_for("schedule",**{"from":return_date,"to":return_date}))
 c.close();return render_template("schedule_form.html",mode="edit",row=row,drivers=drivers,vehicles=vehicles,lines=lines,return_date=return_date)

@app.post("/raspored/<int:id>/obrisi")
def delete_schedule(id):
 c=db();c.execute("DELETE FROM schedules WHERE id=?",(id,));c.commit();c.close();flash("Vožnja je obrisana.","success");return redirect(url_for("schedule"))

@permission_required("schedule_view")
@app.route("/dnevni")
def daily():
 from datetime import datetime
 d=request.args.get("date") or date.today().isoformat()
 try: selected=datetime.strptime(d,"%Y-%m-%d").date()
 except ValueError: selected=date.today()
 c=db(); rows,next_day=planned_rows_for_dates(c,selected)
 phone_rows=c.execute("SELECT name,phone FROM drivers").fetchall()
 phones={str(x["name"]).strip():(x["phone"] or "") for x in phone_rows if x["name"]}
 for r in rows:
  r["p1"]=phones.get((r.get("driver1") or "").strip(),"")
  r["p2"]=phones.get((r.get("driver2") or "").strip(),"")
 c.close()
 return render_template("daily.html",rows=rows,selected_date=selected.isoformat(),
                        tomorrow_date=next_day.isoformat())

@permission_required("drivers_export")
@app.route("/vozači/izvoz")
def export_drivers():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    c=db()
    rows=c.execute("SELECT * FROM drivers ORDER BY name").fetchall()
    c.close()
    wb=Workbook(); ws=wb.active; ws.title="Vozači"; ws.sheet_view.showGridLines=False
    headers=["Ime i prezime","Mobitel","Prijava","Baza","Grad","Adresa","Status","Napomena"]
    for col,h in enumerate(headers,1):
        x=ws.cell(1,col,h); x.font=Font(name="Arial",bold=True,size=12,color="FFFFFF")
        x.fill=PatternFill("solid",fgColor="2F75B5"); x.alignment=Alignment(horizontal="center")
    thin=Side(style="thin",color="AAB7C4")
    for rr,r in enumerate(rows,2):
        vals=[r["name"],r["phone"],r["prijava"],r["baza"],r["grad"],r["adresa"],r["active"],r["note"]]
        for cc,v in enumerate(vals,1):
            x=ws.cell(rr,cc,"" if v is None or str(v).lower()=="none" else v)
            x.font=Font(name="Arial",size=11); x.border=Border(left=thin,right=thin,top=thin,bottom=thin)
            x.alignment=Alignment(vertical="center",wrap_text=True)
    widths=[25,18,14,18,20,35,14,35]
    for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width=w
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:H{max(1,ws.max_row)}"
    bio=io.BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio,as_attachment=True,download_name="vozaci.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@permission_required("vehicles_export")
@app.route("/vozila/izvoz")
def export_vehicles():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    c=db()
    rows=c.execute("SELECT * FROM vehicles ORDER BY registration").fetchall()
    c.close()
    wb=Workbook(); ws=wb.active; ws.title="Vozila"; ws.sheet_view.showGridLines=False
    headers=["Registracija","Tip vozila","Broj šasije","Osovine","Broj sjedala","Godina proizvodnje","Datum registracije","Istek registracije","Tahograf","Istek tahografa","Istek periodičnog","Istek PP aparata","Status","Napomena"]
    # Use only columns that exist in the current database.
    dbcols={r[1] for r in db().execute("PRAGMA table_info(vehicles)").fetchall()} if False else set()
    for col,h in enumerate(headers,1):
        x=ws.cell(1,col,h); x.font=Font(name="Arial",bold=True,size=12,color="FFFFFF")
        x.fill=PatternFill("solid",fgColor="2F75B5"); x.alignment=Alignment(horizontal="center")
    thin=Side(style="thin",color="AAB7C4")
    for rr,r in enumerate(rows,2):
        def val(*names):
            for n in names:
                try:
                    v=r[n]
                    if v is not None and str(v).lower()!="none": return v
                except (KeyError,IndexError): pass
            return ""
        vals=[val("registration"),val("vehicle_type","type"),val("chassis_number"),val("axles"),val("seats","seat_count"),val("production_year","year"),val("registration_date"),val("registration_expiry"),val("tachograph_type"),val("tachograph_expiry"),val("periodic_expiry"),val("fire_extinguisher_expiry"),val("active"),val("note")]
        for cc,v in enumerate(vals,1):
            x=ws.cell(rr,cc,v); x.font=Font(name="Arial",size=11); x.border=Border(left=thin,right=thin,top=thin,bottom=thin)
            x.alignment=Alignment(vertical="center",wrap_text=True)
    for i,w in enumerate([20,24,16,22,20,20,14,40],1): ws.column_dimensions[chr(64+i)].width=w
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:H{max(1,ws.max_row)}"
    bio=io.BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio,as_attachment=True,download_name="vozila.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@permission_required("drivers_view")
@app.route("/vozači")
def drivers():
 q=request.args.get("q","").strip();c=db();rows=clean_rows(c.execute("SELECT * FROM drivers WHERE name LIKE ? ORDER BY name",("%"+q+"%",)).fetchall());c.close();return render_template("drivers.html",rows=rows,q=q)
@permission_required("vehicles_view")
@app.route("/vozila")
def vehicles():
 q=request.args.get("q","").strip()
 sort=request.args.get("sort","registration")
 direction=request.args.get("direction","asc")
 allowed={
     "registration":"registration",
     "vehicle_type":"vehicle_type",
     "seats":"CAST(COALESCE(seats,0) AS INTEGER)",
     "production_year":"CAST(COALESCE(production_year,0) AS INTEGER)",
     "axles":"CAST(COALESCE(axles,0) AS INTEGER)",
     "registration_expiry":"registration_expiry"
 }
 if sort not in set(allowed) | {"registration_days_left"}: sort="registration"
 direction="desc" if direction=="desc" else "asc"
 c=db()
 if sort=="registration_days_left":
  rows=clean_rows(c.execute(
      "SELECT * FROM vehicles WHERE registration LIKE ? ORDER BY registration ASC",
      ("%"+q+"%",)
  ).fetchall())
 else:
  rows=clean_rows(c.execute(
      f"SELECT * FROM vehicles WHERE registration LIKE ? ORDER BY {allowed[sort]} {direction}, registration ASC",
      ("%"+q+"%",)
  ).fetchall())
 c.close()
 from datetime import datetime as _dt, date as _date
 today=_date.today()
 for r in rows:
  raw=(r["registration_expiry"] or "").strip()
  remaining=None
  if raw:
   try:
    remaining=(_dt.strptime(raw,"%Y-%m-%d").date()-today).days
   except ValueError:
    try:
     remaining=(_dt.strptime(raw,"%d.%m.%Y").date()-today).days
    except ValueError:
     remaining=None
  r["registration_days_left"]=remaining
 if sort=="registration_days_left":
  # Vozila bez valjanog datuma isteka ostaju na dnu.
  rows.sort(
   key=lambda r: (r["registration_days_left"] is None, r["registration_days_left"] if r["registration_days_left"] is not None else 0),
   reverse=(direction=="desc")
  )
  # Kod silaznog sortiranja None ponovno premjesti na kraj.
  if direction=="desc":
   rows.sort(key=lambda r: r["registration_days_left"] is None)
 return render_template("vehicles.html",rows=rows,q=q,sort=sort,direction=direction)
@permission_required("lines_view")
@app.route("/linije")
def lines():
 q=request.args.get("q","").strip()
 c=db()
 rows=clean_rows(c.execute("SELECT * FROM lines WHERE name LIKE ? ORDER BY name",("%"+q+"%",)).fetchall())
 for r in rows:
  r["permits"]=get_line_permits(c,r["id"])
 c.close()
 return render_template("lines.html",rows=rows,q=q)

@permission_required("drivers_edit")
@app.route("/vozači/dodaj",methods=["GET","POST"])
def add_driver():
 c=db()
 if request.method=="POST":
  name=request.form.get("name","").strip()
  phone=request.form.get("phone","").strip()
  prijava=request.form.get("prijava","").strip()
  baza=request.form.get("baza","").strip()
  grad=request.form.get("grad","").strip()
  adresa=request.form.get("adresa","").strip()
  active=request.form.get("active","Da")
  note=request.form.get("note","").strip()
  if not name:
   flash("Ime i prezime je obavezno.","danger")
  else:
   c.execute("""INSERT INTO drivers(name,phone,prijava,baza,grad,adresa,active,note)
                VALUES(?,?,?,?,?,?,?,?)""",(name,phone,prijava,baza,grad,adresa,active,note))
   c.commit();c.close();flash("Vozač je dodan.","success");return redirect(url_for("drivers"))
 c.close()
 return render_template("driver_form.html",mode="add",row={},
                        prijave=["GTM","CBZ","VANJSKI"],
                        baze=["Međugorje","Kiseljak","Zagreb"])

@permission_required("drivers_view", "drivers_edit")
@app.route("/vozači/<int:id>/uredi",methods=["GET","POST"])
def edit_driver(id):
 c=db();row=c.execute("SELECT * FROM drivers WHERE id=?",(id,)).fetchone()
 if not row:
  c.close();return "Nije pronađeno",404
 if request.method=="POST":
  name=request.form.get("name","").strip()
  phone=request.form.get("phone","").strip()
  prijava=request.form.get("prijava","").strip()
  baza=request.form.get("baza","").strip()
  grad=request.form.get("grad","").strip()
  adresa=request.form.get("adresa","").strip()
  active=request.form.get("active","Da")
  note=request.form.get("note","").strip()
  if not name:
   flash("Ime i prezime je obavezno.","danger")
  else:
   c.execute("""UPDATE drivers SET name=?,phone=?,prijava=?,baza=?,grad=?,adresa=?,active=?,note=?
                WHERE id=?""",(name,phone,prijava,baza,grad,adresa,active,note,id))
   c.commit();c.close();flash("Vozač je izmijenjen.","success");return redirect(url_for("drivers"))
 c.close()
 return render_template("driver_form.html",mode="edit",row=row,
                        prijave=["GTM","CBZ","VANJSKI"],
                        baze=["Međugorje","Kiseljak","Zagreb"],
                        readonly=not has_permission(auth_user(),"drivers_edit"))

@permission_required("vehicles_edit")
@app.route("/vozila/dodaj",methods=["GET","POST"])
def add_vehicle():
 c=db()
 if request.method=="POST":
  reg=request.form.get("registration","").strip()
  vehicle_type=request.form.get("vehicle_type","").strip()
  seats=request.form.get("seats","").strip() or None
  production_year=request.form.get("production_year","").strip() or None
  registration_date=request.form.get("registration_date","").strip() or None
  chassis_number=request.form.get("chassis_number","").strip() or None
  axles=request.form.get("axles","").strip() or None
  tachograph_type=request.form.get("tachograph_type","").strip() or None
  tachograph_expiry=request.form.get("tachograph_expiry","").strip() or None
  periodic_expiry=request.form.get("periodic_expiry","").strip() or None
  fire_extinguisher_expiry=request.form.get("fire_extinguisher_expiry","").strip() or None
  active=request.form.get("active","Da")
  note=request.form.get("note","").strip()
  permit_file_name=None
  upload=request.files.get("permit_file")
  if upload and upload.filename:
   if not permit_file_allowed(upload.filename):
    flash("Dozvoljeni formati su PDF, JPG, JPEG, PNG i WEBP.","danger")
   else:
    ext=upload.filename.rsplit(".",1)[1].lower()
    permit_file_name=f"{uuid.uuid4().hex}.{ext}"
    upload.save(os.path.join(UPLOAD_DIR,permit_file_name))
  registration_expiry=""
  if registration_date:
   try:
    from datetime import datetime,timedelta
    registration_expiry=(datetime.strptime(registration_date,"%Y-%m-%d").date()+timedelta(days=365)).isoformat()
   except ValueError:
    flash("Datum registracije nije ispravan.","danger")
    registration_date=None
  if not reg: flash("Registracija vozila je obavezna.","danger")
  elif c.execute("SELECT 1 FROM vehicles WHERE UPPER(TRIM(registration))=UPPER(TRIM(?))",(reg,)).fetchone(): flash("Vozilo već postoji.","danger")
  else:
   c.execute("""INSERT INTO vehicles(registration,vehicle_type,seats,production_year,registration_date,registration_expiry,chassis_number,axles,tachograph_type,tachograph_expiry,periodic_expiry,fire_extinguisher_expiry,active,note,permit_file_name)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",(reg,vehicle_type,seats,production_year,registration_date,registration_expiry,chassis_number,axles,tachograph_type,tachograph_expiry,periodic_expiry,fire_extinguisher_expiry,active,note,permit_file_name))
   c.commit();c.close();flash("Vozilo je dodano.","success");return redirect(url_for("vehicles"))
 c.close();return render_template("vehicle_form.html",mode="add",row={})

@permission_required("vehicles_view", "vehicles_edit")
@app.route("/vozila/<int:id>/uredi",methods=["GET","POST"])
def edit_vehicle(id):
 c=db();row=c.execute("SELECT * FROM vehicles WHERE id=?",(id,)).fetchone()
 if not row:c.close();return "Nije pronađeno",404
 if request.method=="POST":
  reg=request.form.get("registration","").strip()
  vehicle_type=request.form.get("vehicle_type","").strip()
  seats=request.form.get("seats","").strip() or None
  production_year=request.form.get("production_year","").strip() or None
  registration_date=request.form.get("registration_date","").strip() or None
  chassis_number=request.form.get("chassis_number","").strip() or None
  axles=request.form.get("axles","").strip() or None
  tachograph_type=request.form.get("tachograph_type","").strip() or None
  tachograph_expiry=request.form.get("tachograph_expiry","").strip() or None
  periodic_expiry=request.form.get("periodic_expiry","").strip() or None
  fire_extinguisher_expiry=request.form.get("fire_extinguisher_expiry","").strip() or None
  active=request.form.get("active","Da")
  note=request.form.get("note","").strip()
  permit_file_name=row["permit_file_name"] if "permit_file_name" in row.keys() else None
  upload=request.files.get("permit_file")
  if upload and upload.filename:
   if not permit_file_allowed(upload.filename):
    flash("Dozvoljeni formati su PDF, JPG, JPEG, PNG i WEBP.","danger")
   else:
    if permit_file_name: delete_permit_file(permit_file_name)
    ext=upload.filename.rsplit(".",1)[1].lower()
    permit_file_name=f"{uuid.uuid4().hex}.{ext}"
    upload.save(os.path.join(UPLOAD_DIR,permit_file_name))
  registration_expiry=""
  if registration_date:
   try:
    from datetime import datetime,timedelta
    registration_expiry=(datetime.strptime(registration_date,"%Y-%m-%d").date()+timedelta(days=365)).isoformat()
   except ValueError:
    flash("Datum registracije nije ispravan.","danger")
    registration_date=None
  if not reg: flash("Registracija vozila je obavezna.","danger")
  else:
   dup=c.execute("SELECT id FROM vehicles WHERE UPPER(TRIM(registration))=UPPER(TRIM(?)) AND id<>?",(reg,id)).fetchone()
   if dup: flash("Vozilo već postoji.","danger")
   else:
    c.execute("""UPDATE vehicles SET registration=?,vehicle_type=?,seats=?,production_year=?,
                 registration_date=?,registration_expiry=?,chassis_number=?,axles=?,tachograph_type=?,tachograph_expiry=?,periodic_expiry=?,fire_extinguisher_expiry=?,active=?,note=?,permit_file_name=? WHERE id=?""",
              (reg,vehicle_type,seats,production_year,registration_date,registration_expiry,chassis_number,axles,tachograph_type,tachograph_expiry,periodic_expiry,fire_extinguisher_expiry,active,note,permit_file_name,id))
    c.commit();c.close();flash("Vozilo je izmijenjeno.","success");return redirect(url_for("vehicles"))
 c.close();return render_template("vehicle_form.html",mode="edit",row=clean_row(row),readonly=not has_permission(auth_user(),"vehicles_edit"))

@permission_required("drivers_delete")
@app.route("/vozači/<int:id>/izbrisi",methods=["POST"])
def delete_driver(id):
 c=db()
 row=c.execute("SELECT name FROM drivers WHERE id=?",(id,)).fetchone()
 if not row:
  c.close(); flash("Vozač nije pronađen.","danger"); return redirect(url_for("drivers"))
 name=row["name"]
 c.execute("DELETE FROM drivers WHERE id=?",(id,))
 c.commit(); c.close()
 flash(f"Vozač {name} je izbrisan.","success")
 return redirect(url_for("drivers"))

@permission_required("vehicles_view")
@app.route("/vozila/<int:id>/dozvola/pregled")
def view_vehicle_permit(id):
 c=db()
 row=c.execute("SELECT registration,permit_file_name FROM vehicles WHERE id=?",(id,)).fetchone()
 c.close()
 if not row or not row["permit_file_name"]: abort(404)
 filename=os.path.basename(row["permit_file_name"])
 if not os.path.isfile(os.path.join(UPLOAD_DIR,filename)): abort(404)
 return send_from_directory(UPLOAD_DIR,filename)

@permission_required("vehicles_edit")
@app.route("/vozila/<int:id>/dozvola/obrisi")
def delete_vehicle_permit(id):
 c=db()
 row=c.execute("SELECT permit_file_name FROM vehicles WHERE id=?",(id,)).fetchone()
 if not row:
  c.close(); return "Nije pronađeno",404
 delete_permit_file(row["permit_file_name"])
 c.execute("UPDATE vehicles SET permit_file_name=NULL WHERE id=?",(id,))
 c.commit(); c.close()
 flash("Skenirana dozvola vozila je obrisana.","success")
 return redirect(url_for("edit_vehicle",id=id))

@permission_required("vehicles_delete")
@app.route("/vozila/<int:id>/izbrisi",methods=["POST"])
def delete_vehicle(id):
 c=db()
 row=c.execute("SELECT registration FROM vehicles WHERE id=?",(id,)).fetchone()
 if not row:
  c.close(); flash("Vozilo nije pronađeno.","danger"); return redirect(url_for("vehicles"))
 reg=row["registration"]
 c.execute("DELETE FROM vehicles WHERE id=?",(id,))
 c.commit(); c.close()
 flash(f"Vozilo {reg} je izbrisano.","success")
 return redirect(url_for("vehicles"))

def ensure_line_permits_table(c):
    c.execute("""CREATE TABLE IF NOT EXISTS line_permits (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        line_id INTEGER NOT NULL,
        permit TEXT NOT NULL,
        valid_until TEXT,
        file_name TEXT
    )""")
    cols=[r[1] for r in c.execute("PRAGMA table_info(line_permits)").fetchall()]
    if "valid_until" not in cols:
        c.execute("ALTER TABLE line_permits ADD COLUMN valid_until TEXT")
    if "file_name" not in cols:
        c.execute("ALTER TABLE line_permits ADD COLUMN file_name TEXT")

def get_line_permits(c, line_id):
    try:
        ensure_line_permits_table(c)
        return [dict(r) for r in c.execute(
            "SELECT id,permit,valid_until,file_name FROM line_permits WHERE line_id=? ORDER BY id",(line_id,)
        ).fetchall()]
    except Exception:
        return []

def save_line_permits(c, line_id, permits):
    ensure_line_permits_table(c)
    existing={r["id"]:r["file_name"] for r in c.execute(
        "SELECT id,file_name FROM line_permits WHERE line_id=?",(line_id,)
    ).fetchall()}
    c.execute("DELETE FROM line_permits WHERE line_id=?",(line_id,))
    for item in permits:
        permit=str(item.get("permit") or "").strip()
        valid_until=str(item.get("valid_until") or "").strip() or None
        file_name=item.get("file_name") or None
        if permit and permit.lower() not in ("none","null"):
            c.execute("INSERT INTO line_permits(line_id,permit,valid_until,file_name) VALUES(?,?,?,?)",
                      (line_id,permit,valid_until,file_name))

def delete_permit_file(file_name):
    if file_name:
        path=os.path.join(UPLOAD_DIR,os.path.basename(file_name))
        try:
            if os.path.isfile(path): os.remove(path)
        except OSError:
            pass

@permission_required("lines_view")
@app.route("/dozvole/<int:permit_id>/pregled")
def view_permit(permit_id):
    c=db()
    row=c.execute("SELECT file_name,permit FROM line_permits WHERE id=?",(permit_id,)).fetchone()
    c.close()
    if not row or not row["file_name"]: abort(404)
    path=os.path.join(UPLOAD_DIR,os.path.basename(row["file_name"]))
    if not os.path.isfile(path): abort(404)
    return send_from_directory(UPLOAD_DIR,os.path.basename(row["file_name"]))

@permission_required("lines_view")
@app.route("/dozvole/<int:permit_id>/download")
def download_permit(permit_id):
    c=db()
    row=c.execute("SELECT file_name,permit FROM line_permits WHERE id=?",(permit_id,)).fetchone()
    c.close()
    if not row or not row["file_name"]: abort(404)
    path=os.path.join(UPLOAD_DIR,os.path.basename(row["file_name"]))
    if not os.path.isfile(path): abort(404)
    ext=os.path.splitext(row["file_name"])[1]
    safe_name="Dozvola_"+re.sub(r"[^A-Za-z0-9_-]+","_",row["permit"] or "dokument")+ext
    return send_from_directory(UPLOAD_DIR,os.path.basename(row["file_name"]),as_attachment=True,download_name=safe_name)

@permission_required("lines_edit")
@app.route("/dozvole/<int:permit_id>/obrisi",methods=["GET","POST"])
def delete_permit_document(permit_id):
    c=db()
    row=c.execute("SELECT file_name,line_id FROM line_permits WHERE id=?",(permit_id,)).fetchone()
    if not row:
        c.close(); return "Nije pronađeno",404
    delete_permit_file(row["file_name"])
    c.execute("UPDATE line_permits SET file_name=NULL WHERE id=?",(permit_id,))
    c.commit(); c.close()
    flash("Skenirana dozvola je obrisana.","success")
    return redirect(url_for("edit_line",id=row["line_id"]))

@permission_required("lines_edit")
@app.route("/linije/dodaj",methods=["GET","POST"])
def add_line():
 c=db()
 if request.method=="POST":
  name=request.form.get("name","").strip();origin=request.form.get("origin","").strip();destination=request.form.get("destination","").strip()
  departure=request.form.get("departure","").strip();active=request.form.get("active","Da")
  group_type=request.form.get("group_type","D0").strip()
  internal_return=1 if request.form.get("internal_return") else 0
  duration_days=max(1,int(request.form.get("duration_days","1") or 1))
  duration_hours_int=max(0,int(request.form.get("duration_hours_int","0") or 0))
  duration_minutes=max(0,min(59,int(request.form.get("duration_minutes","0") or 0)))
  duration_hours=duration_hours_int + duration_minutes/60
  try:
   distance_km=max(0,float((request.form.get("distance_km","0") or "0").replace(",", ".")))
  except ValueError:
   distance_km=0
  days=[x for x,_ in WEEKDAYS if request.form.get("day_"+x)]
  # New lines default to every day; unchecked boxes are the exceptions.
  if not days:
   days=[x for x,_ in WEEKDAYS]
  schedule_day=",".join(days)
  permit_values=request.form.getlist("permit")
  permit_dates=request.form.getlist("permit_valid_until")
  permit_files=request.files.getlist("permit_file")
  permit_existing=request.form.getlist("permit_existing_file")
  permits=[]
  for i,pv in enumerate(permit_values):
      pv=pv.strip()
      if not pv:
          continue
      file_name=permit_existing[i] if i < len(permit_existing) and permit_existing[i] else None
      if i < len(permit_files) and permit_files[i] and permit_files[i].filename:
          f=permit_files[i]
          if not permit_file_allowed(f.filename):
              flash("Dozvoljeni formati su PDF, JPG, JPEG, PNG i WEBP.","danger")
              continue
          ext=f.filename.rsplit(".",1)[1].lower()
          file_name=f"{uuid.uuid4().hex}.{ext}"
          f.save(os.path.join(UPLOAD_DIR,file_name))
      permits.append({"permit":pv,"valid_until":permit_dates[i].strip() if i < len(permit_dates) else "","file_name":file_name})
  if not name or not departure: flash("Naziv linije i vrijeme polaska su obavezni.","danger")
  elif group_type not in ("D0","D+1"): flash("Grupa mora biti D0 ili D+1.","danger")
  elif not days: flash("Odaberi barem jedan dan prometovanja.","danger")
  elif c.execute("SELECT 1 FROM lines WHERE LOWER(TRIM(name))=LOWER(TRIM(?))",(name,)).fetchone(): flash("Linija već postoji.","danger")
  else:
   cur=c.execute("""INSERT INTO lines(name,origin,destination,departure,schedule_day,active,group_type,internal_return,duration_days,duration_hours,duration_hours_int,duration_minutes,distance_km)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",(name,origin,destination,departure,schedule_day,active,group_type,internal_return,duration_days,duration_hours,duration_hours_int,duration_minutes,distance_km))
   save_line_permits(c,cur.lastrowid,permits)
   c.commit();c.close();flash("Linija je dodana.","success");return redirect(url_for("lines"))
 c.close();return render_template("line_form.html",mode="add",row={},weekdays=WEEKDAYS,permits=[])

@permission_required("lines_view", "lines_edit")
@app.route("/linije/<int:id>/uredi",methods=["GET","POST"])
def edit_line(id):
 c=db();row=c.execute("SELECT * FROM lines WHERE id=?",(id,)).fetchone()
 if not row:c.close();return "Nije pronađeno",404
 if request.method=="POST":
  name=request.form.get("name","").strip();origin=request.form.get("origin","").strip();destination=request.form.get("destination","").strip()
  departure=request.form.get("departure","").strip();active=request.form.get("active","Da")
  group_type=request.form.get("group_type","D0").strip()
  internal_return=1 if request.form.get("internal_return") else 0
  duration_days=max(1,int(request.form.get("duration_days","1") or 1))
  duration_hours_int=max(0,int(request.form.get("duration_hours_int","0") or 0))
  duration_minutes=max(0,min(59,int(request.form.get("duration_minutes","0") or 0)))
  duration_hours=duration_hours_int + duration_minutes/60
  try:
   distance_km=max(0,float((request.form.get("distance_km","0") or "0").replace(",", ".")))
  except ValueError:
   distance_km=0
  days=[x for x,_ in WEEKDAYS if request.form.get("day_"+x)]
  schedule_day=",".join(days)
  permit_values=request.form.getlist("permit")
  permit_dates=request.form.getlist("permit_valid_until")
  permit_files=request.files.getlist("permit_file")
  permit_existing=request.form.getlist("permit_existing_file")
  permits=[]
  for i,pv in enumerate(permit_values):
      pv=pv.strip()
      if not pv:
          continue
      file_name=permit_existing[i] if i < len(permit_existing) and permit_existing[i] else None
      if i < len(permit_files) and permit_files[i] and permit_files[i].filename:
          f=permit_files[i]
          if not permit_file_allowed(f.filename):
              flash("Dozvoljeni formati su PDF, JPG, JPEG, PNG i WEBP.","danger")
              continue
          ext=f.filename.rsplit(".",1)[1].lower()
          file_name=f"{uuid.uuid4().hex}.{ext}"
          f.save(os.path.join(UPLOAD_DIR,file_name))
      permits.append({"permit":pv,"valid_until":permit_dates[i].strip() if i < len(permit_dates) else "","file_name":file_name})
  if not name or not departure: flash("Naziv linije i vrijeme polaska su obavezni.","danger")
  elif group_type not in ("D0","D+1"): flash("Grupa mora biti D0 ili D+1.","danger")
  elif not days: flash("Odaberi barem jedan dan prometovanja.","danger")
  else:
   old_name=row["name"]
   c.execute("""UPDATE lines SET name=?,origin=?,destination=?,departure=?,schedule_day=?,active=?,group_type=?,internal_return=?,duration_days=?,duration_hours=?,duration_hours_int=?,duration_minutes=?,distance_km=? WHERE id=?""",
             (name,origin,destination,departure,schedule_day,active,group_type,internal_return,duration_days,duration_hours,duration_hours_int,duration_minutes,distance_km,id))
   c.execute("UPDATE schedules SET time=?,line=? WHERE TRIM(line)=TRIM(?)",(departure,name,old_name))
   save_line_permits(c,id,permits)
   c.commit();c.close();flash("Linija je izmijenjena.","success");return redirect(url_for("lines"))
 permits=get_line_permits(c,id)
 c.close()
 return render_template("line_form.html",mode="edit",row=row,weekdays=WEEKDAYS,permits=permits)

@permission_required("schedule_export")
@app.route("/izvoz/excel")
def export_excel():
    from datetime import datetime, timedelta
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.page import PageMargins

    d1s=request.args.get("from") or date.today().isoformat()
    # Daily schedule export is an explicit snapshot of the selected date.
    # The UI may pass a "to" date for display purposes; snapshot=1 means
    # ignore it and export selected date's D0 + D+1 only.
    snapshot=request.args.get("snapshot")=="1"
    d2s=request.args.get("to") or d1s
    try:
        d1=datetime.strptime(d1s,"%Y-%m-%d").date()
        # If the caller asks for the same from/to day, this is always a
        # daily snapshot. A daily snapshot contains that day's D0 + D+1,
        # never a separately-built schedule for the next calendar day.
        d2=d1 if snapshot or d2s==d1s else datetime.strptime(d2s,"%Y-%m-%d").date()
    except ValueError:
        d1=date.today(); d2=d1

    if d1 == d2:
        dates=[d1,d1+timedelta(days=1)]
        title_date=f"{d1.day:02d}./{(d1+timedelta(days=1)).day:02d}."
        months={"JANUARY":"SIJEČANJ","FEBRUARY":"VELJAČA","MARCH":"OŽUJAK","APRIL":"TRAVANJ",
                "MAY":"SVIBANJ","JUNE":"LIPANJ","JULY":"SRPANJ","AUGUST":"KOLOVOZ",
                "SEPTEMBER":"RUJAN","OCTOBER":"LISTOPAD","NOVEMBER":"STUDENI","DECEMBER":"PROSINAC"}
        month=months.get(d1.strftime("%B").upper(),d1.strftime("%B").upper())
        subtitle=f"RASPORED ZA DATUM {title_date} {month} {d1.year}"
    else:
        dates=[]
        cur=d1
        while cur<=d2:
            dates.append(cur); cur+=timedelta(days=1)
        subtitle=f"RASPORED ZA DATUM {d1.day:02d}./{d2.day:02d}. {d1.strftime('%B').upper()} {d1.year}"

    c=db()
    rows=[]
    if d1 == d2:
        # Build the exact same list shown on /raspored for the selected day.
        # Do not query the next day's saved snapshot separately.
        planned, _ = planned_rows_for_dates(c,d1)
        seen=set()
        for r in planned:
            # Match the screen: only entered rows are exported.
            if not ((r.get("driver1") or "").strip() or
                    (r.get("driver2") or "").strip() or
                    (r.get("vehicle") or "").strip()):
                continue
            norm_line=" ".join((r.get("line") or "").replace("–","-").replace("—","-").split()).lower()
            key=(r.get("date"),norm_line,(r.get("time") or "").strip())
            if key in seen:
                continue
            seen.add(key)
            rows.append(r)
    else:
        # Exact screen snapshot for every selected date.
        for cur in (d1 + timedelta(days=i) for i in range((d2-d1).days+1)):
            planned, _ = planned_rows_for_dates(c,cur)
            seen=set()
            for r in planned:
                if not ((r.get("driver1") or "").strip() or
                        (r.get("driver2") or "").strip() or
                        (r.get("vehicle") or "").strip()):
                    continue
                norm_line=" ".join((r.get("line") or "").replace("–","-").replace("—","-").split()).lower()
                key=(r.get("date"),norm_line,(r.get("time") or "").strip())
                if key in seen:
                    continue
                seen.add(key)
                rows.append(r)

    phone_rows=c.execute("SELECT name,phone FROM drivers").fetchall()
    phones={str(r["name"]).strip():(r["phone"] or "") for r in phone_rows if r["name"]}
    c.close()

    wb=Workbook()
    ws=wb.active
    ws.title="Dnevni raspored"
    ws.sheet_view.showGridLines=False

    navy="17365D"; blue="2F75B5"; light_blue="D9EAF7"; very_light="F5F9FD"
    gold="FFF2CC"; white="FFFFFF"; border_color="AAB7C4"; dark="1F2937"

    ws.merge_cells("A1:H1")
    ws["A1"]="GLOBTOUR MEĐUGORJE   •   DNEVNI RASPORED VOŽNJI   •   CROATIA BUS"
    ws["A1"].font=Font(name="Arial",bold=True,size=16,color=white)
    ws["A1"].fill=PatternFill(fill_type="solid",fgColor=navy)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=32

    ws.merge_cells("A2:H2")
    ws["A2"]=subtitle
    ws["A2"].font=Font(name="Arial",bold=True,size=13,color=navy)
    ws["A2"].fill=PatternFill(fill_type="solid",fgColor=light_blue)
    ws["A2"].alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[2].height=27

    headers=["Datum","Linija","Vrijeme","Vozilo","Vozač 1","Vozač 2","Mobitel Vozač 1","Mobitel Vozač 2"]
    for col,h in enumerate(headers,1):
        cell=ws.cell(3,col,h)
        cell.font=Font(name="Arial",bold=True,size=11,color=white)
        cell.fill=PatternFill(fill_type="solid",fgColor=blue)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.row_dimensions[3].height=34

    thin=Side(style="thin",color=border_color)
    medium=Side(style="medium",color=navy)
    border=Border(left=thin,right=thin,top=thin,bottom=thin)

    def excel_text(value):
        if value is None:
            return ""
        text=str(value).strip()
        return "" if text.lower()=="none" else text

    for rnum,r in enumerate(rows,4):
        # "date" is the schedule-building date. The actual operating date is
        # D0 = selected date, D+1 = selected date + one day.
        actual_date=d1 + timedelta(days=1) if (r["group_name"] or "").strip()=="Sutra" else d1
        disp=actual_date.strftime("%d.%m.%Y.")

        d1_name=excel_text(r["driver1"])
        d2_name=excel_text(r["driver2"])
        vehicle_name=excel_text(r["vehicle"])
        vals=[disp,excel_text(r["line"]),excel_text(r["time"]),vehicle_name,
              d1_name,d2_name,
              excel_text(phones.get(d1_name,"")) if d1_name else "",
              "" if not d2_name else excel_text(phones.get(d2_name,""))]

        for col,val in enumerate(vals,1):
            cell=ws.cell(rnum,col,val)
            # All operational data is larger and important fields are bold.
            cell.font=Font(name="Arial",size=12,bold=(col in (1,2,4,5,6,7,8)),color=dark)
            cell.border=border
            cell.alignment=Alignment(
                horizontal="center" if col in (1,3) else "left",
                vertical="center",
                wrap_text=(col==4)
            )
            if rnum % 2 == 1:
                cell.fill=PatternFill(fill_type="solid",fgColor=very_light)
            if col==1:
                cell.fill=PatternFill(fill_type="solid",fgColor=gold)
            if col==2:
                cell.font=Font(name="Arial",size=12,bold=True,color=navy)
        vehicle_text=excel_text(r["vehicle"])
        ws.row_dimensions[rnum].height=38 if len(vehicle_text)>35 else 25

    # Širina kolone Vozilo se automatski prilagođava kada je uneseno
    # više vozila, npr. "A31 O 873 / A49 K 508".
    max_vehicle_len=max([len(excel_text(r["vehicle"])) for r in rows] or [0])
    # Excel width is based on the longest actual vehicle value.
    # Keep enough room for two registrations such as:
    # "A31 O 873 / A49 K 508"
    vehicle_width=max(22,min(55,max_vehicle_len+5))
    widths=[15,34,11,vehicle_width,27,27,22,22]
    for i,wid in enumerate(widths,1):
        ws.column_dimensions[chr(64+i)].width=wid

    ws.freeze_panes="A4"
    # Namjerno BEZ auto_filtera – izvoz treba izgledati kao čisti raspored.
    ws.auto_filter.ref=None
    ws.page_setup.orientation="landscape"
    ws.page_setup.paperSize=ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth=1
    ws.page_setup.fitToHeight=0
    ws.sheet_properties.pageSetUpPr.fitToPage=True
    ws.page_margins=PageMargins(left=0.25,right=0.25,top=0.35,bottom=0.35,header=0.15,footer=0.15)
    ws.print_title_rows="1:3"
    ws.print_area=f"A1:H{max(3,3+len(rows))}"
    ws.oddFooter.center.text="Stranica &P / &N"

    bio=io.BytesIO()
    wb.save(bio); bio.seek(0)
    return send_file(bio,as_attachment=True,
                     download_name=f"Dnevni_raspored_{d1.strftime('%d-%m-%Y')}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/raspored")
def api(): 
 d=request.args.get("date") or date.today().isoformat();c=db();r=[dict(x) for x in c.execute("SELECT * FROM schedules WHERE date=? ORDER BY time,line",(d,)).fetchall()];c.close();return jsonify(r)

@app.route("/klijenti/<int:id>/obrisi", methods=["POST"])
def customer_delete(id):
    c = db()
    customer = c.execute("SELECT * FROM customers WHERE id=?", (id,)).fetchone()
    if not customer:
        c.close()
        flash("Klijent nije pronađen.", "error")
        return redirect(url_for("customers"))
    try:
        c.execute("DELETE FROM customers WHERE id=?", (id,))
        c.commit()
        flash("Klijent je uspješno obrisan.", "success")
    except Exception as e:
        c.rollback()
        flash("Klijent se ne može obrisati.", "error")
    finally:
        c.close()
    return redirect(url_for("customers"))

if __name__=="__main__":app.run(host="127.0.0.1",port=5000,debug=False)
